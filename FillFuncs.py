from re import A
import pandas as pd
import numpy as np
import openpyxl as opxl
import os.path
import pycountry
import pycountry_convert as pc
import datetime as dt
import json
import os.path
import mysql.connector

from Spider import task1
from Spider_Pubchem import runMerged
from difflib import SequenceMatcher

def tidyDescription(desc):
    desc = desc.replace('&TRADE', '')
    desc = desc.replace('Â®', '')
    desc = desc.replace('Î¼', '')
    desc = desc.replace('& Growâ„¢', '')
    desc = desc.replace('â‰', '')
    desc = desc.replace('&Beta', '')
    desc = desc.replace('Â‰¥', '')
    desc = desc.replace('Â', '')
    desc = desc.replace('<em>', '')
    desc = desc.replace('</em>', '')
    desc = desc.replace('®','')
    desc = desc.replace('<F128>','')
    desc = desc.replace('<130>','')
    desc = desc.replace('<F255>','')
    desc = desc.replace('&deg;', '°')
    desc = desc.replace('deg ', '°')
    if desc.endswith(','):
        desc = desc[:len(desc)-1]
    return desc

def columnize(df, primary_row, secondary_row):
    temp = df.fillna('')
    columns = []
    for i in range(len(temp.columns)):
        if temp.iloc[primary_row, i] == '':
            columns.append(temp.iloc[secondary_row, i])
        else:
            columns.append(temp.iloc[primary_row, i])
    df.columns = columns
    return df

def chooseSearchName(sku, magento):
    product_info = magento.loc[magento['sku'] == sku]

    search_name = ''

    if not product_info.empty:
        cas_number = product_info['cas_number'].values[0]

        if type(cas_number) == str and len(cas_number) > 0:
            search_name = cas_number
        else:
            pack_size_joined = product_info['pack_size_joined'].values[0]
            name = product_info['name'].values[0]
            if type(name) == str:
                if type(pack_size_joined) == str:
                    name = name.replace(pack_size_joined, '')
                search_name = name

    return search_name

def getAbcamData(sku, magento):
    search_name = chooseSearchName(sku, magento)
    product_info = magento.loc[magento['sku'] == sku]
    if os.path.exists('Abcam/'+str(search_name)+'.json'):
        with open('Abcam/'+str(search_name)+'.json', 'r') as f:
            data = json.load(f)
        
        return chooseDataAbcam(data, product_info)
    elif len(search_name) > 0:
        search_name = search_name.replace('/', '')
        task1('Abcam', [search_name])
        # print(os.path.exists('Abcam/'+str(search_name)+'.json'))
        if os.path.exists('Abcam/'+str(search_name)+'.json'):
            with open('Abcam/'+str(search_name)+'.json', 'r') as f:
                data = json.load(f)
            
            return chooseDataAbcam(data, product_info)
        else:
            return None

    else:
        return None

def getPubchemData(sku, magento):
    search_name = chooseSearchName(sku, magento)
    if len(search_name) > 0:
        runMerged([search_name])

        try:
            with open('Pubchem/result.json', 'r') as f:
                data = json.load(f)

            for item in data:
                if item['search_name'] == search_name:
                    return item
        except:
            print('pubchem error...')

    return None

# def getDatabase(host, user, password, database):
#     mydb = mysql.connector.connect(
#         host=host,
#         user=user,
#         password=password,
#         database=database
#     )

#     return mydb

# def getDatabaseData(mydb, sku, table_name):
#     cursor = mydb.cursor()
#     cursor.execute("SELECT * FROM " + table_name + " WHERE sku = '" + sku + "';")
#     result = cursor.fetchall()

#     if len(result) > 0:
#         return result[0]
#     else:
#         return None

# def isColInDB(mydb, table_name, val_name):
#     cursor = mydb.cursor()
#     cursor.execute("SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = '" + table_name + "';")
#     columns = cursor.fetchall()

#     val_name = str(val_name).replace(' ', '_')
#     val_name = str(val_name).replace('/', '_')
#     for i in columns:
#         if i[0] == val_name:
#             return True

#     return False

# def SKUINDB(mydb, sku, table_name):
#     cursor = mydb.cursor()
#     cursor.execute("SELECT sku FROM " + table_name + " WHERE sku = '" + sku + "';")
#     result = cursor.fetchall()
#     if not len(result) > 0:
#         cursor.execute("INSERT INTO " + table_name + "(sku) VALUES ('" + sku + "');")
#     mydb.commit()

# def getValueFromResult(mydb, result, val_name, table_name):
#     cursor = mydb.cursor()
#     cursor.execute("SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = '" + table_name + "';")
#     columns = cursor.fetchall()
#     index = 0
#     val_name = str(val_name).replace(' ', '_')
#     val_name = str(val_name).replace('/', '_')
#     for i in columns:
#         if i[0] == val_name:
#             return result[index]
#         index += 1
    
#     addColToDB(mydb, val_name, table_name)
#     return None

# def addColToDB(mydb, val_name, table_name):
#     val_name = str(val_name).replace(' ', '_')
#     val_name = str(val_name).replace('/', '_')
#     cursor = mydb.cursor()
#     try:
#         cursor.execute("ALTER TABLE " + table_name + " ADD " + val_name + " VARCHAR(300);")
#         cursor.execute("ALTER TABLE " + table_name + " ALTER " + val_name + " SET DEFAULT 'None';")
#     except:
#         print("Could not add col")
#     mydb.commit()

# def addValToDB(mydb, sku, val_name, table_name, val):
    # cursor = mydb.cursor()
    # SKUINDB(mydb, sku, table_name)
    # val_name = str(val_name).replace(' ', '_')
    # val_name = str(val_name).replace('/', '_')
    # cursor.execute("UPDATE " + table_name + " SET " + val_name + " = '" + str(val) + "' WHERE sku = '" + sku + "';")
    # mydb.commit()

def chooseDataAbcam(data, product_info):
    if not product_info.empty:
        for item in data:
            if item['search_name'] == product_info['cas_number'].values[0]:
                return item
            else:
                antibody_type = product_info['antibody_type'].values[0]
                antibody_type = str(antibody_type)
                if 'Polyclonal' in antibody_type or 'polycolnal' in antibody_type:
                    antibody_type = 'Polyclonal'
                elif 'Monoclonal' in antibody_type or 'monoclonal' in antibody_type:
                    antibody_type = 'Monoclonal'
                else:
                    antibody_type = ''

                similarity = SequenceMatcher(None, item['search_name'], item['product_name']).ratio()

                if item['Clonality'] == antibody_type and similarity > 0.6:
                    return item

    return None

def catSheetVWR(sheet_name, cell_str):
    if cell_str == '':
        cell_str = sheet_name
    else:
        cell_str += ', ' + sheet_name

    return cell_str

def VWREnrichmentDriver(filename, magento, prms, magento_may, categories):
    wb = opxl.load_workbook(filename)
    skus = wb.active
    # antibodies = pd.read_excel('forms/vwr_enrichment_antibodies.xlsx', dtype=str)
    antibodies = pd.read_excel('forms/GlobalProductEnrichmentFile_Antibodies (New).xlsx', dtype=str)
    # ppe = pd.read_excel('forms/vwr_enrichment_ppe.xlsx', dtype=str)
    ppe = pd.read_excel('forms/GlobalProductEnrichmentFile_Proteins_Peptides_Enzymes (New).xlsx', dtype=str)
    # sera = pd.read_excel('forms/vwr_enrichment_sera.xlsx', dtype=str)
    sera = pd.read_excel('forms/GlobalProductEnrichmentFile_Sera (Old).xlsx', dtype=str)
    ccm = pd.read_excel('forms/vwr_enrichment_ccm.xlsx', dtype=str)
    # chemicals = pd.read_excel('forms/vwr_enrichment_chemicals.xlsx', dtype=str)
    chemicals = pd.read_excel('forms/GlobalProductEnrichmentFile_Chemicals (New).xlsx', dtype=str)
    # antibodies.columns = np.arange(len(antibodies.columns))
    # ppe.columns = np.arange(len(ppe.columns))
    # sera.columns = np.arange(len(sera.columns))
    # ccm.columns = np.arange(len(ccm.columns))
    # chemicals.columns = np.arange(len(chemicals.columns))
    ccm = columnize(ccm, 7, 6)
    sera = columnize(sera, 7, 6)
    chemicals.columns = chemicals.iloc[11]
    # antibodies.columns = antibodies.iloc[1]
    antibodies = columnize(antibodies, 6, 5)
    # ppe.columns = ppe.iloc[0]
    ppe = columnize(ppe, 6, 5)
    new_columns = [i.strip() if type(i) == str else i for i in ccm.columns]
    ccm.columns = new_columns
    new_columns = [i.strip() if type(i) == str else i for i in sera.columns]
    sera.columns = new_columns
    new_columns = [i.strip() if type(i) == str else i for i in chemicals.columns]
    chemicals.columns = new_columns
    new_columns = [i.strip() if type(i) == str else i for i in antibodies.columns]
    antibodies.columns = new_columns
    new_columns = [i.strip() if type(i) == str else i for i in ppe.columns]
    ppe.columns = new_columns

    categories.columns = categories.iloc[0]
    
    num_antibodies = 0
    num_ppe = 0
    num_sera = 0
    num_ccm = 0
    num_chemicals = 0
    
    for i in range(2, skus.max_row+1):
        sku = str(skus['A'+str(i)].value)
        product_info = magento.loc[magento['sku'] == sku]
        category_info = categories.loc[categories['SKU'] == sku]

        if not category_info.empty:
            category = category_info['Class Name'].values[0]
        else:
            category = ''

        if not product_info.empty:
            name = product_info['name'].values[0]
            cas_number = product_info['cas_number'].values[0]
        else:
            name = ''
            cas_number = ''

        skus['B'+str(i)].value = ''
        if sku.startswith('11'):
            if (category == 'MEDIA' or sku.startswith('1130') or sku.startswith('1131') or sku.startswith('1133') or sku.startswith('1140') or sku.startswith('1141') or sku.startswith('1144') or sku.startswith('1145') or sku.startswith('1148') or sku.startswith('1151')):
                skus['B'+str(i)].value = catSheetVWR('Cell Culture Media', skus['B'+str(i)].value)
                ccm.loc[num_ccm+11, 'Supplier Part No.'] = sku
                num_ccm = num_ccm + 1
            if category == 'BIOCHEMICALS':
                skus['B'+str(i)].value = catSheetVWR('Chemicals', skus['B'+str(i)].value)
                chemicals.loc[num_chemicals+22, 'Supplier Part No.'] = sku
                num_chemicals = num_chemicals + 1
            # if category == 'PCR' or sku.startswith('11EB') or sku.startswith('11EP') or sku.startswith('11MSTP') or sku.startswith('11RTO') or sku.startswith('1199'):

            
        else:
            if (sku.startswith('08') or 'anti-' in name or 'Anti-' in name or 'antibody' in name or 'Antibody' in name) and not sku.startswith('07'):
                skus['B'+str(i)].value = catSheetVWR('Antibodies', skus['B'+str(i)].value)
                antibodies.loc[num_antibodies+14, 'Supplier Cat. No.'] = sku
                num_antibodies = num_antibodies + 1
            elif 'serum' in name or 'Serum' in name:
                    skus['B'+str(i)].value = catSheetVWR('Sera', skus['B'+str(i)].value)
                    sera.loc[num_sera+11, 'Supplier Part No.'] = sku
                    num_sera = num_sera + 1
            if sku.startswith('02') or 'ase' in name:
                if not (sku.startswith('02') and ('ChLiA' in name or 'peptide' in name or 'Peptide' in name)):
                    skus['B'+str(i)].value = catSheetVWR('Protein, Peptides, Enzymes', skus['B'+str(i)].value)
                    ppe.loc[num_ppe+14, 'Supplier Cat. No.'] = sku
                    num_ppe = num_ppe + 1
            if 'media' in name or 'Media' in name or 'medium' in name or 'Medium' in name or 'RPMI' in name or (sku.startswith('09') and not 'serum' in name):
                skus['B'+str(i)].value = catSheetVWR('Cell Culture Media', skus['B'+str(i)].value)
                ccm.loc[num_ccm+11, 'Supplier Part No.'] = sku
                num_ccm = num_ccm + 1
            if (type(cas_number) == str and len(cas_number) > 0 and not cas_number == 'Not applicable') or sku.startswith('02'):
                skus['B'+str(i)].value = catSheetVWR('Chemicals', skus['B'+str(i)].value)
                chemicals.loc[num_chemicals+22, 'Supplier Part No.'] = sku
                num_chemicals = num_chemicals + 1

    wb.save(filename)

    fillVWR_Enrichment_Antibodies(antibodies, magento, prms)
    fillVWR_Enrichmnent_PPE(ppe, magento, prms)
    fillVWR_Enrichmnent_Sera(sera, magento)
    # fillVWR_Enrichmnent_CCM(ccm, magento, prms)
    fillVWR_Enrichmnent_Chemicals(chemicals, magento, prms)
    fillVWR_Enrichment(filename, magento_may)

def fillVWR_Old(filename, magento, new_magento, lot_master, prms, unspsc_codes, origin):
    wb = opxl.load_workbook(filename)
    skus = wb.active
    vwr = pd.read_excel('forms/vwr_form.xlsx', dtype=object)
    vwr.columns = vwr.iloc[1]
    for i in range(2, skus.max_row+1):
        vwr.loc[i+2, ' Vendor Part Number'] = str(skus['A'+str(i)].value)
        # vwr[' Vendor Part Number'][i+2] = str(skus['A'+str(i)].value)
        # print(str(skus['A'+str(i)].value))

    for i in range(4, len(vwr)):
        sku = str(vwr[' Vendor Part Number'][i])
        product_info = magento.loc[magento['sku'] == sku]
        lot_info = lot_master.loc[lot_master['Product number'] == sku]
        prms_info = prms.loc[prms['SKU'] == sku]
        unspsc_info = unspsc_codes.loc[unspsc_codes['Part Number'] == sku]
        origin_info = origin.loc[origin['Product number'] == sku]
        product_info_july = new_magento.loc[new_magento['sku'] == sku]

        if not product_info_july.empty:
            tariff_code = product_info_july['tariff_code'].values[0]
        else:
            tariff_code = ''
        if not product_info.empty:
            if not prms_info.empty:
                name = prms_info['Product Name'].values[0]
                price = prms_info['USD List Price'].values[0]
                un_num = prms_info['UN#'].values[0]
                packing_group = prms_info['Packing Group'].values[0]
                alt_storage_temp = prms_info['Storage Temp'].values[0]
                ship_temp = prms_info['Ship Temp'].values[0]
                storage_temp = prms_info['Storage Temp'].values[0]
            else:
                name = product_info['name'].values[0].upper()
                price = product_info['price'].values[0]
                packing_group = product_info['packing_group'].values[0]
                ship_temp = ''
                storage_temp = ''
                    
            if not origin_info.empty:
                country_of_origin = origin_info.loc[origin_info['Expiration date -'] == max(origin_info['Expiration date -'].values)]['Country of Origin'].values[0]
            elif not prms_info.empty:
                country_of_origin = prms_info['Country of Origin (most recent lot)'].values[0]
            elif not lot_info.empty:
                country_of_origin = lot_info['Country of Origin'].values[0]
            else:
                country_of_origin = ''
            
            if country_of_origin == 'GER':
                        country_of_origin = 'DEU'
                
            if type(country_of_origin) == str and len(country_of_origin) > 0:
                country_of_origin = pc.country_alpha3_to_country_alpha2(country_of_origin)
                        
            if not unspsc_info.empty:
                unspsc = unspsc_info['UNSPSC'].values[0]
            else:
                unspsc = product_info['unspsc'].values[0]
            
            short_desc = product_info['short_description'].values[0]
            img_link = product_info['base_image'].values[0]
            pack_size = product_info['pack_size_numeric_value'].values[0]
            unit = product_info['pack_size_unit_of_measure'].values[0]
            quantity = product_info['lk_packaging_facet'].values[0]
            host = product_info['host'].values[0]
            # tariff_code = product_info['tariff_code'].values[0]
            msds_avail = skus['B'+str(i-2)].value
            cas_number = product_info['cas_number'].values[0]
            DOT_PSN = 'N/A'
            hazard_class = product_info['hazard_class'].values[0]
            biochem_physiol_actions = product_info['biochem_physiol_actions'].values[0]
            storage_and_handling = product_info['storage_and_handling'].values[0]
            group_name = product_info['prms_group_name'].values[0]
            # storage_temp = product_info['storage_and_handling'].values[0]
            categories = product_info['categories'].values[0]
            weight_in_lb = new_magento.loc[new_magento['sku'] == sku]['weight'].values[0]
            
            
            if type(quantity) == str:
                quantity = quantity.replace('1 ÂµCi', '')
                quantity = quantity.split('x')[0]
                
            
            if weight_in_lb < 0.5:
                vwr[' UM1 Length'][i] = 7
                vwr[' UM1 Width'][i] = 4
                vwr[' UM1 Height'][i] = 5
            elif weight_in_lb <= 1:
                vwr[' UM1 Length'][i] = 12
                vwr[' UM1 Width'][i] = 7
                vwr[' UM1 Height'][i] = 5
            else:
                vwr[' UM1 Length'][i] = 12
                vwr[' UM1 Width'][i] = 12
                vwr[' UM1 Height'][i] = 12

            if type(name) == str:
                name = tidyDescription(name)
                temp_name = ''
                for c in name:
                    if c.isalpha() or c.isnumeric() or c == ' ':
                        temp_name = temp_name + c
                name = temp_name
                if len(name) <= 40:
                    vwr[' 40 Character Description'][i] = name
                else:
                    vwr[' 40 Character Description'][i] = name[:40]
            if type(short_desc) == str:
                short_desc = tidyDescription(short_desc)
                if len(short_desc) <= 300:
                    vwr[' Enhanced Description'][i] = short_desc + "\n pls work"
                else:
                    vwr[' Enhanced Description'][i] = short_desc[:300]
                    
            if ship_temp == 'CP':
                vwr[' Ship Temp Requirement'][i] = 'B'
                vwr[' If on ice, blue or dry?'][i] = 'Blue'
            elif ship_temp == 'DI':
                vwr[' Ship Temp Requirement'][i] = 'A'
                vwr[' If on ice, blue or dry?'][i] = 'Dry'
            else:
                vwr[' Ship Temp Requirement'][i] = 'C'
                
            if type(host) == str and len(host) > 0:
                vwr[' Animal/Plant or Synthetic'][i] = 'Animal: ' + host
            else:
                vwr[' Animal/Plant or Synthetic'][i] = 'Synthetic'
                
            #if DOT_PSN != 'N/A': 
            vwr[' DOT - If NOS, Technical Name'][i] = name
            vwr[' DOT - UN Identification #'][i] = biochem_physiol_actions
            vwr[' DOT - Packing Group I, II, or III'][i] = packing_group
            vwr[' Reportable Quanitity (RQ)'][i] = 'No'
            vwr[' Limited Quantity'][i] = 'No'
            vwr[' DOT-SP'][i] = 'No'
            vwr[' IATA - Packing Group I, II, or III'][i] = packing_group
                
                
            if type(price) != str and type(categories) == str:
                if 'Biochemicals' in categories or 'Cell Biology' in categories or 'Immunology' in categories or 'Antibody' in categories or 'Chemicals' in categories:
                    vwr[' Vendor Price'][i] = 0.77*price
                    vwr[' Future Vendor Price'][i] = 0.77*price
                elif 'Molecular Biology' in categories or 'SafTest' in categories:
                    vwr[' Vendor Price'][i] = 0.85*price
                    vwr[' Future Vendor Price'][i] = 0.85*price
                elif sku.startswith('02') or sku.startswith('07') or sku.startswith('04'):
                    vwr[' Vendor Price'][i] = 0.77*price
                    vwr[' Future Vendor Price'][i] = 0.77*price
                elif sku.startswith('09') or sku.startswith('08') or sku.startswith('11'):
                    vwr[' Vendor Price'][i] = 0.85*price
                    vwr[' Future Vendor Price'][i] = 0.85*price
                
            tariff_code = str(tariff_code)
            if len(tariff_code) >= 4:
                vwr[' US Harmonization Code'][i] = tariff_code
            
            vwr[' Purchase Unit of Measure'][i] = 'EA'
            vwr[' Minimum Order?'][i] = 1
            vwr[' Vendor Currency'][i] = 'USD'
            vwr[' Future Pricing Effectivity Date'][i] = '1/1/2021'
            vwr[' lot controlled'][i] = 'No'
            vwr[' Component Quantity'][i] = quantity
            vwr[' Component Measure'][i] = unit
            vwr[' Component Size'][i] = pack_size
            vwr[' Selling UM1'][i] = 'EA'
            vwr[' UM1 List Price'][i] = price
            vwr[' Future UM1 List Price'][i] = price
            vwr[' UM1 Weight'][i] = weight_in_lb
            vwr[' UM1 Ship as Is?'][i] = 'Y'
            vwr[' Country of Origin'][i] = country_of_origin
            vwr[' Eligible for NAFTA'][i] = 'No'
            vwr[' Other US or CA Free Trade Agreements (FTAs)'][i] = 'No'
            vwr[' If Animal Origin - Confirm Genus/Species & CITES compliance'][i] = host
            vwr[' MSDS available'][i] = msds_avail
            vwr[' WHMIS Compliant'][i] = msds_avail
            vwr[' DOT Proper Ship Name'][i] = DOT_PSN
            vwr[' Type of Outside Container Used'][i] = 'Box'
            vwr[' Inner Bottles Pressure Tested?'][i] = 'No'
            vwr[' Inner Bottle Material'][i] = 'Plastic'
            vwr[' CAS#'][i] = cas_number
            vwr[' If a Chemical, CWC?'][i] = 'No'
            vwr[' Regulated by DEA or TTB?'][i] = 'No'
            vwr[' Regulated by Health Canada as a Precursor Chemical?'][i] = 'No'
            vwr[' Certificate of Analysis'][i] = 'Y'
            vwr[' Certificate of Sterility'][i] = 'No'
            vwr[' Certificate of Quality/Conformance'][i] = 'No'
            vwr[' Electrical Certification Code'][i] = 'N/A'
            vwr[' Electrical w/Motors'][i] = 'N/A'
            vwr[' Drop Ship Fee?'][i] = 'No'
            vwr[' Air Eligible?'][i] = 'Y'
            
            if not lot_info.empty:
                creation_date = np.datetime64(lot_info['Creation date -'].values[0])
                expiration_date = np.datetime64(lot_info['Expiration date -'].values[0])
                shelf_life = expiration_date - creation_date
                shelf_life = shelf_life.astype('timedelta64[M]')/np.timedelta64(1, 'M')
                if shelf_life > 0:
                    vwr[' Dated Shelf Life?'][i] = shelf_life
            
            vwr[' If dated, returnable?'][i] = 'No'
            
            if type(storage_temp) == str:
                if storage_temp == 'AM':
                    vwr[' Storage Temp Requirement'][i] = 'C'
                elif storage_temp == 'FR':
                    vwr[' Storage Temp Requirement'][i] = 'A'
                # elif '-70' in storage_temp or '-80' in storage_temp:
                #     vwr[111][i] = 'D'
                elif storage_temp == 'RF':
                    vwr[' Storage Temp Requirement'][i] = 'B'
                elif storage_temp == '70' or storage_temp == '80':
                    vwr[' Storage Temp Requirement'][i] = 'D'
                # else:
                #     vwr[111][i] = storage_temp
            
            vwr[" Children's Product?"][i] = 'N'


    new_vwr = opxl.load_workbook('forms/vwr_form.xlsx')
    vwr_sheet = new_vwr.active
    if vwr_sheet.max_row < skus.max_row+4:
        for j in range(skus.max_row - vwr_sheet.max_row + 4):
            vwr_sheet.insert_rows(vwr_sheet.max_row-1)
    i = 0
    for row in vwr_sheet.iter_rows(min_row=5):
        if i < skus.max_row:
            for j in range(4, len(vwr.columns)-1):
                # print(i, j)
                row[j].value = vwr.iloc[i+3, j]
        else:
            break
        i = i+1

    # new_vwr.save('../../outputs/old_product_outputs/old_vwr_output.xlsx')
    new_vwr.save('outputs/old_product_outputs/old_vwr_output.xlsx')

    return vwr

def fillThomas_Old(filename, magento, new_magento, lot_master, prms, unspsc_codes, origin, magento_sept):
    thomas = pd.read_excel('forms/thomas_form.xlsx', dtype = object)
    thomas = columnize(thomas, 4, 3)
    new_columns = [i.strip() if type(i) == str else i for i in thomas.columns]
    thomas.columns = new_columns
    wb = opxl.load_workbook(filename)
    skus = wb.active
    for i in range(2, skus.max_row+1):
        thomas.loc[i+19, 'MFR. NUMBER'] = str(skus['A'+str(i)].value)
        
    for i in range(21, skus.max_row+20):
        sku = thomas['MFR. NUMBER'][i]
        product_info = magento.loc[magento['sku'] == sku]
        lot_info = lot_master.loc[lot_master['Product number'] == sku]
        origin_info = origin.loc[origin['Product number'] == sku]
        prms_info = prms.loc[prms['SKU'] == sku]
        unspsc_info = unspsc_codes.loc[unspsc_codes['Part Number'] == sku]
        product_info_july = new_magento.loc[new_magento['sku'] == sku]
        product_info_sept = magento_sept.loc[magento_sept['sku'] == sku]

        if not product_info_july.empty:
            tariff_code = product_info_july['tariff_code'].values[0]
        else:
            tariff_code = ''

        if not product_info.empty:
            if not product_info_sept.empty:
                pack_size_joined = product_info_sept['pack_size_joined'].values[0]
                product_type = product_info_july['product_type'].values[0]
                description = product_info_sept['description'].values[0]
                clone_name = product_info_sept['clone_name'].values[0]
                conjugate = product_info_sept['conjugate'].values[0]
                host = product_info_sept['host'].values[0]
                isotype = product_info_sept['isotype'].values[0]
                solubility = product_info_sept['solubility'].values[0]
                source = product_info_sept['source'].values[0]
                species_reactivity = product_info_sept['species_reactivity'].values[0]
                immunogen = product_info_sept['immunogen'].values[0]
                concentration = product_info_sept['concentration'].values[0]
                specificity = product_info_sept['specificity'].values[0]
                sterility = product_info_sept['sterility'].values[0]
                components = product_info_sept['components'].values[0]
                format_value = product_info_sept['format'].values[0]
                name = product_info_sept['name'].values[0]

            else:
                pack_size_joined = product_info['pack_size_joined'].values[0]
                product_type = product_info['product_type'].values[0]
                description = product_info['description'].values[0]
                clone_name = ''
                conjugate = ''
                host = ''
                isotype = ''
                solubility = ''
                source = ''
                species_reactivity = ''
                immunogen = ''
                concentration = product_info['concentration'].values[0]
                specificity = ''
                sterility = ''
                components = ''
                format_value = ''
                name = ''

            if not prms_info.empty:
                price = prms_info['USD List Price'].values[0]
                un_num = prms_info['UN#'].values[0]
                packing_group = prms_info['Packing Group'].values[0]
                storage_temp = prms_info['Storage Temp'].values[0]
                ship_temp = prms_info['Ship Temp'].values[0]
    #             country_of_origin = prms_info['Country of Origin (most recent lot)'].values[0]
            else:
                price = product_info['price'].values[0]
                packing_group = product_info['packing_group'].values[0]
                ship_temp = ''
                storage_temp = ''
    #             if not lot_info.empty:
    #                 country_of_origin = lot_info['Country of Origin'].values[0]
    #             else:
    #                 country_of_origin = ''
                    
            if not origin_info.empty:
                country_of_origin = origin_info.loc[origin_info['Expiration date -'] == max(origin_info['Expiration date -'].values)]['Country of Origin'].values[0]
            elif not prms_info.empty:
                country_of_origin = prms_info['Country of Origin (most recent lot)'].values[0]
            elif not lot_info.empty:
                country_of_origin = lot_info['Country of Origin'].values[0]
            else:
                country_of_origin = ''
                        
            if not unspsc_info.empty:
                unspsc = unspsc_info['UNSPSC'].values[0]
            else:
                unspsc = product_info['unspsc'].values[0]
            
            #name = product_info['name'].values[0].upper()
            # description = product_info['description'].values[0]
            short_desc = product_info['short_description'].values[0]
            img_link = product_info['base_image'].values[0]
            #price = product_info['price'].values[0]
            pack_size = product_info['pack_size_numeric_value'].values[0]
            unit = product_info['pack_size_unit_of_measure'].values[0]
            quantity = product_info['lk_packaging_facet'].values[0]
            #packing_group = product_info['packing_group'].values[0]
            #ship_temp = product_info['ship_conditions'].values[0]
            host = product_info['host'].values[0]
            # tariff_code = product_info['tariff_code'].values[0]
            msds_avail = skus['B'+str(i-1)].value
            cas_number = product_info['cas_number'].values[0]
            DOT_PSN = 'N/A'
            hazard_class = product_info['hazard_class'].values[0]
            biochem_physiol_actions = product_info['biochem_physiol_actions'].values[0]
            storage_and_handling = product_info['storage_and_handling'].values[0]
            group_name = product_info['prms_group_name'].values[0]
            img_link = product_info['base_image'].values[0]
            hazard_statements = product_info['hazard_statements'].values[0]
            keywords = product_info['meta_keywords'].values[0]
            grade = product_info['grade'].values[0]
            ph = product_info['ph'].values[0]
            weight_in_lb = new_magento.loc[new_magento['sku'] == sku]['weight'].values[0]
            purity = product_info['purity'].values[0]
            molecular_weight = product_info['molecular_weight'].values[0]
            key_applications = product_info['key_applications'].values[0]
            
            if type(name) == str:
                name = tidyDescription(name)
                if (str(pack_size) + str(unit)) in name:
                    name = name.replace((str(pack_size) + str(unit)), '')
                elif str(pack_size_joined) in name:
                    name = name.replace(str(pack_size_joined), '')
                if name.endswith(' '):
                    name = name[:len(name)-1]
                if name.endswith(','):
                    name = name[:len(name)-1]
                if len(name) > 40:
                    name = name[:40]
                thomas.loc[i, 'INVENTORY DESCRIPTION\n(40 character maximum including spaces)\nMUST include voltage of equipment\nand/or package size to clarify differences amongst products. \nWE CAN NOT ACCEPT DUPLICATE DESCRIPTIONS FOR DIFFERENT PRODUCTS - see Examples Tab'] = name
            
            thomas.loc[i, 'PURCHASE UNIT OF MEASURE\n(EA PK CS ETC.)'] = 'EA'
            thomas.loc[i, 'SIZE QTY\nNumber needed Only'] = 1
            
            if type(price) != str:
                if str(sku).startswith('11'):
                    thomas.loc[i, 'PUoM COST'] = 0.89*price
                else:
                    thomas.loc[i, 'PUoM COST'] = 0.8*price
                thomas.loc[i, 'PUoM LIST'] = price
            
            if ship_temp == 'DI':
                thomas.loc[i, 'Type (Hazardous, Handling/Processing, Ice/Dry Ice, Container, Other)'] = 'Ice/Dry Ice'
                thomas.loc[i, 'Amount'] = '$10.00'
                thomas.loc[i, 'Per Item/Order?'] = 'ITEM'
                thomas.loc[i, 'SHIPPING CONDITIONS \n(RT ICE DRY ICE)'] = 'DRY ICE'
            elif ship_temp == 'CP':
                thomas.loc[i, 'SHIPPING CONDITIONS \n(RT ICE DRY ICE)'] = 'ICE'
            elif ship_temp == 'AM':
                thomas.loc[i, 'SHIPPING CONDITIONS \n(RT ICE DRY ICE)'] = 'RT'
            
            if type(cas_number) == str and len(cas_number) > 0:
                thomas.loc[i, 'Section\n(Supplies,\nInstruments,\nEquipment,\nChemicals)'] = 'Chemicals'
                thomas.loc[i, 'Grade(s)'] = grade
                # thomas.loc[i, 53] = pack_size_joined
                thomas.loc[i, 'Pkg. Type\n(Bottle Poly)'] = 'Bottle'
                thomas.loc[i, 'CAS #'] = cas_number
                thomas.loc[i, 'pH\n(for Buffers & Standards)'] = ph
                
            thomas.loc[i, 'DROP SHIP?\n(Y or N)'] = 'Y'
            thomas.loc[i, 'Shipping or Drop minimums or penalties?\nIf yes please state'] = 'NONE'
            thomas.loc[i, 'COUNTRY OF ORIGIN\n(Need only 1)'] = country_of_origin
            
            tariff_code = str(tariff_code)
            if len(tariff_code) >= 4:
                thomas.loc[i, 'HARMONIZATION CODE'] = tariff_code
            
            if weight_in_lb < 0.5:
                thomas.loc[i, 'HEIGHT (IN.)'] = 7
                thomas.loc[i, 'WIDTH (IN.)'] = 4
                thomas.loc[i, 'LENGTH (IN.)'] = 5
                thomas.loc[i, 'CUBIC FEET'] = 0.081
            elif weight_in_lb <= 1:
                thomas.loc[i, 'HEIGHT (IN.)'] = 12
                thomas.loc[i, 'WIDTH (IN.)'] = 7
                thomas.loc[i, 'LENGTH (IN.)'] = 5
                thomas.loc[i, 'CUBIC FEET'] = 0.243
            else:
                thomas.loc[i, 'HEIGHT (IN.)'] = 12
                thomas.loc[i, 'WIDTH (IN.)'] = 12
                thomas.loc[i, 'LENGTH (IN.)'] = 12
                thomas.loc[i, 'CUBIC FEET'] = 1
                
            thomas.loc[i, 'lbs'] = weight_in_lb
            
            if type(storage_temp) == str:
                if storage_temp == 'AM':
                    thomas.loc[i, 'STORAGE CONDITIONS \n(RT 4°C -20°C -80°C)'] = 'RT'
                    thomas.loc[i, 'Refrigeration Requirements\n(Refrigerator or Freezer)'] = 'N/A'
                elif storage_temp == 'FR':
                    thomas.loc[i, 'STORAGE CONDITIONS \n(RT 4°C -20°C -80°C)'] = '-20°C'
                    thomas.loc[i, 'Refrigeration Requirements\n(Refrigerator or Freezer)'] = 'Freezer'
                elif storage_temp == '70' or storage_temp == '80':
                    thomas.loc[i, 'STORAGE CONDITIONS \n(RT 4°C -20°C -80°C)'] = '-80°C'
                    thomas.loc[i, 'Refrigeration Requirements\n(Refrigerator or Freezer)'] = 'Freezer'
                elif storage_temp == 'RF':
                    thomas.loc[i, 'STORAGE CONDITIONS \n(RT 4°C -20°C -80°C)'] = '4°C'
                    thomas.loc[i, 'Refrigeration Requirements\n(Refrigerator or Freezer)'] = 'Refrigerator'
                # else:
                #     thomas.loc[i, 32] = storage_temp
            #thomas.loc[i, 32] = storage_temp
            
            if not lot_info.empty:
                creation_date = np.datetime64(lot_info['Creation date -'].values[0])
                expiration_date = np.datetime64(lot_info['Expiration date -'].values[0])
                shelf_life = expiration_date - creation_date
                shelf_life = shelf_life.astype('timedelta64[D]')/np.timedelta64(1, 'D')
                if shelf_life > 0:
                    thomas.loc[i, 'SHELF LIFE'] = str(shelf_life) + ' days'
            
            thomas.loc[i, 'MSDS CODE \n( A B C OR D)'] = 'D'
            thomas.loc[i, 'REPLACES  PART#\n(if applicable enter PT# being replaced)'] = 'No'
            thomas.loc[i, 'Pkg. Size\n(4 L)'] = pack_size_joined
            
            if type(name) == str:
                name = tidyDescription(name)
                if (not (str(pack_size) + str(unit)) in name) and not ((str(pack_size_joined) in name)):
                    # if len(name) > 40-len(str(pack_size_joined)):
                    #     name = name[:39-len(pack_size_joined)] + ' ' + pack_size_joined
                    # else:
                    name = name + ' ' + str(pack_size_joined)
                # if product_type == 'configurable':
                #     thomas.loc[i, 'WEBSITE DESCRIPTION\n\nThis is a simple line listing description.\ninclude voltage of equipment \nWE CAN NOT ACCEPT DUPLICATE DESCRIPTIONS FOR DIFFERENT PRODUCTS - see Examples Tab'] = name + ' ' + str(pack_size_joined)
                # else:
                thomas.loc[i, 'WEBSITE DESCRIPTION\n\nThis is a simple line listing description.\ninclude voltage of equipment \nWE CAN NOT ACCEPT DUPLICATE DESCRIPTIONS FOR DIFFERENT PRODUCTS - see Examples Tab'] = name
            
            thomas.loc[i, 'If Yes please provide mfr. number or product to group with'] = 'N/A'

            if type(description) == str and len(description) > 0:
                thomas.loc[i, "Plain text or HTML format as shown in the example.\nLarger product adds (more than 5000 individual items) may require a Data Pull. For questions, contact Mike Kortonick in Web Operations: MikeK@thomassci.com or 856-340-8166\nCOPY refers to the overall write up for the entire product block on Thomas' website.\nSee examples below."] = description
            elif not product_info_sept.empty:
                thomas.loc[i, "Plain text or HTML format as shown in the example.\nLarger product adds (more than 5000 individual items) may require a Data Pull. For questions, contact Mike Kortonick in Web Operations: MikeK@thomassci.com or 856-340-8166\nCOPY refers to the overall write up for the entire product block on Thomas' website.\nSee examples below."] = product_info_sept['application_notes'].values[0]
            
            specifications = ''
            if type(purity) == str and len(purity) > 0:
                specifications = specifications + 'Purity: ' + purity.replace('â‰¥', '')
            if type(molecular_weight) == str and len(molecular_weight) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Molecular Weight: ' + molecular_weight
            if type(key_applications) == str and len(key_applications) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Key Applications: ' + key_applications
            if type(concentration) == str and len(concentration) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Concentration: ' + concentration
            if type(clone_name) == str and len(clone_name) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Clone Name: ' + clone_name
            if type(conjugate) == str and len(conjugate) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Conjugate: ' + conjugate
            if type(host) == str and len(host) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Host: ' + host
            if type(isotype) == str and len(isotype) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Isotype: ' + isotype
            if type(solubility) == str and len(solubility) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Solubility: ' + solubility
            if type(source) == str and len(source) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Source: ' + source
            if type(species_reactivity) == str and len(species_reactivity) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Species Reactivity: ' + species_reactivity
            if type(immunogen) == str and len(immunogen) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Immunogen: ' + immunogen
            if type(specificity) == str and len(specificity) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Specificity: ' + specificity
            if type(sterility) == str and len(sterility) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Sterility: ' + sterility
            if type(components) == str and len(components) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Components: ' + components
            if type(format_value) == str and len(format_value) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Format: ' + format_value
            
            thomas.loc[i, 'Plain text or HTML format; \nor prodvide hyperlink to specifications'] = specifications
            
            if type(keywords) == str:
                thomas.loc[i, 'Synonyms Nouns and Adjectives used to find the product in a catalog or website'] = keywords.replace(',', ' ')
            
            # thomas.loc[i, 69] = img_link
            
    new_thomas = opxl.load_workbook('forms/thomas_form.xlsx')
    thomas_sheet = new_thomas.active
    if thomas_sheet.max_row < skus.max_row:
        for j in range(skus.max_row - thomas_sheet.max_row):
            thomas_sheet.insert_rows(thomas_sheet.max_row-1)
    i = 1
    for row in thomas_sheet.iter_rows(min_row=23):
        if i < skus.max_row:
            for j in range(len(thomas.columns)):
                row[j].value = thomas.iloc[i+2, j]
        else:
            break
        i = i+1

    # new_thomas.save('../../outputs/old_product_outputs/old_thomas_output.xlsx')
    new_thomas.save('outputs/old_product_outputs/old_thomas_output.xlsx')

def fillFisher_Old(filename, magento, new_magento, lot_master, prms, unspsc_codes, origin, magento_sept):
    fisher_file = pd.ExcelFile('forms/fisher_form.xlsx')
    fisher = pd.read_excel(fisher_file, 'General Info', dtype=object)
    regulatory = pd.read_excel(fisher_file, 'Regulatory', dtype=object)
    new_columns = [i.strip() for i in regulatory.columns]
    regulatory.columns = new_columns
    new_columns = [i.strip() for i in fisher.columns]
    fisher.columns = new_columns
    wb = opxl.load_workbook(filename)
    skus = wb.active
    for i in range(2, skus.max_row+1):
        regulatory.loc[i-2, 'Supplier Catalog Number'] = str(skus['A'+str(i)].value)
    for i in range(2, skus.max_row+1):
        fisher.loc[i-2, 'Supplier Catalog Number'] = str(skus['A'+str(i)].value)

    for i in range(len(regulatory)):
        sku = regulatory['Supplier Catalog Number'][i]
        product_info = magento.loc[magento['sku'] == sku]
        lot_info = lot_master.loc[lot_master['Product number'] == sku]
        prms_info = prms.loc[prms['SKU'] == sku]
        unspsc_info = unspsc_codes.loc[unspsc_codes['Part Number'] == sku]
        product_info_july = new_magento.loc[new_magento['sku'] == sku]
        product_info_sept = magento_sept.loc[magento_sept['sku'] == sku]     

        if not product_info_july.empty:
            tariff_code = product_info_july['tariff_code'].values[0]
        else:
            tariff_code = ''
        if not product_info.empty:
            if not product_info_sept.empty:
                cas_number = product_info_sept['cas_number'].values[0]
                pack_size = product_info_sept['pack_size_joined'].values[0]
            else:
                cas_number = product_info['cas_number'].values[0]
                pack_size = product_info['pack_size_joined'].values[0]

            if not prms_info.empty:
                price = prms_info['USD List Price'].values[0]
                hazard_statements = prms_info['UN#'].values[0]
                packing_group = prms_info['Packing Group'].values[0]
                alt_storage_temp = prms_info['Storage Temp'].values[0]
                ship_temp = prms_info['Ship Temp'].values[0]
                country_of_origin = prms_info['Country of Origin (most recent lot)'].values[0]
                cas_number = product_info['cas_number'].values[0]
                storage_temp = prms_info['Storage Temp'].values[0]
                
            else:
                price = product_info['price'].values[0]
                packing_group = product_info['packing_group'].values[0]
                ship_temp = product_info['ship_conditions'].values[0]
                hazard_statements = product_info['hazard_statements'].values[0]
                storage_temp = ''
                
            if not unspsc_info.empty:
                unspsc = unspsc_info['UNSPSC'].values[0]
            else:
                unspsc = product_info['unspsc'].values[0]
            
            short_desc = product_info['short_description'].values[0]

            quantity = product_info['lk_packaging_facet'].values[0]
            host = product_info['host'].values[0]
            msds_avail = skus['B'+str(i+1)].value
            DOT_PSN = 'N/A'
            hazard_class = product_info['hazard_class'].values[0]
            biochem_physiol_actions = product_info['biochem_physiol_actions'].values[0]
            storage_and_handling = product_info['storage_and_handling'].values[0]
            group_name = product_info['prms_group_name'].values[0]
            img_link = product_info['base_image'].values[0]
            keywords = product_info['meta_keywords'].values[0]

            if type(pack_size) == str:
                unit = ''.join([i for i in pack_size if not i.isdigit()])
            else:
                unit = ''
            
            if sku.startswith('07') or sku.startswith('08') or sku.startswith('09') or sku.startswith('11'):
                regulatory['Primary Commodity Code'][i] = 'Diagnostics'
            elif type(cas_number) == str and len(cas_number) > 0:
                regulatory['Primary Commodity Code'][i] = 'Chemicals'
            elif 'preps' in unit or 'Preps' in unit:
                regulatory['Primary Commodity Code'][i] = 'Consumables'
            elif sku.startswith('02') or sku.startswith('03') or sku.startswith('04') or sku.startswith('05'):
                regulatory['Primary Commodity Code'][i] = 'Diagnostics'

            regulatory['Safety Data Sheet Code'][i] = 99998
            
            if type(storage_temp) == str:
                if storage_temp == 'AM':
                    regulatory['Storage Code'][i] = 'GWN4'
                elif storage_temp == 'FR' or storage_temp == '70' or storage_temp == '80':
                    regulatory['Storage Code'][i] = 'DFD1'
                elif storage_temp == 'RF':
                    regulatory['Storage Code'][i] = 'RFC2'
                else:
                    regulatory['Storage Code'][i] = storage_temp
            else:
                regulatory['Storage Code'][i] = 'GWN4'
            
            regulatory['STERILE'][i] = 'N'
            regulatory['Proposition 65'][i] = 'N'
            regulatory['Latex'][i] = 'N'
            regulatory['Lithium Battery'][i] = 'N'
            regulatory['Medical Device'][i] = 'None'
            regulatory['UPC - Standard'][i] = 'NA'
            
            tariff_code = str(tariff_code).replace('.', '')
            if len(tariff_code) >= 4:
                regulatory['Harmonized Tariff Schedule Code'][i] = tariff_code[:4] + '999999'
                
            regulatory['REACH Compliant'][i] = 'N'
            regulatory['RoHS Compliant'][i] = 'N'
            regulatory['Mercury'][i] = 'N'
            regulatory['Benzene %'][i] = 'N'
            regulatory['Asbestos'][i] = 'N'
            regulatory['Iodine'][i] = 'N'
            regulatory['Radioactive Materials'][i] = 'N'
            regulatory['Pesticides'][i] = 'N'
            regulatory['Ethyl Alcohol'][i] = 'N'
            regulatory['DEA List 1 Chemical or Drug'][i] = 'N'
            regulatory['WHMIS Regulated?\n(Y/N)'][i] = 'N'
            regulatory['Marine Pollutant'][i] = 'N'
            regulatory['Is USMCA/CUSMA Certificate available?'][i] = 'N'
            regulatory["Federal  or Int'l Regulations"][i] = 'NA'
                
            if type(host) == str and len(host) > 0:
                regulatory['Animal or Human Origin'][i] = 'Y'
            else:
                regulatory['Animal or Human Origin'][i] = 'N'
            
            regulatory['MDL Number'][i] = 'NA'

            if type(cas_number) == str and not cas_number == 'Not applicable':
                regulatory['Chemical Abstracts Number (CAS 1)'][i] = cas_number
                regulatory['Chemical Abstracts Percentage(1)'][i] = 100

    for i in range(len(fisher)):
        sku = fisher['Supplier Catalog Number'][i]
        product_info = magento.loc[magento['sku'] == sku]
        lot_info = lot_master.loc[lot_master['Product number'] == sku]
        prms_info = prms.loc[prms['SKU'] == sku]
        unspsc_info = unspsc_codes.loc[unspsc_codes['Part Number'] == sku]
        product_info_sept = magento_sept.loc[magento_sept['sku'] == sku]


        if not product_info.empty:
            if not product_info_sept.empty:
                name = product_info_sept['name'].values[0].upper()
                pack_size = product_info_sept['pack_size_joined'].values[0]
            else:
                name = product_info['name'].values[0].upper()
                pack_size = product_info['pack_size_joined'].values[0]


            if not prms_info.empty:
                price = prms_info['USD List Price'].values[0]
                hazard_statements = prms_info['UN#'].values[0]
                packing_group = prms_info['Packing Group'].values[0]
                alt_storage_temp = prms_info['Storage Temp'].values[0]
                ship_temp = prms_info['Ship Temp'].values[0]
                country_of_origin = prms_info['Country of Origin (most recent lot)'].values[0]
                if type(country_of_origin) == str and len(country_of_origin) > 0:
                    country_of_origin = pc.country_alpha3_to_country_alpha2(country_of_origin)
                    fisher['Canada Availability'][i] = 'Y'
                    fisher['Currency'][i] = 'USD'
                else:
                    fisher['Canada Availability'][i] = 'N'
                
            else:
                price = product_info['price'].values[0]
                packing_group = product_info['packing_group'].values[0]
                ship_temp = product_info['ship_conditions'].values[0]
                hazard_statements = product_info['hazard_statements'].values[0]
                
            if not unspsc_info.empty:
                unspsc = unspsc_info['UNSPSC'].values[0]
            else:
                unspsc = product_info['unspsc'].values[0]
            
            short_desc = product_info['short_description'].values[0]
            img_link = product_info['base_image'].values[0]
            quantity = product_info['lk_packaging_facet'].values[0]
            host = product_info['host'].values[0]
            tariff_code = product_info['tariff_code'].values[0]
            cas_number = product_info['cas_number'].values[0]
            DOT_PSN = 'NA'
            hazard_class = product_info['hazard_class'].values[0]
            biochem_physiol_actions = product_info['biochem_physiol_actions'].values[0]
            storage_and_handling = product_info['storage_and_handling'].values[0]
            group_name = product_info['prms_group_name'].values[0]
            img_link = product_info['base_image'].values[0]
            keywords = product_info['meta_keywords'].values[0]
            storage_temp = product_info['storage_and_handling'].values[0]
            if not new_magento.loc[new_magento['sku'] == sku].empty:
                weight_in_lb = new_magento.loc[new_magento['sku'] == sku]['weight'].values[0]
                
                fisher['Standard Unit Weight'][i] = weight_in_lb
            
                if weight_in_lb < 0.5:
                    fisher['Standard Unit Length'][i] = 7
                    fisher['Standard Unit Width'][i] = 4
                    fisher['Standard Unit Height'][i] = 5
                    fisher['Units on a Pallet'][i] = 60
                    fisher['Pallet layers'][i] = 36
                elif weight_in_lb <= 1:
                    fisher['Standard Unit Length'][i] = 12
                    fisher['Standard Unit Width'][i] = 7
                    fisher['Standard Unit Height'][i] = 5
                    fisher['Units on a Pallet'][i] = 18
                    fisher['Pallet layers'][i] = 36
                else:
                    fisher['Standard Unit Length'][i] = 12
                    fisher['Standard Unit Width'][i] = 12
                    fisher['Standard Unit Height'][i] = 12
                    fisher['Units on a Pallet'][i] = 12
                    fisher['Pallet layers'][i] = 15
            

            if type(name) == str:
                name = tidyDescription(name)
                temp_name = ''
                for c in name:
                    if c.isalpha() or c.isnumeric() or c == ' ':
                        temp_name = temp_name + c
                name = temp_name
                if len(name) <= 30:
                    fisher['Item Description - 30 Characters'][i] = name
                else:
                    fisher['Item Description - 30 Characters'][i] = name[:30]
                
                if len(name) <= 240:
                    fisher['Item Description - 240 Characters'][i] = name
                else:
                    fisher['Item Description - 240 Characters'][i] = name[:240]
                    
            fisher['Standard Unit'][i] = 'EA'
            fisher['Standard Unit Quantity'][i] = 1
            fisher['Unit of Use'][i] = 'EA'

            if type(pack_size) == str:
                quant = ''.join([i for i in pack_size if i.isdigit()])
                unit = ''.join([i for i in pack_size if not i.isdigit()])
                fisher['Package Size Quantity'][i] = quant
                if 'preps' in unit or 'Preps' in unit:
                    unit = 'PP'
                elif 'mL' in unit or 'ml' in unit:
                    unit = 'ML'
                elif 'lb' in unit:
                    unit = 'LB'
                elif 'tests' in unit or 'Tests' in unit:
                    unit = 'TS'
                elif 'KU' in unit or 'ku' in unit:
                    unit = 'KU'
                elif 'mg' in unit:
                    unit = 'MG'
                elif 'U' in unit or 'wells' in unit or 'Bags' in unit or 'Bottle' in unit or 'Each' in unit or 'Kit' in unit or 'mCI' in unit:
                    unit = 'UN'
                elif unit == 'g':
                    unit = 'GR'
                elif unit == 'l' or unit == 'L' or 'Liter' in unit:
                    unit = 'LT'
                else:
                    unit = 'UN'
                fisher['Package Size Unit'][i] = unit
            
            if type(price) != str:
                fisher['Standard Unit Cost'][i] = price*0.7
                fisher['Standard Unit List'][i] = price
            
            fisher['Discount %'][i] = '30'
            fisher['Pricing Expiration Date'][i] = '31/DEC/2020'
            
            fisher['UNSPSC Code'][i] = unspsc
            
            if type(hazard_statements) == str and len(hazard_statements) > 0:
                fisher['Hazardous'][i] = 'Y'
            else:
                fisher['Hazardous'][i] = 'N'
            
            fisher['UN/NA#'][i] = hazard_statements
            fisher['Hazard Class'][i] = hazard_class
            fisher['Packing Group'][i] = packing_group

            fisher['Shelf Life'][i] = 'N'
            fisher['Shelf Life  (Days)'][i] = 0
            
            if type(keywords) == str:
                keywords = keywords.split(',')
                for j in range(len(keywords)):
                    if len(keywords[j]) > 11:
                        keywords[j] = keywords[j][:11]
                keywords = list(set(keywords))
                j = 0
                while j<5 and j<len(keywords):
                    fisher.iloc[i, j+6] = keywords[j]
                    fisher.iloc[i, j+6] = keywords[j]
                    j = j+1
                    
            fisher['Manufacturer Lead-Time'][i] = '25'
            fisher['Green Product'][i] = 'N'
            fisher['Country of Origin'][i] = country_of_origin
            fisher['Build to Stock or Build to Order'][i] = 'Build to Order'
            fisher['Serial / Lot Control'][i] = 'NA'
            fisher['Expiration'][i] = 'Y'
            fisher['Product Info on Shipping Box'][i] = 'N'
            fisher['Expiration Info on Shipping Box'][i] = 'Y'
            fisher['Minimum Order Quantity'][i] = 1
            fisher['Certificates Available'][i] = 'Y'
            fisher['Certificate of Analysis'][i] = 'Y'
            fisher['Installation'][i] = 'N'
            
            
    new_fisher = opxl.load_workbook('forms/fisher_form.xlsx')
    gen_sheet = new_fisher['General Info']
    regulatory_sheet = new_fisher['Regulatory']
    if gen_sheet.max_row < skus.max_row:
        for j in range(skus.max_row - gen_sheet.max_row+1):
            gen_sheet.insert_rows(gen_sheet.max_row-1)
    if regulatory_sheet.max_row < skus.max_row:
        regulatory_sheet.append([''])
        for j in range(skus.max_row - regulatory_sheet.max_row + 1):
            regulatory_sheet.insert_rows(regulatory_sheet.max_row)
    i = 0
    for row in gen_sheet.iter_rows(min_row=2):
        if i < len(fisher):
            for j in range(len(fisher.columns)-1):
                if not fisher.iloc[i, j] == 'None':
                    row[j].value = fisher.iloc[i, j]
        else:
            break
        i = i+1
    i = 0
    for row in regulatory_sheet.iter_rows(min_row=2):
        if i < len(regulatory):
            for j in range(len(regulatory.columns)-1):
                if not fisher.iloc[i, j] == 'None':
                    row[j].value = regulatory.iloc[i, j]
        else:
            break
        i = i+1

    new_fisher.save('outputs/old_product_outputs/old_fisher_output.xlsx')
    return fisher
    # new_fisher.save('../../outputs/old_product_outputs/old_fisher_output.xlsx')

def attributeLookup(attribute, product_info, product_info_sept, prms_info, lot_info, unspsc_info, origin_info, abcam_info, sku, magento):

    if attribute == 'Form':
        try:
            pubchem_data = getPubchemData(sku, magento)
            if not pubchem_data == None:
                form = pubchem_data['Color/Form']
                return form + " | pulled from Pubchem"
        except:
            print('Pubchem Error')

    if not prms_info.empty:
        storage_condition = prms_info['Storage Temp'].values[0]
        ship_temp = prms_info['Ship Temp'].values[0]
        # un_num = prms_info['UN#'].values[0]
        
        if storage_condition == 'AM':
            storage_condition = 'Room Temperature'
            storage_code = 'GWN4'
        elif storage_condition == 'RF':
            storage_condition = 'Refrigerated'
            storage_code = 'RFC2'
        elif storage_condition == 'FR':
            storage_condition = 'Frozen'
            storage_code = 'DFD1'
        else:
            storage_code = ''
            
        if ship_temp == 'AM':
            ship_temp = 'Ambient'
        elif ship_temp == 'CP':
            ship_temp = 'Cold Pack'
        elif ship_temp == 'DI':
            ship_temp = 'Dry Ice'
    else:
        storage_condition = ''
        storage_code = ''
        ship_temp = ''
        # un_num = ''
        
    if not lot_info.empty:
        creation_date = np.datetime64(lot_info['Creation date -'].values[0])
        expiration_date = np.datetime64(lot_info['Expiration date -'].values[0])
        shelf_life = expiration_date - creation_date
        shelf_life = shelf_life.astype('timedelta64[D]')/np.timedelta64(1, 'D')
    else:
        shelf_life = ''
        
    if not origin_info.empty:
        country_of_origin = origin_info.loc[origin_info['Expiration date -'] == max(origin_info['Expiration date -'].values)]['Country of Origin'].values[0]
    elif not prms_info.empty:
        country_of_origin = prms_info['Country of Origin (most recent lot)'].values[0]
    elif not lot_info.empty:
        country_of_origin = lot_info['Country of Origin'].values[0]
    else:
        country_of_origin = ''

    if not product_info_sept.empty:
        antibody_type = product_info_sept['antibody_type'].values[0]
        antibody_type = str(antibody_type)
        if 'Polyclonal' in antibody_type or 'polycolnal' in antibody_type:
            antibody_type = 'Polyclonal'
        elif 'Monoclonal' in antibody_type or 'monoclonal' in antibody_type:
            antibody_type = 'Monoclonal'
        else:
            antibody_type = ''

        keywords = product_info_sept['keywords'].values[0]
        keywords = str(keywords)
        if 'Primary' in keywords or 'primary' in keywords:
            keywords = 'Primary'
        elif 'Secondary' in keywords or 'secondary' in keywords:
            keywords = 'Secondary'
        else:
            keywords = ''

        short_desc = product_info_sept['short_description'].values[0]
        short_desc = tidyDescription(str(short_desc))

        application_notes = product_info_sept['application_notes'].values[0]
        application_notes = str(application_notes)
        if 'strain' in application_notes or 'Strain' in application_notes:
            application_notes = 'Strain'
        elif 'genotype' in application_notes or 'Genotype' in application_notes:
            application_notes = 'Genotype'
        else:
            application_notes = ''

        host = product_info_sept['host'].values[0]

        un_number = product_info_sept['un_number'].values[0]
        if not str(un_number).startswith('UN'):
            un_number = ''

    else:
        antibody_type = ''
        keywords = ''
        short_desc = ''
        application_notes = ''
        host = ''
        un_number = ''

    if not abcam_info == None:
        purity = abcam_info['Purity'] + " | pulled from Abcam"
        immunogen = abcam_info['Immunogen'] + " | pulled from Abcam"
        isotype = abcam_info['Isotype'] + " | pulled from Abcam"
        function = abcam_info['Function'] + " | pulled from Abcam"
        concentration = abcam_info['Concentration'] + " | pulled from Abcam"
        clonality = abcam_info['Clonality'] + " | pulled from Abcam"
        try:
            host_species = abcam_info['Host species'] + " | pulled from Abcam"
        except:
            host_species = abcam_info['Host Species'] + " | pulled from Abcam"
        clone_number = abcam_info['Clone number'] + " | pulled from Abcam"
        light_chain_type = abcam_info['Light chain type'] + " | pulled from Abcam"
        species_reactivity = abcam_info['Species reactivity'] + " | pulled from Abcam"
    else:
        purity = ''
        immunogen = ''
        isotype = ''
        function = ''
        concentration = ''
        clonality = ''
        host_species = ''
        clone_number = ''
        light_chain_type = ''
        species_reactivity = ''

    if type(light_chain_type) == str and len(light_chain_type) > 0:
        light_chain_type = 'Chains'
    else:
        light_chain_type = ''
        
    attribute_dict = {}

    if not product_info_sept.empty:
        attribute_dict['Sterility'] = product_info_sept['sterilization_of_solutions'].values[0]
        attribute_dict['Synonym'] = product_info_sept['alternate_names'].values[0]
        attribute_dict['Formula Weight'] = product_info_sept['molecular_weight'].values[0]
        attribute_dict['For Use With'] = product_info_sept['application_notes'].values[0]
        attribute_dict['Linear Formula'] = product_info_sept['molecular_formula'].values[0]
        attribute_dict['absorbance'] = product_info_sept['uv_visible_absorbance'].values[0]
        attribute_dict['Absorbance'] = product_info_sept['uv_visible_absorbance'].values[0]
        attribute_dict['Ignition Point'] = product_info_sept['auto_ignition'].values[0]
        attribute_dict['Enzyme'] = product_info_sept['protein_or_enzyme_type'].values[0]
        attribute_dict['Validated Application'] = product_info_sept['application_notes'].values[0]
        attribute_dict['Research Category'] = product_info_sept['key_applications'].values[0]
        attribute_dict['Conjugate'] = product_info_sept['conjugate'].values[0]
        attribute_dict['Grade'] = product_info_sept['grade'].values[0]
        attribute_dict['Protein Family'] = product_info_sept['protein_or_enzyme_type'].values[0]
        attribute_dict['pH'] = product_info_sept['ph'].values[0]
        attribute_dict['Sample Size'] = product_info_sept['pack_size_joined'].values[0]
        attribute_dict['Molecular Formula'] = product_info_sept['molecular_formula'].values[0]
        attribute_dict['Molecular Weight (g/mol)'] = product_info_sept['molecular_weight'].values[0]
        attribute_dict['Sample Volume'] = product_info_sept['pack_size_joined'].values[0]
        attribute_dict['Density'] = product_info_sept['density'].values[0]
        attribute_dict['Sufficient For'] = product_info_sept['key_applications'].values[0]
        attribute_dict['Flash Point'] = product_info_sept['flash_point'].values[0]
        attribute_dict['Packaging Quantity'] = product_info_sept['pack_size_joined'].values[0]
        attribute_dict['Melting Point'] = product_info_sept['melting_point'].values[0]
        attribute_dict['pH Range'] = product_info_sept['ph'].values[0]
        attribute_dict['Sensitivity'] = product_info_sept['sensitivity'].values[0]
        attribute_dict['Test Sensitivity'] = product_info_sept['sensitivity'].values[0]
        attribute_dict['Monoclonal or Polyclonal'] = antibody_type
        attribute_dict['Formulation'] = product_info_sept['formulation'].values[0]
        attribute_dict['Green Features'] = 'NA'
        attribute_dict['CAS'] = product_info_sept['cas_number'].values[0]
        attribute_dict['Melting Temp_Begin'] = product_info_sept['melting_point'].values[0]
        attribute_dict['Primary or Secondary'] = keywords
        attribute_dict['Description'] = short_desc
        attribute_dict['Strain or Genotype'] = application_notes
        attribute_dict['Shipping Condition'] = ship_temp
        attribute_dict['Host Species'] = host
        attribute_dict['CAS Max %'] = '100'
        attribute_dict['Molecular Weight of Antigen'] = product_info_sept['molecular_weight'].values[0]
        attribute_dict['Solubility'] = product_info_sept['solubility'].values[0]
        attribute_dict['Boiling Point'] = product_info_sept['boiling_point'].values[0]
        attribute_dict['UN Number'] = un_number
        attribute_dict['Sterility'] = product_info_sept['sterility'].values[0]
        attribute_dict['CAS Min %'] = '0'
        attribute_dict['Antibody Molecular Weight'] = product_info_sept['molecular_weight'].values[0]
        attribute_dict['Origin'] = product_info_sept['concentration'].values[0]
        attribute_dict['Boiling Range'] = product_info_sept['boiling_point'].values[0]
        attribute_dict['For Use With (Application)'] = product_info_sept['upsell_position'].values[0]

    if not prms_info.empty:
        attribute_dict['Content And Storage'] = storage_condition
        attribute_dict['Storage Requirements'] = storage_condition
        attribute_dict['Storage'] = storage_condition

    if not lot_info.empty:
        attribute_dict['Manufacturing Origin'] = country_of_origin
        attribute_dict['Country of Origin'] = country_of_origin

    if not abcam_info == None:
        attribute_dict['Purity'] = purity
        attribute_dict['Purity Grade Notes'] = purity
        attribute_dict['Sub Class Isotype'] = isotype
        attribute_dict['Isotype'] = isotype
        attribute_dict['Key Functions'] = function
        attribute_dict['Target Isotype'] = isotype
        attribute_dict['Glucose Concentration'] = concentration
        attribute_dict['Clonality'] = clonality
        attribute_dict['Host abrv'] = host_species
        attribute_dict['Percent Purity'] = purity
        attribute_dict['Concentration'] = concentration
        attribute_dict['Host Cell'] = host_species
        attribute_dict['Clone ID'] = clone_number
        attribute_dict['Concentration Ratio'] = concentration
        attribute_dict['Chains or Fragments'] = light_chain_type
        attribute_dict['Concentration or Composition (by Analyte or Components)'] = concentration
        attribute_dict['Concentration or Composition Notes'] = concentration
        attribute_dict['Endotoxin Concentration'] = concentration
        attribute_dict['Target Function'] = function
        attribute_dict['Purity or Quality Grade'] = purity
        attribute_dict['Reactivity'] = species_reactivity
        attribute_dict['BIS MSB Concentration'] = concentration
        attribute_dict['Enzyme Function'] = function

    
    if attribute in attribute_dict:
        return attribute_dict[attribute]
    else:
        return ''

def fillFisher_Enrichment(filename, magento, new_magento, lot_master, prms, magento_sept, unspsc_codes, origin, images):
    authoring_file = pd.ExcelFile(filename)

    authoring = pd.read_excel(authoring_file, 'Core_Content', dtype=object)
    authoring.columns = authoring.iloc[2]
    authoring.columns = [i.strip() for i in authoring.columns]

    attributes = pd.read_excel(authoring_file, 'Category_Attributes', dtype=object)
    attributes.columns = attributes.iloc[2]
    attributes.columns = [i.strip() for i in attributes.columns]
        
    for i in range(3, len(authoring)):
        sku = authoring['manufacturerPartNumber'][i]
        product_info = magento.loc[magento['sku'] == sku]
        product_info_sept = magento_sept.loc[magento_sept['sku'] == sku]
        prms_info = prms.loc[prms['SKU'] == sku]
        images_info = images.loc[images['sku'] == sku]
        if not product_info.empty:
            name = product_info['name'].values[0]
            short_description = product_info['short_description'].values[0]
            description = product_info['description'].values[0]
            keywords = product_info['meta_keywords'].values[0]
            pack_size_joined = product_info['pack_size_joined'].values[0]

            if not images_info.empty:
                image = images['base_image'].values[0]
            else:
                image = ''
            
            if type(pack_size_joined) == str:
                authoring['productTitle'][i] = name.replace(pack_size_joined, '')
            else:
                authoring['productTitle'][i] = name
                
            authoring['skuDifferentiatorText'][i] = name
            authoring['teaserText'][i] = short_description
            authoring['productFeatures'][i] = description
            authoring['Image'][i] = image
            authoring['keywords'][i] = keywords
        
        if not prms_info.empty:
            un_number = prms_info['UN#'].values[0]
            
            authoring['alerts'][i] = un_number
    
    for i in range(3, len(attributes)-1):
        sku = attributes['manufacturerPartNumber'][i]
        product_info = new_magento.loc[new_magento['sku'] == sku]
        product_info_sept = magento_sept.loc[magento_sept['sku'] == sku]
        prms_info = prms.loc[prms['SKU'] == sku]
        lot_info = lot_master.loc[lot_master['Product number'] == sku]
        unspsc_info = unspsc_codes.loc[unspsc_codes['Part Number'] == sku]
        origin_info = origin.loc[origin['Product number'] == sku]
        
        attribute_name = attributes['Attribute_Name'][i]
        
        # abcam_info = getAbcamData(sku, magento_sept)
        abcam_info = None # replace this

        attributes['Values'][i] = attributeLookup(attribute_name, product_info, product_info_sept, prms_info, lot_info, unspsc_info, origin_info, abcam_info, sku, magento)
        # add this back

    wb_enrichment = opxl.load_workbook(filename)
    core_content = wb_enrichment['Core_Content']
    i = 3
    for row in core_content.iter_rows(min_row=5):
        for j in range(len(authoring.columns)):
            row[j].value = authoring.iloc[i, j]
        i = i+1
    
    attribute_sheet = wb_enrichment['Category_Attributes']
    i = 3
    for row in attribute_sheet.iter_rows(min_row=5):
        if i < len(attributes):
            row[11].value = attributes['Values'][i]
        else:
            break
        i = i+1

    wb_enrichment.save('outputs/enrichment_outputs/fisher_enrichment_output.xlsx')
    # wb_enrichment.save('../../outputs/old_product_outputs/fisher_enrichment_output.xlsx')

def fillVWR_Enrichment_Antibodies(enrichment, magento, prms):
    for i in range(14, len(enrichment)):
        sku = enrichment.loc[i, 'Supplier Cat. No.']
        product_info = magento.loc[magento['sku'] == sku]
        prms_info = prms.loc[prms['SKU'] == sku]
        
        if not prms_info.empty:
            ship_temp = prms_info['Ship Temp'].values[0]
            storage_temp = prms_info['Storage Temp'].values[0]
            
            if ship_temp == 'CP':
                enrichment.loc[i, 'Shipping Temperature'] = '2C to 8C'
            elif ship_temp == 'DI':
                enrichment.loc[i, 'Shipping Temperature'] = '-30C to -2C'
            elif ship_temp == 'AM':
                enrichment.loc[i, 'Shipping Temperature'] = '15C to 30C'
                
            if storage_temp == 'AM':
                enrichment.loc[i, 'Storage Temperature'] = '15C to 30C'
            elif storage_temp == 'FR':
                enrichment.loc[i, 'Storage Temperature'] = '-30C to -2C'
            elif storage_temp == 'RF':
                enrichment.loc[i, 'Storage Temperature'] = '2C to 8C'
            elif storage_temp == '70' or storage_temp == '80':
                enrichment.loc[i, 'Storage Temperature'] = '-70C'
        
        if not product_info.empty:
            name = product_info['name'].values[0]
            # description = product_info['description'].values[0]
            # short_desc = product_info['short_description'].values[0]
            pack_size_joined = product_info['pack_size_joined'].values[0]
            antibody_type = product_info['antibody_type'].values[0]
            host = product_info['host'].values[0]
            conjugate = product_info['conjugate'].values[0]
            clone_name = product_info['clone_name'].values[0]
            concentration = product_info['concentration'].values[0]
            immunogen = product_info['immunogen'].values[0]
            molecular_weight = product_info['molecular_weight'].values[0]
            purity = product_info['purity'].values[0]
            # application_notes = product_info['application_notes'].values[0]

            # enrichment.loc[i, 'Title / Short Description / Antibody Name'] = description

            # if type(short_desc) == str and len(short_desc) > 30:
            #     enrichment.loc[i, 'Short Description'] = short_desc[:30]
            # else:
            #     enrichment.loc[i, 'Short Description'] = short_desc

            # enrichment.loc[i, 'Long Text Description'] = description
            enrichment.loc[i, 'Size with unit'] = pack_size_joined

            if type(antibody_type) == str:
                enrichment.loc[i, 'Type'] = str(antibody_type).replace(' Antibody', '')
                enrichment.loc[i, 'Clonality'] = str(antibody_type).replace(' Antibody', '')

            enrichment.loc[i, 'Host'] = host
            enrichment.loc[i, 'Conjugation'] = conjugate
            
            if type(clone_name) == str and len(clone_name) > 0:
                enrichment.loc[i, 'Clone'] = 'Clone: ' + clone_name
                
            enrichment.loc[i, 'Reactivity'] = host
            enrichment.loc[i, 'Concentration'] = concentration
            enrichment.loc[i, 'Immunogen'] = immunogen
            enrichment.loc[i, 'Molecular Weight'] = molecular_weight
            enrichment.loc[i, 'Purification Method'] = purity

            if 'ELISA' in name or 'elisa' in name:
                enrichment.loc[i, 'ELISA'] = 'Yes'
            else:
                enrichment.loc[i, 'ELISA'] = 'No'

            # enrichment['Application Notes'][i] = application_notes

    new_enrichment = opxl.load_workbook('forms/GlobalProductEnrichmentFile_Antibodies (New).xlsx')
    sheet = new_enrichment.active
    if sheet.max_row < len(enrichment):
        sheet.append([''])
        for j in range(len(enrichment) - sheet.max_row + 4):
            sheet.insert_rows(sheet.max_row)
    i = 14
    for row in sheet.iter_rows(min_row=16):
        if i < (len(enrichment)):
            for j in range(len(enrichment.columns)):
                row[j].value = enrichment.iloc[i, j]
        else:
            break
        i += 1
            
    new_enrichment.save('outputs/enrichment_outputs/vwr_enrichment_antibody_output.xlsx')

def fillVWR_Enrichmnent_PPE(enrichment, magento, prms):
    for i in range(14, len(enrichment)):
        sku = enrichment.loc[i, 'Supplier Cat. No.']
        product_info = magento.loc[magento['sku'] == sku]
        prms_info = prms.loc[prms['SKU'] == sku]
        
        if not prms_info.empty:
            ship_temp = prms_info['Ship Temp'].values[0]
            storage_temp = prms_info['Storage Temp'].values[0]
            
            if ship_temp == 'CP':
                enrichment.loc[i, 'Shipping temperature'] = '2C to 8C'
            elif ship_temp == 'DI':
                enrichment.loc[i, 'Shipping temperature'] = '-30C to -2C'
            elif ship_temp == 'AM':
                enrichment.loc[i, 'Shipping temperature'] = '15C to 30C'
                
            if storage_temp == 'AM':
                enrichment.loc[i, 'Storage conditions'] = '15C to 30C'
            elif storage_temp == 'FR':
                enrichment.loc[i, 'Storage conditions'] = '-30C to -2C'
            elif storage_temp == 'RF':
                enrichment.loc[i, 'Storage conditions'] = '2C to 8C'
            elif storage_temp == '70' or storage_temp == '80':
                enrichment.loc[i, 'Storage conditions'] = '-70C'
        
        if not product_info.empty:
            name = product_info['name'].values[0]
            description = product_info['description'].values[0]
            # short_desc = product_info['short_description'].values[0]
            pack_size_joined = product_info['pack_size_joined'].values[0]
            # antibody_type = product_info['antibody_type'].values[0]
            host = product_info['host'].values[0]
            conjugate = product_info['conjugate'].values[0]
            # clone_name = product_info['clone_name'].values[0]
            concentration = product_info['concentration'].values[0]
            # immunogen = product_info['immunogen'].values[0]
            molecular_weight = product_info['molecular_weight'].values[0]
            purity = product_info['purity'].values[0]
            # application_notes = product_info['application_notes'].values[0]
            cas_number = product_info['cas_number'].values[0]
            
            # if type(short_desc) == str:
            #     short_desc = tidyDescription(short_desc)
            #     short_desc = ''.join([i for i in short_desc if i.isalnum() or i == ' '])
            #     if len(short_desc) > 30:
            #         enrichment.loc[i, 'short Description'] = short_desc[:30]
            #     else:
            #         enrichment.loc[i, 'short Description'] = short_desc
                    
            # if type(description) == str:
            #     description = tidyDescription(description)
            #     description = ''.join([i for i in description if i.isalnum() or i == ' '])
            #     enrichment['long Description'] = description
                
            if 'recombinant' in str(description) or 'Recombinant' in str(description):
                enrichment.loc[i, 'Protein/ Peptide/ Enzyme Type'] = 'recombinant'

            if type(name) == str:
                # enrichment.loc[i, 'Product Title'][i] = name + ' MP Biomedical'
                if 'protein' in name:
                    enrichment.loc[i, 'Product Class'] = 'protein'
                elif 'enzyme' in name:
                    enrichment.loc[i, 'Product Class'] = 'enzyme'
                elif 'peptide' in name:
                    enrichment.loc[i, 'Product Class'] = 'peptide'
                
            enrichment.loc[i, 'Size with unit'] = pack_size_joined
                
            enrichment.loc[i, 'Species'] = host
            enrichment.loc[i, 'Conjugation'] = conjugate
            enrichment.loc[i, 'CAS No'] = cas_number
            enrichment.loc[i, 'Purity'] = purity
            enrichment.loc[i, 'Molecular Weight'] = molecular_weight
            enrichment.loc[i, 'Concentration'] = concentration
            
    new_enrichment = opxl.load_workbook('forms/GlobalProductEnrichmentFile_Proteins_Peptides_Enzymes (New).xlsx')
    sheet = new_enrichment.active
    if sheet.max_row < len(enrichment):
        sheet.append([''])
        for j in range(len(enrichment) - sheet.max_row + 4):
            sheet.insert_rows(sheet.max_row)
    i = 14
    for row in sheet.iter_rows(min_row=16):
        if i < (len(enrichment)):
            for j in range(len(enrichment.columns)):
                row[j].value = enrichment.iloc[i, j]
        else:
            break
        i += 1

    new_enrichment.save('outputs/enrichment_outputs/vwr_enrichment_ppe_output.xlsx')

    return enrichment

def fillVWR_Enrichmnent_Sera(enrichment, magento):

    for i in range(11, len(enrichment)):
        sku = enrichment['Supplier Part No.'][i]
        product_info = magento.loc[magento['sku'] == sku]
        
        if not product_info.empty:
            pack_size_joined = product_info['pack_size_joined'].values[0]
            host = product_info['host'].values[0]
            concentration = product_info['concentration'].values[0]
            purity = product_info['purity'].values[0]
            application_notes = product_info['application_notes'].values[0]
            sterility = product_info['sterility'].values[0]
            keywords = product_info['keywords'].values[0]
            
            # if host == 'human' or host == 'Human':
            #     enrichment.loc[i, 'Serum Source'] = 'Human'
            # elif type(host) == str and len(host) > 0:
            #     enrichment.loc[i, 'Serum Source'] = 'Animal'
            if type(host) == str and len(host) > 0 and host != 'human' or host != "Human":
                enrichment.loc[i, 'Animal Serum Type'] = host
            
            enrichment.loc[i, 'Sterility'] = sterility
            enrichment.loc[i, 'Size'] = pack_size_joined
            enrichment.loc[i, 'Application'] = application_notes
            enrichment.loc[i, 'Concentration'] = concentration
            enrichment.loc[i, 'Purification'] = purity
            
            if type(keywords) == str:
                keywords = keywords.replace(', ', ';')
                keywords = keywords.replace(',', ';')
                keywords = keywords.replace(' |', ';')
                keywords = keywords.replace('| ', ';')
                keywords = keywords.replace(' | ', ';')
                keywords = keywords.replace('|', ';')
                enrichment.loc[i, 'Search Keywords'] = keywords
                
    new_enrichment = opxl.load_workbook('forms/vwr_enrichment_sera.xlsx')
    sheet = new_enrichment.active
    if sheet.max_row < len(enrichment):
        sheet.append([''])
        for j in range(len(enrichment) - sheet.max_row + 12):
            sheet.insert_rows(sheet.max_row)
    i = 11
    for row in sheet.iter_rows(min_row=13):
        if i < (len(enrichment)):
            for j in range(len(enrichment.columns)):
                row[j].value = enrichment.iloc[i, j]
        else:
            break
        i += 1

    new_enrichment.save('outputs/enrichment_outputs/vwr_enrichment_sera_output.xlsx')

    return enrichment

def fillVWR_Enrichmnent_CCM(enrichment, magento, prms):
    for i in range(11, len(enrichment)):
        sku = enrichment.loc[i, 'Supplier Part No.']
        product_info = magento.loc[magento['sku'] == sku]
        prms_info = prms.loc[prms['SKU'] == sku]
        
        if not prms_info.empty:
            # ship_temp = prms_info['Ship Temp'].values[0]
            storage_temp = prms_info['Storage Temp'].values[0]
                
            if storage_temp == 'AM':
                enrichment.loc[i, 'Storage and stability'] = '15C to 30C'
            elif storage_temp == 'FR':
                enrichment.loc[i, 'Storage and stability'] = '-30C to -2C'
            elif storage_temp == 'RF':
                enrichment.loc[i, 'Storage and stability'] = '2C to 8C'
            elif storage_temp == '70' or storage_temp == '80':
                enrichment.loc[i, 'Storage and stability'] = '-70C'
        
        if not product_info.empty:
            name = product_info['name'].values[0]
            description = product_info['description'].values[0]
            short_desc = product_info['short_description'].values[0]
            pack_size_joined = product_info['pack_size_joined'].values[0]
            # host = product_info['host'].values[0]
            # concentration = product_info['concentration'].values[0]
            # purity = product_info['purity'].values[0]
            application_notes = product_info['application_notes'].values[0]
            # sterility = product_info['sterility'].values[0]
            # keywords = product_info['keywords'].values[0]
            culture_media_type = product_info['culture_media_type'].values[0]
            formulation = product_info['formulation'].values[0]

            
            enrichment.loc[i, 'Cell Culture Media Type'] = culture_media_type
            enrichment.loc[i, 'Media Formulation'] = formulation
            enrichment.loc[i, 'Media Format'] = formulation
            enrichment.loc[i, 'Size'] = pack_size_joined
            enrichment.loc[i, 'Application'] = application_notes
            
            if type(short_desc) == str:
                short_desc = tidyDescription(short_desc)
                enrichment.loc[i, 'Description'] = short_desc
                enrichment.loc[i, 'Quick Summary Text\n(max. 100 words)'] = short_desc
            
            enrichment.loc[i, 'Brand Name'] = 'MP Bio'
            enrichment.loc[i, 'Product Title\n(max. 100 characters)'] = name
            
            if type(description) == str:
                description = tidyDescription(description)
                enrichment.loc[i, 'Extended Exposition Text'] = description
    new_enrichment = opxl.load_workbook('forms/vwr_enrichment_ccm.xlsx')
    sheet = new_enrichment.active
    if sheet.max_row < len(enrichment):
        sheet.append([''])
        for j in range(len(enrichment) - sheet.max_row + 12):
            sheet.insert_rows(sheet.max_row)
    i = 11
    for row in sheet.iter_rows(min_row=13):
        if i < (len(enrichment)):
            for j in range(len(enrichment.columns)):
                row[j].value = enrichment.iloc[i, j]
        else:
            break
        i += 1

    new_enrichment.save('outputs/enrichment_outputs/vwr_enrichment_ccm_output.xlsx')

    return enrichment

def fillVWR_Enrichmnent_Chemicals(enrichment, magento, prms):

    for i in range(22, len(enrichment)):
        sku = enrichment['Supplier Part No.'][i]
        product_info = magento.loc[magento['sku'] == sku]
        prms_info = prms.loc[prms['SKU'] == sku]
        try:
            pubchem_data = getPubchemData(sku, magento)
        except:
            pubchem_data = None
        # pubchem_data = None #change this

        if not pubchem_data == None:
            density = pubchem_data['Density']
            boiling_point = pubchem_data['Boiling Point']

            if (not density == None) and len(density) > 0 and (not 'None' in density):
                enrichment.loc[i, 'Density'] = density + " | pulled from Pubchem"
            if (not boiling_point == None) and len(boiling_point) > 0 and (not 'None' in boiling_point):
                boiling_point = tidyDescription(boiling_point)
                enrichment.loc[i, 'Boiling Point'] = boiling_point + " | pulled from Pubchem"
        
        if not prms_info.empty:
            pack_size = prms_info['Pack Size'].values[0]
            storage_temp = prms_info['Storage Temp'].values[0]
    
            enrichment.loc[i, 'Size'] = pack_size
        
            if storage_temp == 'AM':
                enrichment.loc[i, 'Storage Temperature'] = 'Ambient'
            elif storage_temp == 'FR':
                enrichment.loc[i, 'Storage Temperature'] = 'Dry Ice'
            elif storage_temp == 'RF':
                enrichment.loc[i, 'Storage Temperature'] = 'Cold Pack'
            elif storage_temp == '70' or storage_temp == '80':
                enrichment.loc[i, 'Storage Temperature'] = 'Dry Ice'
        
        if not product_info.empty:
            name = product_info['name'].values[0]
            description = product_info['description'].values[0]
            keywords = product_info['keywords'].values[0]
            cas_number = product_info['cas_number'].values[0]
            melting_point = product_info['melting_point'].values[0]
            if type(melting_point) == str:
                melting_point = tidyDescription(melting_point)

            if type(keywords) == str:
                keywords = keywords.replace(', ', ';')
                keywords = keywords.replace(',', ';')
                keywords = keywords.replace(' |', ';')
                keywords = keywords.replace('| ', ';')
                keywords = keywords.replace(' | ', ';')
                keywords = keywords.replace('|', ';')
                enrichment.loc[i, 'Search Keywords'] = keywords
            
            if type(name) == str:
                name = tidyDescription(name)
                name = ''.join([i for i in name if i.isalnum() or i == ' '])
                enrichment.loc[i, 'Product Name (full)'] = name
                
            if type(description) == str:
                description = tidyDescription(description)
                description = ''.join([i for i in description if i.isalnum() or i == ' '])
                enrichment.loc[i, 'Product Text'] = description
                
            enrichment.loc[i, 'CAS'] = cas_number
            enrichment.loc[i, 'Melting Point'] = melting_point
    new_enrichment = opxl.load_workbook('forms/GlobalProductEnrichmentFile_Chemicals (New).xlsx')
    sheet = new_enrichment.active
    if sheet.max_row < len(enrichment):
        sheet.append([''])
        for j in range(len(enrichment) - sheet.max_row + 18):
            sheet.insert_rows(sheet.max_row)
    i = 22
    for row in sheet.iter_rows(min_row=24):
        if i < (len(enrichment)):
            for j in range(len(enrichment.columns)):
                row[j].value = enrichment.iloc[i, j]
        else:
            break
        i += 1

    new_enrichment.save('outputs/enrichment_outputs/vwr_enrichment_chemicals_output.xlsx')

    return enrichment

def fillVWR_Enrichment(filename, magento):
    # enrichment = pd.read_excel('forms/vwr_enrichment_form.xlsx', dtype = object)
    # enrichment.columns = np.arange(len(enrichment.columns))
    # enrichment.columns = enrichment.iloc[6]
    enrichment = pd.read_excel('forms/GlobalProductEnrichmentFile (New).xlsx')
    enrichment.columns = enrichment.iloc[6]
    new_columns = [i.strip() if type(i) == str else i for i in enrichment.columns]
    enrichment.columns = new_columns
    wb = opxl.load_workbook(filename)
    skus = wb.active
    for i in range(2, skus.max_row+1):
        enrichment.loc[i+9, 'Supplier Part No.'] = str(skus['A'+str(i)].value)
        
    for i in range(11, skus.max_row+10):
        sku = enrichment['Supplier Part No.'][i]
        product_info = magento.loc[magento['sku'] == sku]
        
        if not product_info.empty:
            short_desc = product_info['short_description'].values[0]
            keywords = product_info['meta_keywords'].values[0]
            name = product_info['name'].values[0].upper()
            description = product_info['description'].values[0]
            
            enrichment.loc[i, 'Brand Name'] = 'MP Biomedical'
            enrichment.loc[i, 'Supplier Name'] = 'MP Biomedical'
            enrichment.loc[i, 'Product Title\n(max. 100 characters)'] = tidyDescription(str(name))
            enrichment.loc[i, 'Quick Summary Text\n(max. 100 words)'] = tidyDescription(str(short_desc))
            enrichment.loc[i, 'Key Features/Benefits'] = tidyDescription(str(description))
            enrichment.loc[i, 'Extended Exposition Text'] = keywords
            
    new_enrichment = opxl.load_workbook('forms/vwr_enrichment_form.xlsx')
    regulatory_sheet = new_enrichment.active
    if regulatory_sheet.max_row < skus.max_row:
        regulatory_sheet.append([''])
        for j in range(skus.max_row - regulatory_sheet.max_row + 11):
            regulatory_sheet.insert_rows(regulatory_sheet.max_row)
    i = 11
    for row in regulatory_sheet.iter_rows(min_row=13):
        if i < (skus.max_row+11):
            for j in range(len(enrichment.columns)):
                row[j].value = enrichment.iloc[i, j]
        else:
            break
        i = i+1

    new_enrichment.save('outputs/enrichment_outputs/vwr_enrichment_output.xlsx')
    # new_enrichment.save('../../outputs/enrichment_outputs/vwr_enrichment_output.xlsx')


def fillVWR_New(product_manager, prms2, e_marketing):
    vwr = pd.read_excel('forms/vwr_form.xlsx', dtype=object)
    vwr.columns = np.arange(len(vwr.columns))
    for i in range(1, len(product_manager)):
        vwr.loc[i+3, 4] = str(product_manager['sku'][i])

        
    for i in range(4, len(vwr)):
        sku = str(vwr[4][i])
        product_info = product_manager.loc[product_manager['sku'] == sku]
        marketing_info = e_marketing.loc[e_marketing['sku'] == sku]
        prms_info = prms2.loc[prms2['sku'] == sku]
        
        if not product_info.empty:
            name = product_info['Name'].values[0].upper().replace(',', '')
            full_desc = product_info['Description'].values[0]
            short_desc = product_info['short_description'].values[0]
            price = product_info['Price information '].values[0]
            shipping_condition = product_info['shipping condition '].values[0]
            storage_condition = product_info['storage condition '].values[0]
            country_of_origin = product_info['country_of_manufacture'].values[0]
            shelf_life = product_info['shelf life '].values[0]
            pkg_size = product_info['Package size '].values[0]
            pkg_unit = product_info['Package size unit'].values[0]
            tariff_code = product_info['tariff_code'].values[0]
            host = product_info['host'].values[0]
            keywords = product_info['keywords'].values[0]
            unspsc = product_info['unspsc'].values[0]
            hazard_statements = product_info['hazard_statements'].values[0]
            hazard_class = product_info['Hazard class '].values[0]
            psn = product_info['proper shipping name '].values[0]
            un_num = product_info['UN/NA #'].values[0]
            packing_group = product_info['Package group'].values[0]
            safety_symbol = product_info['safety_symbol'].values[0]
            cas_number = product_info['cas_number'].values[0]
            tariff_code = product_info['tariff_code'].values[0]
            ph = product_info['ph'].values[0]
            protein_or_enzyme_type = product_info['protein_or_enzyme_type'].values[0]
            #msds_avail = skus['B'+str(i+1)].value
            biochem_physiol_actions = product_info['biochem_physiol_actions'].values[0]
            
            if type(name) == str:
                name = tidyDescription(name)
                if len(name) <= 40:
                    vwr[5][i] = name
                else:
                    vwr[5][i] = name[:40]
            if type(short_desc) == str:
                short_desc = tidyDescription(short_desc)
                if len(short_desc) <= 300:
                    vwr[6][i] = short_desc
                else:
                    vwr[6][i] = short_desc[:300]
                    
            if shipping_condition == 'Cold Pack' or shipping_condition == 'Cold pack':
                vwr[107][i] = 'B'
                vwr[108][i] = 'Blue'
            elif shipping_condition == 'Dry Ice':
                vwr[107][i] = 'A'
                vwr[108][i] = 'Dry'
            else:
                vwr[107][i] = 'C'
                
            if type(storage_condition) == str:
                if 'Room Temperature' in storage_condition or '15-30' in storage_condition or 'ROOM TEMPERATURE' in storage_condition:
                    vwr[111][i] = 'C'
                elif '-20' in storage_condition:
                    vwr[111][i] = 'A'
                elif '-70' in storage_condition or '-80' in storage_condition:
                    vwr[111][i] = 'D'
                elif '2-8' in storage_condition or '0-5' in storage_condition or '2 - 8' in storage_condition or '0' in storage_condition or '4' in storage_condition:
                    vwr[111][i] = 'B'
                else:
                    vwr[111][i] = storage_condition
                
            if type(host) == str and host == 'N/A':
                vwr[61][i] = 'Synthetic'
            else:
                vwr[61][i] = 'Animal: ' + str(host)
            
            #if DOT_PSN != 'N/A': 
            vwr[68][i] = name
            vwr[69][i] = biochem_physiol_actions
            vwr[70][i] = packing_group
            vwr[71][i] = 'No'
            vwr[72][i] = 'No'
            vwr[73][i] = 'No'
            vwr[77][i] = packing_group
            
            tariff_code = str(tariff_code)
            if len(tariff_code) >= 4:
                vwr[54][i] = tariff_code[:4] + '.99.9999'
            else:
                vwr[54][i] = tariff_code
                
            vwr[7][i] = 'EA'
            vwr[8][i] = 1
            vwr[10][i] = 'USD'
            vwr[12][i] = '1/1/2021'
            vwr[15][i] = 'No'
            vwr[18][i] = pkg_unit
            vwr[19][i] = pkg_size
            vwr[20][i] = 'EA'
            vwr[21][i] = price
            vwr[22][i] = price
            vwr[43][i] = 'Y'
            vwr[58][i] = country_of_origin
            vwr[59][i] = 'No'
            vwr[60][i] = 'No'
            vwr[62][i] = host
            #vwr[65][i] = msds_avail
            #vwr[66][i] = msds_avail
            vwr[67][i] = psn
            vwr[82][i] = 'Box'
            vwr[83][i] = 'No'
            vwr[84][i] = 'Plastic'
            vwr[85][i] = cas_number
            vwr[86][i] = 'No'
            vwr[87][i] = 'No'
            vwr[88][i] = 'No'
            vwr[89][i] = 'Y'
            vwr[90][i] = 'No'
            vwr[91][i] = 'No'
            vwr[92][i] = 'N/A'
            vwr[93][i] = 'N/A'
            vwr[105][i] = 'No'
            vwr[106][i] = 'Y'
            vwr[110][i] = 'No'
            vwr[164][i] = 'N'
            vwr[109][i] = shelf_life
        
        if not marketing_info.empty:
            
            weight = marketing_info['weight'].values[0]
            meta_keywords = marketing_info['meta_keywords'].values[0]
            meta_description = marketing_info['meta_description'].values[0]
            
            vwr[39][i] = weight
            
            if weight < 0.5:
                vwr[40][i] = 7
                vwr[41][i] = 4
                vwr[42][i] = 5
            elif weight <= 1:
                vwr[40][i] = 12
                vwr[41][i] = 7
                vwr[42][i] = 5
            else:
                vwr[40][i] = 12
                vwr[41][i] = 12
                vwr[42][i] = 12
            
        if not prms_info.empty:
            pkg_size_joined = prms_info['pack_size_joined'].values[0]
            ship_conditions = prms_info['ship_conditions'].values[0]
            un_number = prms_info['un_number'].values[0]
            pack_group = prms_info['packing_group'].values[0]
            ship_hazard_code = prms_info['ship_hazard_code'].values[0]

    new_vwr = opxl.load_workbook('forms/vwr_form.xlsx')
    vwr_sheet = new_vwr.active
    if vwr_sheet.max_row < len(product_manager)+4:
        for j in range(len(product_manager) - vwr_sheet.max_row + 4):
            vwr_sheet.insert_rows(vwr_sheet.max_row-1)
    # for j in range(product_manager.max_row):
    #     vwr_sheet.insert_rows(6+j)
    i = 1
    for row in vwr_sheet.iter_rows(min_row=6):
        if i < len(product_manager):
            for j in range(2, len(vwr.columns)):
                row[j].value = vwr[j][i+3]
        else:
            break
        i = i+1

    new_vwr.save('../../outputs/new_product_outputs/new_vwr_output.xlsx')

def fillThomas_New(product_manager, prms2, e_marketing):
    thomas = pd.read_excel('forms/thomas_form.xlsx', dtype = object)
    thomas.columns = np.arange(len(thomas.columns))
    for i in range(21, len(product_manager)+20):
        thomas[1][i] = str(product_manager['sku'][i-20])
        
    for i in range(21, len(product_manager)+20):
        sku = thomas[1][i]
        product_info = product_manager.loc[product_manager['sku'] == sku]
        marketing_info = e_marketing.loc[e_marketing['sku'] == sku]
        prms_info = prms2.loc[prms2['sku'] == sku]
        if not product_info.empty:
            name = product_info['Name'].values[0].upper().replace(',', '')
            full_desc = product_info['Description'].values[0]
            short_desc = product_info['short_description'].values[0]
            price = product_info['Price information '].values[0]
            shipping_condition = product_info['shipping condition '].values[0]
            storage_condition = product_info['storage condition '].values[0]
            country_of_origin = product_info['country_of_manufacture'].values[0]
            shelf_life = product_info['shelf life '].values[0]
            pkg_size = product_info['Package size '].values[0]
            pkg_unit = product_info['Package size unit'].values[0]
            tariff_code = product_info['tariff_code'].values[0]
            host = product_info['host'].values[0]
            keywords = product_info['keywords'].values[0]
            unspsc = product_info['unspsc'].values[0]
            hazard_statements = product_info['hazard_statements'].values[0]
            hazard_class = product_info['Hazard class '].values[0]
            psn = product_info['proper shipping name '].values[0]
            un_num = product_info['UN/NA #'].values[0]
            packing_group = product_info['Package group'].values[0]
            safety_symbol = product_info['safety_symbol'].values[0]
            cas_number = product_info['cas_number'].values[0]
            tariff_code = product_info['tariff_code'].values[0]
            ph = product_info['ph'].values[0]
            protein_or_enzyme_type = product_info['protein_or_enzyme_type'].values[0]
            # msds_avail = skus['B'+str(i-1)].value
            purity = product_info['purity'].values[0]
            grade = product_info['grade'].values[0]
            molecular_weight = product_info['molecular_weight'].values[0]
            key_applications = product_info['key_applications'].values[0]
            concentration = product_info['concentration'].values[0]
            
            thomas[3][i] = 'EA'
            thomas[4][i] = 1
            
            if type(price) != str:
                if str(sku).startswith('11'):
                    thomas[5][i] = 0.89*price
                else:
                    thomas[5][i] = 0.8*price
                thomas[6][i] = price
                
            if shipping_condition == 'Dry Ice':
                thomas[13][i] = 'Ice/Dry Ice'
                thomas[14][i] = '$10.00'
                thomas[15][i] = 'ITEM'
                thomas[31][i] = 'DRY ICE'
            elif shipping_condition == 'Cold Pack' or shipping_condition == 'Cold pack':
                thomas[31][i] = 'ICE'
            elif shipping_condition == 'Ambient':
                thomas[31][i] = 'RT'
                
            if type(name) == str:
                name = tidyDescription(name)
                if (not (str(pkg_size) + str(pkg_unit)) in name) and not ((str(pkg_size) + ' ' + str(pkg_unit)) in name):
                    if len(name) > 40:
                        name = name[:39-len(str(pkg_size) + str(pkg_unit))] + ' ' + str(pkg_size) + str(pkg_unit)
                    else:
                        name = name + ' ' + str(pkg_size) + str(pkg_unit)
                if len(name) > 40:
                    name = name[:40]
                thomas[2][i] = name
                
            if type(cas_number) == str and len(cas_number) > 0:
                thomas[17][i] = 'Chemicals'
                thomas[52][i] = grade
                thomas[53][i] = str(pkg_size) + str(pkg_unit)
                thomas[54][i] = 'Bottle'
                thomas[55][i] = cas_number
                thomas[59][i] = ph
                
            thomas[21][i] = 'Y'
            thomas[22][i] = 'NONE'
            thomas[24][i] = country_of_origin
            
            tariff_code = str(tariff_code)
            if len(tariff_code) >= 4:
                thomas[23][i] = tariff_code[:4] + '.99.9999'
                
            if shelf_life > 0:
                thomas[34][i] = str(shelf_life) + ' days'
                
            if type(storage_condition) == str:
                if 'Room Temperature' in storage_condition or '15-30' in storage_condition or 'ROOM TEMPERATURE' in storage_condition:
                    thomas[32][i] = 'RT'
                    thomas[33][i] = 'N/A'
                elif '-20' in storage_condition:
                    thomas[32][i] = '-20°C'
                    thomas[33][i] = 'Freezer'
                elif '-70' in storage_condition or '-80' in storage_condition:
                    thomas[32][i] = '-80°C'
                    thomas[33][i] = 'Freezer'
                elif '2-8' in storage_condition or '0-5' in storage_condition or '2 - 8' in storage_condition or '0' in storage_condition or '4' in storage_condition:
                    thomas[32][i] = '4°C'
                    thomas[33][i] = 'Refrigerator'
                else:
                    thomas[32][i] = storage_condition
                    
            if type(name) == str:
                name = tidyDescription(name)
                thomas[64][i] = name + ' ' + str(pkg_size) + str(pkg_unit)
                thomas[66][i] = name + ' ' + str(pkg_size) + str(pkg_unit)
                
            if type(keywords) == str:
                thomas[68][i] = keywords.replace(',', ' ')
                
            specifications = ''
            if type(purity) == str and len(purity) > 0:
                specifications = specifications + 'Purity: ' + purity + '\n'
            if type(molecular_weight) == str and len(molecular_weight) > 0:
                specifications = specifications + 'Molecular Weight: ' + molecular_weight + '\n'
            if type(key_applications) == str and len(key_applications) > 0:
                specifications = specifications + 'Key Applications: ' + key_applications + '\n'
            if type(concentration) == str and len(concentration) > 0:
                specifications = specifications + 'Concentration: ' + concentration
            thomas[67][i] = specifications
            
            thomas[44][i] = 'No'
            thomas[65][i] = 'N/A'
                
    #         thomas[69][i] = img_link
            
        if not marketing_info.empty:
            weight = marketing_info['weight'].values[0]
            meta_keywords = marketing_info['meta_keywords'].values[0]
            meta_description = marketing_info['meta_description'].values[0]
            
            if weight < 0.5:
                thomas[25][i] = 7
                thomas[26][i] = 4
                thomas[27][i] = 5
                thomas[28][i] = 0.081
            elif weight <= 1:
                thomas[25][i] = 12
                thomas[26][i] = 7
                thomas[27][i] = 5
                thomas[28][i] = 0.243
            else:
                thomas[25][i] = 12
                thomas[26][i] = 12
                thomas[27][i] = 12
                thomas[28][i] = 1
                
            thomas[29][i] = weight
        
        if not prms_info.empty:
            pkg_size_joined = prms_info['pack_size_joined'].values[0]
            ship_conditions = prms_info['ship_conditions'].values[0]
            un_number = prms_info['un_number'].values[0]
            pack_group = prms_info['packing_group'].values[0]
            ship_hazard_code = prms_info['ship_hazard_code'].values[0]
            
    new_thomas = opxl.load_workbook('forms/thomas_form.xlsx')
    thomas_sheet = new_thomas.active
    if thomas_sheet.max_row < len(product_info):
        for j in range(len(product_manager) - thomas_sheet.max_row+22):
            thomas_sheet.insert_rows(thomas_sheet.max_row-1)
    i = 1
    for row in thomas_sheet.iter_rows(min_row=23):
        if i < len(product_manager)+22:
            for j in range(len(thomas.columns)):
                row[j].value = thomas[j][i+2]
        else:
            break
        i = i+1

    new_thomas.save('../../outputs/new_product_outputs/new_thomas_output.xlsx')

def fillFisher_New(product_manager, prms2, e_marketing):
    fisher_file = pd.ExcelFile('forms/fisher_form.xlsx')
    regulatory = pd.read_excel(fisher_file, 'Regulatory', dtype=object)
    regulatory.columns = np.arange(len(regulatory.columns))
    fisher = pd.read_excel(fisher_file, 'General Info', dtype=object)
    fisher.columns = np.arange(len(fisher.columns))

    for i in range(1, len(product_manager)):
        regulatory.loc[i, 1] = product_manager['sku'][i]
        fisher.loc[i, 1] = product_manager['sku'][i]

    for i in range(1, len(regulatory[1])):
        sku = regulatory[1][i]
        product_info = product_manager.loc[product_manager['sku'] == sku]
        marketing_info = e_marketing.loc[e_marketing['sku'] == sku]
        prms_info = prms2.loc[prms2['sku'] == sku]

        if not product_info.empty:
            name = product_info['Name'].values[0].upper().replace(',', '')
            full_desc = product_info['Description'].values[0]
            short_desc = product_info['short_description'].values[0]
            price = product_info['Price information '].values[0]
            shipping_condition = product_info['shipping condition '].values[0]
            storage_condition = product_info['storage condition '].values[0]
            country_of_origin = product_info['country_of_manufacture'].values[0]
            shelf_life = product_info['shelf life '].values[0]
            pkg_size = product_info['Package size '].values[0]
            pkg_unit = product_info['Package size unit'].values[0]
            tariff_code = product_info['tariff_code'].values[0]
            host = product_info['host'].values[0]
            keywords = product_info['keywords'].values[0]
            unspsc = product_info['unspsc'].values[0]
            hazard_statements = product_info['hazard_statements'].values[0]
            hazard_class = product_info['Hazard class '].values[0]
            psn = product_info['proper shipping name '].values[0]
            un_num = product_info['UN/NA #'].values[0]
            packing_group = product_info['Package group'].values[0]
            safety_symbol = product_info['safety_symbol'].values[0]
            cas_number = product_info['cas_number'].values[0]
            ph = product_info['ph'].values[0]
            protein_or_enzyme_type = product_info['protein_or_enzyme_type'].values[0]
            biochem_physiol_actions = product_info['biochem_physiol_actions'].values[0]
            
            if sku.startswith('07') or sku.startswith('08') or sku.startswith('09') or sku.startswith('11'):
                regulatory[3][i] = 'Diagnostics'
            elif len(str(cas_number)) > 0:
                regulatory[3][i] = 'Chemicals'
            elif pkg_unit == 'PP' or pkg_unit == 'pp':
                regulatory[3][i] = 'Consumables'
            elif sku.startswith('02') or sku.startswith('03') or sku.startswith('04') or sku.startswith('05'):
                regulatory[3][i] = 'Diagnostics'
            

            if type(storage_condition) == str:
                if 'Room Temperature' in storage_condition or '15-30' in storage_condition or 'ROOM TEMPERATURE' in storage_condition or 'room Temperature' in storage_condition:
                    regulatory[6][i] = 'GWN4'
                elif '-20' in storage_condition:
                    regulatory[6][i] = 'DFD1'
                elif '-70' in storage_condition or '-80' in storage_condition:
                    regulatory[6][i] = 'DFD1'
                elif '2-8' in storage_condition or '0-5' in storage_condition or '2 - 8' in storage_condition or '0' in storage_condition or '4' in storage_condition or 'freezer' in storage_condition:
                    regulatory[6][i] = 'RFN1'
                else:
                    regulatory[6][i] = storage_condition
            else:
                regulatory[6][i] = 'GWN4'
            
            regulatory[8][i] = 'N'
            regulatory[9][i] = 'N'
            regulatory[11][i] = 'N'
            regulatory[12][i] = 'N'
            
            tariff_code = str(tariff_code)
            hts = ''.join([i for i in tariff_code if (i.isdigit() or i == '.' or i == ' ') ])
            if len(hts) >= 4:
                regulatory[29][i] = hts[:4] + '999999'
            elif len(hts) > 0:
                regulatory[29][i] = hts + '999999'
                
            regulatory[30][i] = 'N'
            regulatory[31][i] = 'N'
            regulatory[32][i] = 'N'
            regulatory[36][i] = 'N'
            regulatory[37][i] = 'N'
            regulatory[39][i] = 'N'
            regulatory[40][i] = 'N'
            regulatory[43][i] = 'N'
                
            if type(host) == str and len(host) > 0:
                regulatory[44][i] = 'Y'
            else:
                regulatory[44][i] = 'N'

            if len(str(cas_number)) > 0:
                regulatory[48][i] = cas_number
                regulatory[49][i] = 100
            
            regulatory[58][i] = 'N'
            regulatory[61][i] = 'N'
            regulatory[62][i] = 'N'
            regulatory[63][i] = 'N'
            regulatory[64][i] = 'NA'

    for i in range(1, len(fisher[1])):
        sku = fisher[1][i]
        product_info = product_manager.loc[product_manager['sku'] == sku]
        marketing_info = e_marketing.loc[e_marketing['sku'] == sku]
        prms_info = prms2.loc[prms2['sku'] == sku]
            
        if not product_info.empty:
            name = product_info['Name'].values[0].upper().replace(',', '')
            short_name = product_info['30 character name '].values[0]
            full_desc = product_info['Description'].values[0]
            short_desc = product_info['short_description'].values[0]
            price = product_info['Price information '].values[0]
            shipping_condition = product_info['shipping condition '].values[0]
            storage_condition = product_info['storage condition '].values[0]
            country_of_origin = product_info['country_of_manufacture'].values[0]
            shelf_life = product_info['shelf life '].values[0]
            pkg_size = product_info['Package size '].values[0]
            pkg_unit = product_info['Package size unit'].values[0]
            tariff_code = product_info['tariff_code'].values[0]
            host = product_info['host'].values[0]
            keywords = product_info['keywords'].values[0]
            unspsc = product_info['unspsc'].values[0]
            hazard_statements = product_info['hazard_statements'].values[0]
            hazard_class = product_info['Hazard class '].values[0]
            psn = product_info['proper shipping name '].values[0]
            un_num = product_info['UN/NA #'].values[0]
            packing_group = product_info['Package group'].values[0]
            safety_symbol = product_info['safety_symbol'].values[0]
            cas_number = product_info['cas_number'].values[0]
            tariff_code = product_info['tariff_code'].values[0]
            ph = product_info['ph'].values[0]
            protein_or_enzyme_type = product_info['protein_or_enzyme_type'].values[0]
            biochem_physiol_actions = product_info['biochem_physiol_actions'].values[0]

            # if type(name) == str:
            #     name = tidyDescription(name)
            #     temp_name = ''
            #     for c in name:
            #         if c.isalpha() or c.isnumeric() or c == ' ':
            #             temp_name = temp_name + c
            #     name = temp_name
            #     if len(name) <= 240:
            #         fisher[3][i] = name
            #     else:
            #         fisher[3][i] = name[:240]

            #     pack_size_joined = str(pkg_size) + ' ' + str(pkg_unit)
            #     if len(name) <= 30-len(pack_size_joined)-1:
            #         fisher[2][i] = name + ' ' + pack_size_joined
            #     else:
            #         fisher[2][i] = name[:(30-len(pack_size_joined)-1)] + ' ' + pack_size_joined

            if type(short_name) == str:
                short_name = tidyDescription(short_name)
                if len(short_name) <= 30:
                    fisher[2][i] = short_name
                else:
                    if str(pkg_size) not in short_name[:30]:
                        pack_size_joined = str(pkg_size) + ' ' + str(pkg_unit)
                        fisher[2][i] = short_name[:(30-len(pack_size_joined)-1)] + ' ' + pack_size_joined
                    else:
                        fisher[2][i] = short_name[:30]
            if type(short_desc) == str:
                short_desc = tidyDescription(short_desc)
                if len(short_desc) < 240:
                    fisher[3][i] = short_desc
                else:
                    fisher[3][i] = short_desc[:240]
                    
            fisher[15][i] = 'EA'
            fisher[16][i] = 1
            fisher[17][i] = 'EA'
            fisher[22][i] = pkg_size

            if type(pkg_unit) == str:
                unit = pkg_unit
                if 'preps' in unit or 'Preps' in unit:
                    unit = 'PP'
                elif 'mL' in unit or 'ml' in unit:
                    unit = 'ML'
                elif 'lb' in unit:
                    unit = 'LB'
                elif 'tests' in unit or 'Tests' in unit:
                    unit = 'TS'
                elif 'KU' in unit or 'ku' in unit:
                    unit = 'KU'
                elif 'mg' in unit:
                    unit = 'MG'
                elif 'U' in unit or 'wells' in unit or 'Bags' in unit or 'Bottle' in unit or 'Each' in unit or 'Kit' in unit or 'mCI' in unit:
                    unit = 'UN'
                elif unit == 'g':
                    unit = 'GR'
                elif unit == 'l' or unit == 'L' or 'Liter' in unit:
                    unit = 'LT'
                else:
                    unit = 'UN'
                fisher[23][i] = unit
            
            if type(hazard_statements) == str and len(hazard_statements) > 0:
                fisher[43][i] = 'Y'
            else:
                fisher[43][i] = 'N'
            
            if type(price) != str:
                fisher[32][i] = price*0.7
                fisher[33][i] = price
                
            fisher[35][i] = '30'
            # fisher[36][i] = '31/DEC/2020'
            fisher[41][i] = unspsc
            
            # if msds_avail == 'Y':
            #     fisher[27][i] = '99998'
            # else:
            #     fisher[27][i] = '00000'
                
            fisher[44][i] = hazard_statements
            fisher[45][i] = hazard_class
            fisher[47][i] = packing_group
            
            # if type(storage_condition) == str:
            #     if 'Room Temperature' in storage_condition or '15-30' in storage_condition or 'ROOM TEMPERATURE' in storage_condition:
            #         fisher[36][i] = 'GWN4'
            #     elif '-20' in storage_condition:
            #         fisher[36][i] = 'DFD1'
            #     elif '-70' in storage_condition or '-80' in storage_condition:
            #         fisher[36][i] = 'DFD1'
            #     elif '2-8' in storage_condition or '0-5' in storage_condition or '2 - 8' in storage_condition or '0' in storage_condition or '4' in storage_condition:
            #         fisher[36][i] = 'RFN1'
            #     else:
            #         fisher[36][i] = storage_condition
                    
            fisher[43][i] = 'N'
            fisher[50][i] = 'No'
            
            # if shelf_life > 0:
            #     fisher[54][i] = 'Y'
            #     fisher[55][i] = shelf_life
            # else:
            #     fisher[54][i] = 'N'
            #     fisher[55][i] = 0
            fisher[54][i] = 'N'
            fisher[55][i] = 0
                
            fisher[14][i] = '25'
            fisher[93][i] = 'N'
            fisher[42][i] = country_of_origin
            fisher[51][i] = 'Build to Order'
            fisher[52][i] = 'NA'
            fisher[58][i] = 'Y'
            fisher[59][i] = 'N'
            fisher[61][i] = 'Y'
            fisher[62][i] = 1
            fisher[63][i] = 'Y'
            fisher[65][i] = 'Y'
            fisher[78][i] = 'N'

        if type(country_of_origin) == str and len(country_of_origin) > 0:
            fisher[108][i] = 'Y'
            fisher[109][i] = 'USD'
        else:
            fisher[108][i] = 'N'
            
        if not marketing_info.empty:
            weight = marketing_info['weight'].values[0]
            meta_keywords = marketing_info['meta_keywords'].values[0]
            meta_description = marketing_info['meta_description'].values[0]
            
            fisher[18][i] = weight
            
            if weight < 0.5:
                fisher[19][i] = 7
                fisher[20][i] = 4
                fisher[21][i] = 5
                fisher[49][i] = 60
                fisher[50][i] = 36
            elif weight <= 1:
                fisher[19][i] = 12
                fisher[20][i] = 7
                fisher[21][i] = 5
                fisher[49][i] = 18
                fisher[50][i] = 36
            else:
                fisher[19][i] = 12
                fisher[20][i] = 12
                fisher[21][i] = 12
                fisher[49][i] = 12
                fisher[50][i] = 15
                
            if type(keywords) == str:
                keywords = keywords.split(',')
                j = 0
                while j<5 and j<len(keywords):
                    if keywords[j][0] == ' ':
                        keywords[j] = keywords[j][1:]
                    if len(keywords[j]) <= 11:
                        fisher[j+6][i] = keywords[j]
                    else:
                        fisher[j+6][i] = keywords[j][:11]
                    j = j+1
            
    new_fisher = opxl.load_workbook('forms/fisher_form.xlsx')
    fisher_sheet = new_fisher['General Info']
    regulatory_sheet = new_fisher['Regulatory']
    if fisher_sheet.max_row < len(regulatory[1]):
        for j in range(len(regulatory[1]) - regulatory_sheet.max_row + 1):
            # regulatory_sheet.insert_rows(regulatory_sheet.max_row)
            regulatory_sheet.append([''])
    i = 1
    for row in regulatory_sheet.iter_rows(min_row=2):
        if i < len(regulatory[1]):
            for j in range(1, len(regulatory.columns)):
                row[j].value = regulatory[j][i]
        else:
            break
        i = i+1
    if fisher_sheet.max_row < len(fisher[1]):
        for j in range(len(fisher[1]) - fisher_sheet.max_row + 1):
            # fisher_sheet.insert_rows(fisher_sheet.max_row)
            fisher_sheet.append([''])
    i = 1
    for row in fisher_sheet.iter_rows(min_row=2):
        if i < len(fisher[1]):
            for j in range(1, len(fisher.columns)):
                row[j].value = fisher[j][i]
        else:
            break
        i = i+1

    # new_fisher.save('../../outputs/new_product_outputs/new_fisher_output.xlsx')
    new_fisher.save('outputs/new_product_outputs/new_fisher_output.xlsx')

def fillGlobalProductRevision(filename, magento, new_magento, lot_master, prms, unspsc_codes, origin, magento_sept):
    revision = pd.read_excel('forms/GlobalProductRevisionFile (New).xlsx')
    revision.columns = revision.iloc[7]
    new_columns = [i.strip() if type(i) == str else i for i in revision.columns]
    revision.columns = new_columns
    wb = opxl.load_workbook(filename)
    skus = wb.active
    for i in range(2, skus.max_row+1):
        revision.loc[i+10, 'Supplier Part No.'] = str(skus['A'+str(i)].value)

    for i in range(12, len(revision)):
        sku = revision.loc[i, 'Supplier Part No.']
        product_info_sept = magento_sept.loc[magento_sept['sku'] == sku]

        if not product_info_sept.empty:
            keywords = product_info_sept['keywords'].values[0]
            name = product_info_sept['name'].values[0]
            description = product_info_sept['description'].values[0]

            if type(keywords) == str:
                keywords = keywords.replace(', ', ';')
                keywords = keywords.replace(',', ';')
                keywords = keywords.replace(' |', ';')
                keywords = keywords.replace('| ', ';')
                keywords = keywords.replace(' | ', ';')
                keywords = keywords.replace('|', ';')
                revision.loc[i, 'Search Keywords'] = keywords
            
            revision.loc[i, 'Supplier Name'] = 'MP Biomedicals'

            if type(name) == str:
                revision.loc[i, 'Product Title\n(max. 100 characters)'] = tidyDescription(name)
            
            if type(description) == str:
                revision.loc[i, 'Text related information'] = tidyDescription(description)
                               
    new_revision = opxl.load_workbook('forms/GlobalProductrevisionFile (New).xlsx')
    sheet = new_revision.active
    if sheet.max_row < len(revision):
        sheet.append([''])
        for j in range(len(revision) - sheet.max_row + 4):
            sheet.insert_rows(sheet.max_row)
    i = 12
    for row in sheet.iter_rows(min_row=14):
        if i < (len(revision)):
            for j in range(len(revision.columns)):
                row[j].value = revision.iloc[i, j]
        else:
            break
        i += 1
            
    new_revision.save('outputs/revision_outputs/GlobalProductRevision_output.xlsx')

def fillGlobalProductRevisionChemicals(filename, magento, new_magento, lot_master, prms, unspsc_codes, origin, magento_sept):
    revision = pd.read_excel('forms/GlobalProductRevisionFile_Chemicals (New).xlsx')
    revision.columns = revision.iloc[8]
    new_columns = [i.strip() if type(i) == str else i for i in revision.columns]
    revision.columns = new_columns
    wb = opxl.load_workbook(filename)
    skus = wb.active
    for i in range(2, skus.max_row+1):
        revision.loc[i+11, 'Supplier Part No.'] = str(skus['A'+str(i)].value)

    for i in range(13, len(revision)):
        sku = revision.loc[i, 'Supplier Part No.']
        product_info_sept = magento_sept.loc[magento_sept['sku'] == sku]
        prms_info = prms.loc[prms['SKU'] == sku]

        if not prms_info.empty:
            storage_temp = prms_info['Storage Temp'].values[0]
                
            if storage_temp == 'AM':
                revision.loc[i, 'Storage Temperature'] = '15C to 30C'
            elif storage_temp == 'FR':
                revision.loc[i, 'Storage Temperature'] = '-30C to -2C'
            elif storage_temp == 'RF':
                revision.loc[i, 'Storage Temperature'] = '2C to 8C'
            elif storage_temp == '70' or storage_temp == '80':
                revision.loc[i, 'Storage Temperature'] = '-70C'

        if not product_info_sept.empty:
            keywords = product_info_sept['keywords'].values[0]
            name = product_info_sept['name'].values[0]
            description = product_info_sept['description'].values[0]
            short_description = product_info_sept['short_description'].values[0]
            pack_size_joined = product_info_sept['pack_size_joined'].values[0]
            cas_number = product_info_sept['cas_number'].values[0]
            mdl_number = product_info_sept['mdl_number'].values[0]
            un_number = product_info_sept['un_number'].values[0]
            packing_group = product_info_sept['packing_group'].values[0]
            density = product_info_sept['density'].values[0]
            boiling_point = product_info_sept['boiling_point'].values[0]
            melting_point = product_info_sept['melting_point'].values[0]

            if type(keywords) == str:
                keywords = keywords.replace(', ', ';')
                keywords = keywords.replace(',', ';')
                keywords = keywords.replace(' |', ';')
                keywords = keywords.replace('| ', ';')
                keywords = keywords.replace(' | ', ';')
                keywords = keywords.replace('|', ';')
                revision.loc[i, 'Search Keywords'] = keywords

            if type(name) == str:
                revision.loc[i, 'Product Name'] = tidyDescription(name)
            
            if type(description) == str:
                revision.loc[i, 'Product Text'] = tidyDescription(description)
            
            if type(short_description) == str:
                revision.loc[i, 'Key Features/Benefits'] = tidyDescription(short_description)

            revision.loc[i, 'Supplier Name'] = 'MP Biomedicals'
            revision.loc[i, 'Size'] = pack_size_joined
            revision.loc[i, 'CAS'] = cas_number
            revision.loc[i, 'MDL'] = mdl_number
            revision.loc[i, 'UN'] = un_number
            revision.loc[i, 'Packing Group'] = packing_group
            revision.loc[i, 'Density'] = density
            revision.loc[i, 'Boiling Point'] = boiling_point
            revision.loc[i, 'Melting Point'] = melting_point
    
    revision.columns = revision.columns.fillna('Blank')
                               
    new_revision = opxl.load_workbook('forms/GlobalProductRevisionFile_Chemicals (New).xlsx')
    sheet = new_revision.active
    if sheet.max_row < len(revision):
        sheet.append([''])
        for j in range(len(revision) - sheet.max_row + 4):
            sheet.insert_rows(sheet.max_row)
    i = 14
    for row in sheet.iter_rows(min_row=16):
        if i < (len(revision)):
            for j in range(len(revision.columns)):
                row[j].value = revision.iloc[i, j]
        else:
            break
        i += 1
            
    new_revision.save('outputs/revision_outputs/GlobalProductRevisionChemicals_output.xlsx')