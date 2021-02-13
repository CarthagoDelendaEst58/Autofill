import pandas as pd
import numpy as np
import openpyxl as opxl
import os.path
import pycountry
import pycountry_convert as pc
import datetime as dt

def tidyDescription(desc):
    desc = desc.replace('&TRADE', '')
    desc = desc.replace('Â®', '')
    desc = desc.replace('Î¼', '')
    desc = desc.replace('& Growâ„¢', '')
    desc = desc.replace('â‰', '')
    desc = desc.replace('&Beta', '')
    desc = desc.replace('Â‰¥', '')
    desc = desc.replace('Â', '')
    desc = desc.replace('<em>', '')
    desc = desc.replace('</em>', '')
    desc = desc.replace('®','')
    if desc.endswith(','):
        desc = desc[:len(desc)-1]
    return desc

def fillVWR_Old(filename, magento, new_magento, lot_master, prms, unspsc_codes, origin):
    wb = opxl.load_workbook(filename)
    skus = wb.active
    vwr = pd.read_excel('forms/vwr_form.xlsx', dtype=object)
    vwr.columns = np.arange(len(vwr.columns))
    for i in range(2, skus.max_row+1):
        vwr.loc[i+2, 4] = str(skus['A'+str(i)].value)

    for i in range(4, len(vwr)):
        sku = str(vwr[4][i])
        product_info = magento.loc[magento['sku'] == sku]
        lot_info = lot_master.loc[lot_master['Product number'] == sku]
        prms_info = prms.loc[prms['SKU'] == sku]
        unspsc_info = unspsc_codes.loc[unspsc_codes['Part Number'] == sku]
        origin_info = origin.loc[origin['Product number'] == sku]
        product_info_july = new_magento.loc[new_magento['sku'] == sku]

        if not product_info_july.empty:
            tariff_code = product_info_july['tariff_code'].values[0]
        else:
            tariff_code = ''
        if not product_info.empty:
            if not prms_info.empty:
                name = prms_info['Product Name'].values[0]
                price = prms_info['USD List Price'].values[0]
                un_num = prms_info['UN#'].values[0]
                packing_group = prms_info['Packing Group'].values[0]
                alt_storage_temp = prms_info['Storage Temp'].values[0]
                ship_temp = prms_info['Ship Temp'].values[0]
            else:
                name = product_info['name'].values[0].upper()
                price = product_info['price'].values[0]
                packing_group = product_info['packing_group'].values[0]
                ship_temp = product_info['ship_conditions'].values[0]
                    
            if not origin_info.empty:
                country_of_origin = origin_info.loc[origin_info['Expiration date -'] == max(origin_info['Expiration date -'].values)]['Country of Origin'].values[0]
            elif not prms_info.empty:
                country_of_origin = prms_info['Country of Origin (most recent lot)'].values[0]
            elif not lot_info.empty:
                country_of_origin = lot_info['Country of Origin'].values[0]
            else:
                country_of_origin = ''
            
            if country_of_origin == 'GER':
                        country_of_origin = 'DEU'
                
            if type(country_of_origin) == str and len(country_of_origin) > 0:
                country_of_origin = pc.country_alpha3_to_country_alpha2(country_of_origin)
                        
            if not unspsc_info.empty:
                unspsc = unspsc_info['UNSPSC'].values[0]
            else:
                unspsc = product_info['unspsc'].values[0]
            
            short_desc = product_info['short_description'].values[0]
            img_link = product_info['base_image'].values[0]
            pack_size = product_info['pack_size_numeric_value'].values[0]
            unit = product_info['pack_size_unit_of_measure'].values[0]
            quantity = product_info['lk_packaging_facet'].values[0]
            host = product_info['host'].values[0]
            # tariff_code = product_info['tariff_code'].values[0]
            msds_avail = skus['B'+str(i-2)].value
            cas_number = product_info['cas_number'].values[0]
            DOT_PSN = 'N/A'
            hazard_class = product_info['hazard_class'].values[0]
            biochem_physiol_actions = product_info['biochem_physiol_actions'].values[0]
            storage_and_handling = product_info['storage_and_handling'].values[0]
            group_name = product_info['prms_group_name'].values[0]
            storage_temp = product_info['storage_and_handling'].values[0]
            categories = product_info['categories'].values[0]
            weight_in_lb = new_magento.loc[new_magento['sku'] == sku]['weight'].values[0]
            
            
            if type(quantity) == str:
                quantity = quantity.replace('1 ÂµCi', '')
                quantity = quantity.split('x')[0]
                
            
            if weight_in_lb < 0.5:
                vwr[40][i] = 7
                vwr[41][i] = 4
                vwr[42][i] = 5
            elif weight_in_lb <= 1:
                vwr[40][i] = 12
                vwr[41][i] = 7
                vwr[42][i] = 5
            else:
                vwr[40][i] = 12
                vwr[41][i] = 12
                vwr[42][i] = 12

            if type(name) == str:
                name = tidyDescription(name)
                if len(name) <= 40:
                    vwr[5][i] = name
                else:
                    vwr[5][i] = name[:40]
            if type(short_desc) == str:
                short_desc = tidyDescription(short_desc)
                if len(short_desc) <= 300:
                    vwr[6][i] = short_desc + "\n pls work"
                else:
                    vwr[6][i] = short_desc[:300]
                    
            if ship_temp == 'CP':
                vwr[107][i] = 'B'
                vwr[108][i] = 'Blue'
            elif ship_temp == 'DI':
                vwr[107][i] = 'A'
                vwr[108][i] = 'Dry'
            else:
                vwr[107][i] = 'C'
                
            if type(host) == str and len(host) > 0:
                vwr[61][i] = 'Animal: ' + host
            else:
                vwr[61][i] = 'Synthetic'
                
            #if DOT_PSN != 'N/A': 
            vwr[68][i] = name
            vwr[69][i] = biochem_physiol_actions
            vwr[70][i] = packing_group
            vwr[71][i] = 'No'
            vwr[72][i] = 'No'
            vwr[73][i] = 'No'
            vwr[77][i] = packing_group
                
                
            if type(price) != str and type(categories) == str:
                if 'Biochemicals' in categories or 'Cell Biology' in categories or 'Immunology' in categories or 'Antibody' in categories or 'Chemicals' in categories:
                    vwr[9][i] = 0.77*price
                    vwr[11][i] = 0.77*price
                elif 'Molecular Biology' in categories or 'SafTest' in categories:
                    vwr[9][i] = 0.85*price
                    vwr[11][i] = 0.85*price
                elif sku.startswith('02') or sku.startswith('07') or sku.startswith('04'):
                    vwr[9][i] = 0.77*price
                    vwr[11][i] = 0.77*price
                elif sku.startswith('09') or sku.startswith('08') or sku.startswith('11'):
                    vwr[9][i] = 0.85*price
                    vwr[11][i] = 0.85*price
                
            tariff_code = str(tariff_code)
            if len(tariff_code) >= 4:
                vwr[54][i] = tariff_code
            
            vwr[7][i] = 'EA'
            vwr[8][i] = 1
            vwr[10][i] = 'USD'
            vwr[12][i] = '1/1/2021'
            vwr[15][i] = 'No'
            vwr[17][i] = quantity
            vwr[18][i] = unit
            vwr[19][i] = pack_size
            vwr[20][i] = 'EA'
            vwr[21][i] = price
            vwr[22][i] = price
            vwr[39][i] = weight_in_lb
            vwr[43][i] = 'Y'
            vwr[58][i] = country_of_origin
            vwr[59][i] = 'No'
            vwr[60][i] = 'No'
            vwr[62][i] = host
            vwr[65][i] = msds_avail
            vwr[66][i] = msds_avail
            vwr[67][i] = DOT_PSN
            vwr[82][i] = 'Box'
            vwr[83][i] = 'No'
            vwr[84][i] = 'Plastic'
            vwr[85][i] = cas_number
            vwr[86][i] = 'No'
            vwr[87][i] = 'No'
            vwr[88][i] = 'No'
            vwr[89][i] = 'Y'
            vwr[90][i] = 'No'
            vwr[91][i] = 'No'
            vwr[92][i] = 'N/A'
            vwr[93][i] = 'N/A'
            vwr[105][i] = 'No'
            vwr[106][i] = 'Y'
            
            if not lot_info.empty:
                creation_date = np.datetime64(lot_info['Creation date -'].values[0])
                expiration_date = np.datetime64(lot_info['Expiration date -'].values[0])
                shelf_life = expiration_date - creation_date
                shelf_life = shelf_life.astype('timedelta64[M]')/np.timedelta64(1, 'M')
                if shelf_life > 0:
                    vwr[109][i] = shelf_life
            
            vwr[110][i] = 'No'
            
            if type(storage_temp) == str:
                if 'Room Temperature' in storage_temp or '15-30' in storage_temp or 'ROOM TEMPERATURE' in storage_temp:
                    vwr[111][i] = 'C'
                elif '-20' in storage_temp:
                    vwr[111][i] = 'A'
                elif '-70' in storage_temp or '-80' in storage_temp:
                    vwr[111][i] = 'D'
                elif '2-8' in storage_temp or '0-5' in storage_temp or '2 - 8' in storage_temp or '0' in storage_temp or '4' in storage_temp:
                    vwr[111][i] = 'B'
                else:
                    vwr[111][i] = storage_temp
            
            vwr[164][i] = 'N'

    new_vwr = opxl.load_workbook('forms/vwr_form.xlsx')
    vwr_sheet = new_vwr.active
    if vwr_sheet.max_row < skus.max_row:
        for j in range(skus.max_row - vwr_sheet.max_row):
            vwr_sheet.insert_rows(vwr_sheet.max_row-1)
    i = 1
    for row in vwr_sheet.iter_rows(min_row=6):
        if i < skus.max_row:
            for j in range(2, len(vwr.columns)):
                row[j].value = vwr[j][i+3]
        else:
            break
        i = i+1

    new_vwr.save('../../outputs/old_product_outputs/old_vwr_output.xlsx')


def fillThomas_Old(filename, magento, new_magento, lot_master, prms, unspsc_codes, origin, magento_sept):
    thomas = pd.read_excel('forms/thomas_form.xlsx', dtype = object)
    thomas.columns = np.arange(len(thomas.columns))
    wb = opxl.load_workbook(filename)
    # wb = opxl.load_workbook('Output for Thomas form.xlsx')
    # skus = wb['Fisher']
    skus = wb.active
    for i in range(2, skus.max_row+1):
        thomas.loc[i+19, 1] = str(skus['A'+str(i)].value)
        
    for i in range(21, skus.max_row+20):
        sku = thomas[1][i]
        product_info = magento.loc[magento['sku'] == sku]
        lot_info = lot_master.loc[lot_master['Product number'] == sku]
        origin_info = origin.loc[origin['Product number'] == sku]
        prms_info = prms.loc[prms['SKU'] == sku]
        unspsc_info = unspsc_codes.loc[unspsc_codes['Part Number'] == sku]
        product_info_july = new_magento.loc[new_magento['sku'] == sku]
        product_info_sept = magento_sept.loc[magento_sept['sku'] == sku]

        if not product_info_july.empty:
            tariff_code = product_info_july['tariff_code'].values[0]
        else:
            tariff_code = ''

        if not product_info.empty:
            if not product_info_sept.empty:
                pack_size_joined = product_info_sept['pack_size_joined'].values[0]
                product_type = product_info_july['product_type'].values[0]
                description = product_info_sept['description'].values[0]
                clone_name = product_info_sept['clone_name'].values[0]
                conjugate = product_info_sept['conjugate'].values[0]
                host = product_info_sept['host'].values[0]
                isotype = product_info_sept['isotype'].values[0]
                solubility = product_info_sept['solubility'].values[0]
                source = product_info_sept['source'].values[0]
                species_reactivity = product_info_sept['species_reactivity'].values[0]
                immunogen = product_info_sept['immunogen'].values[0]
                concentration = product_info_sept['concentration'].values[0]
                specificity = product_info_sept['specificity'].values[0]
                sterility = product_info_sept['sterility'].values[0]
                components = product_info_sept['components'].values[0]
                format_value = product_info_sept['format'].values[0]
                name = product_info_sept['name'].values[0]

            else:
                pack_size_joined = product_info['pack_size_joined'].values[0]
                product_type = product_info['product_type'].values[0]
                description = product_info['description'].values[0]
                clone_name = ''
                conjugate = ''
                host = ''
                isotype = ''
                solubility = ''
                source = ''
                species_reactivity = ''
                immunogen = ''
                concentration = product_info['concentration'].values[0]
                specificity = ''
                sterility = ''
                components = ''
                format_value = ''
                name = ''

            if not prms_info.empty:
                price = prms_info['USD List Price'].values[0]
                un_num = prms_info['UN#'].values[0]
                packing_group = prms_info['Packing Group'].values[0]
                alt_storage_temp = prms_info['Storage Temp'].values[0]
                ship_temp = prms_info['Ship Temp'].values[0]
    #             country_of_origin = prms_info['Country of Origin (most recent lot)'].values[0]
            else:
                price = product_info['price'].values[0]
                packing_group = product_info['packing_group'].values[0]
                ship_temp = product_info['ship_conditions'].values[0]
    #             if not lot_info.empty:
    #                 country_of_origin = lot_info['Country of Origin'].values[0]
    #             else:
    #                 country_of_origin = ''
                    
            if not origin_info.empty:
                country_of_origin = origin_info.loc[origin_info['Expiration date -'] == max(origin_info['Expiration date -'].values)]['Country of Origin'].values[0]
            elif not prms_info.empty:
                country_of_origin = prms_info['Country of Origin (most recent lot)'].values[0]
            elif not lot_info.empty:
                country_of_origin = lot_info['Country of Origin'].values[0]
            else:
                country_of_origin = ''
                        
            if not unspsc_info.empty:
                unspsc = unspsc_info['UNSPSC'].values[0]
            else:
                unspsc = product_info['unspsc'].values[0]
            
            #name = product_info['name'].values[0].upper()
            # description = product_info['description'].values[0]
            short_desc = product_info['short_description'].values[0]
            img_link = product_info['base_image'].values[0]
            #price = product_info['price'].values[0]
            pack_size = product_info['pack_size_numeric_value'].values[0]
            unit = product_info['pack_size_unit_of_measure'].values[0]
            quantity = product_info['lk_packaging_facet'].values[0]
            #packing_group = product_info['packing_group'].values[0]
            #ship_temp = product_info['ship_conditions'].values[0]
            host = product_info['host'].values[0]
            # tariff_code = product_info['tariff_code'].values[0]
            msds_avail = skus['B'+str(i-1)].value
            cas_number = product_info['cas_number'].values[0]
            DOT_PSN = 'N/A'
            hazard_class = product_info['hazard_class'].values[0]
            biochem_physiol_actions = product_info['biochem_physiol_actions'].values[0]
            storage_and_handling = product_info['storage_and_handling'].values[0]
            group_name = product_info['prms_group_name'].values[0]
            #unspsc = product_info['unspsc'].values[0]
            img_link = product_info['base_image'].values[0]
            hazard_statements = product_info['hazard_statements'].values[0]
            keywords = product_info['meta_keywords'].values[0]
            storage_temp = product_info['storage_and_handling'].values[0]
            grade = product_info['grade'].values[0]
            # pack_size_joined = product_info['pack_size_joined'].values[0]
            ph = product_info['ph'].values[0]
            weight_in_lb = new_magento.loc[new_magento['sku'] == sku]['weight'].values[0]
            purity = product_info['purity'].values[0]
            molecular_weight = product_info['molecular_weight'].values[0]
            key_applications = product_info['key_applications'].values[0]
            
    #         if type(quantity) == str:
    #             quantity = int(quantity.split('x')[0])
    #         weight_in_lb = convertWeightToPounds(pack_size, unit, quantity)
            
            if type(name) == str:
                name = tidyDescription(name)
                if (not (str(pack_size) + str(unit)) in name) and not ((str(pack_size_joined) in name)):
                    if len(name) > 40-len(str(pack_size_joined)):
                        name = name[:39-len(pack_size_joined)] + ' ' + pack_size_joined
                    else:
                        name = name + ' ' + str(pack_size_joined)
                if len(name) > 40:
                    name = name[:40]
                thomas[2][i] = name
            
            thomas[3][i] = 'EA'
            thomas[4][i] = 1
            
            if type(price) != str:
                if str(sku).startswith('11'):
                    thomas[5][i] = 0.89*price
                else:
                    thomas[5][i] = 0.8*price
                thomas[6][i] = price
            
            if ship_temp == 'DI':
                thomas[13][i] = 'Ice/Dry Ice'
                thomas[14][i] = '$10.00'
                thomas[15][i] = 'ITEM'
                thomas[31][i] = 'DRY ICE'
            elif ship_temp == 'CP':
                thomas[31][i] = 'ICE'
            elif ship_temp == 'AM':
                thomas[31][i] = 'RT'
            
            if type(cas_number) == str and len(cas_number) > 0:
                thomas[17][i] = 'Chemicals'
                thomas[52][i] = grade
                # thomas[53][i] = pack_size_joined
                thomas[54][i] = 'Bottle'
                thomas[55][i] = cas_number
                thomas[59][i] = ph
                
            thomas[21][i] = 'Y'
            thomas[22][i] = 'NONE'
            thomas[24][i] = country_of_origin
            
            tariff_code = str(tariff_code)
            if len(tariff_code) >= 4:
                thomas[23][i] = tariff_code
            
            if weight_in_lb < 0.5:
                thomas[25][i] = 7
                thomas[26][i] = 4
                thomas[27][i] = 5
                thomas[28][i] = 0.081
            elif weight_in_lb <= 1:
                thomas[25][i] = 12
                thomas[26][i] = 7
                thomas[27][i] = 5
                thomas[28][i] = 0.243
            else:
                thomas[25][i] = 12
                thomas[26][i] = 12
                thomas[27][i] = 12
                thomas[28][i] = 1
                
            thomas[29][i] = weight_in_lb
            
            if type(storage_temp) == str:
                if 'Room Temperature' in storage_temp or '15-30' in storage_temp or 'ROOM TEMPERATURE' in storage_temp:
                    thomas[32][i] = 'RT'
                    thomas[33][i] = 'N/A'
                elif '-20' in storage_temp:
                    thomas[32][i] = '-20°C'
                    thomas[33][i] = 'Freezer'
                elif '-70' in storage_temp or '-80' in storage_temp:
                    thomas[32][i] = '-80°C'
                    thomas[33][i] = 'Freezer'
                elif '2-8' in storage_temp or '0-5' in storage_temp or '2 - 8' in storage_temp or '0' in storage_temp or '4' in storage_temp:
                    thomas[32][i] = '4°C'
                    thomas[33][i] = 'Refrigerator'
                else:
                    thomas[32][i] = storage_temp
            #thomas[32][i] = storage_temp
            
            if not lot_info.empty:
                creation_date = np.datetime64(lot_info['Creation date -'].values[0])
                expiration_date = np.datetime64(lot_info['Expiration date -'].values[0])
                shelf_life = expiration_date - creation_date
                shelf_life = shelf_life.astype('timedelta64[D]')/np.timedelta64(1, 'D')
                if shelf_life > 0:
                    thomas[34][i] = str(shelf_life) + ' days'
            
            thomas[42][i] = 'D'
            thomas[44][i] = 'No'
            thomas[53][i] = pack_size_joined
            
            if type(name) == str:
                name = tidyDescription(name)
                if product_type == 'configurable':
                    thomas[64][i] = name + ' ' + str(pack_size_joined)
                else:
                    thomas[64][i] = name
            
            thomas[65][i] = 'N/A'

            if type(description) == str and len(description) > 0:
                thomas[66][i] = description
            elif not product_info_sept.empty:
                thomas[66][i] = product_info_sept['application_notes'].values[0]
            
            specifications = ''
            if type(purity) == str and len(purity) > 0:
                specifications = specifications + 'Purity: ' + purity.replace('â‰¥', '')
            if type(molecular_weight) == str and len(molecular_weight) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Molecular Weight: ' + molecular_weight
            if type(key_applications) == str and len(key_applications) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Key Applications: ' + key_applications
            if type(concentration) == str and len(concentration) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Concentration: ' + concentration
            if type(clone_name) == str and len(clone_name) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Clone Name: ' + clone_name
            if type(conjugate) == str and len(conjugate) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Conjugate: ' + conjugate
            if type(host) == str and len(host) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Host: ' + host
            if type(isotype) == str and len(isotype) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Isotype: ' + isotype
            if type(solubility) == str and len(solubility) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Solubility: ' + solubility
            if type(source) == str and len(source) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Source: ' + source
            if type(species_reactivity) == str and len(species_reactivity) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Species Reactivity: ' + species_reactivity
            if type(immunogen) == str and len(immunogen) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Immunogen: ' + immunogen
            if type(specificity) == str and len(specificity) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Specificity: ' + specificity
            if type(sterility) == str and len(sterility) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Sterility: ' + sterility
            if type(components) == str and len(components) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Components: ' + components
            if type(format_value) == str and len(format_value) > 0:
                if len(specifications) > 0:
                    specifications = specifications + '\n'
                specifications = specifications + 'Format: ' + format_value
            
            thomas[67][i] = specifications
            
            if type(keywords) == str:
                thomas[68][i] = keywords.replace(',', ' ')
            
            thomas[69][i] = img_link
            
    new_thomas = opxl.load_workbook('forms/thomas_form.xlsx')
    thomas_sheet = new_thomas.active
    if thomas_sheet.max_row < skus.max_row:
        for j in range(skus.max_row - thomas_sheet.max_row):
            thomas_sheet.insert_rows(thomas_sheet.max_row-1)
    i = 1
    for row in thomas_sheet.iter_rows(min_row=23):
        if i < skus.max_row:
            for j in range(len(thomas.columns)):
                row[j].value = thomas[j][i+2]
        else:
            break
        i = i+1

    new_thomas.save('../../outputs/old_product_outputs/old_thomas_output.xlsx')
    # new_thomas.save('outputs/old_product_outputs/old_thomas_output.xlsx')


def fillFisher_Old(filename, magento, new_magento, lot_master, prms, unspsc_codes, origin, magento_sept):
    fisher_file = pd.ExcelFile('forms/fisher_form.xlsx')
    fisher = pd.read_excel(fisher_file, 'General Info', dtype=object)
    regulatory = pd.read_excel(fisher_file, 'Regulatory', dtype=object)
    regulatory.columns = np.arange(len(regulatory.columns))
    fisher.columns = np.arange(len(fisher.columns))
    wb = opxl.load_workbook(filename)
    skus = wb.active
    for i in range(2, skus.max_row+1):
        regulatory.loc[i-2, 1] = str(skus['A'+str(i)].value)
    for i in range(2, skus.max_row+1):
        fisher.loc[i-2, 1] = str(skus['A'+str(i)].value)

    for i in range(len(regulatory)-1):
        sku = regulatory[1][i]
        product_info = magento.loc[magento['sku'] == sku]
        lot_info = lot_master.loc[lot_master['Product number'] == sku]
        prms_info = prms.loc[prms['SKU'] == sku]
        unspsc_info = unspsc_codes.loc[unspsc_codes['Part Number'] == sku]
        product_info_july = new_magento.loc[new_magento['sku'] == sku]
        product_info_sept = magento_sept.loc[magento_sept['sku'] == sku]

        # if not product_info_sept.empty:
        #     name = product_info_sept['name'].values[0].upper()
        #     pack_size = product_info_sept['pack_size_joined'].values[0]
        # else:
        #     name = product_info['name'].values[0].upper()
        #     pack_size = product_info['pack_size_joined'].values[0]
        

        if not product_info_july.empty:
            tariff_code = product_info_july['tariff_code'].values[0]
        else:
            tariff_code = ''
        if not product_info.empty:
            if not product_info_sept.empty:
                cas_number = product_info_sept['cas_number'].values[0]
                pack_size = product_info_sept['pack_size_joined'].values[0]
            else:
                cas_number = product_info['cas_number'].values[0]
                pack_size = product_info['pack_size_joined'].values[0]

            if not prms_info.empty:
                # name = prms_info['Product Name'].values[0]
                price = prms_info['USD List Price'].values[0]
                hazard_statements = prms_info['UN#'].values[0]
                packing_group = prms_info['Packing Group'].values[0]
                alt_storage_temp = prms_info['Storage Temp'].values[0]
                ship_temp = prms_info['Ship Temp'].values[0]
                country_of_origin = prms_info['Country of Origin (most recent lot)'].values[0]
                cas_number = product_info['cas_number'].values[0]
    #             if type(country_of_origin) == str and len(country_of_origin) > 0:
    #                 country_of_origin = pc.country_alpha3_to_country_alpha2(country_of_origin)
    #                 fisher[109][i] = 'Y'
    #             else:
    #                 fisher[109][i] = 'N'
                
            else:
                # name = product_info['name'].values[0].upper()
                price = product_info['price'].values[0]
                packing_group = product_info['packing_group'].values[0]
                ship_temp = product_info['ship_conditions'].values[0]
                hazard_statements = product_info['hazard_statements'].values[0]
                
            if not unspsc_info.empty:
                unspsc = unspsc_info['UNSPSC'].values[0]
            else:
                unspsc = product_info['unspsc'].values[0]
            
            short_desc = product_info['short_description'].values[0]
            img_link = product_info['base_image'].values[0]
            # pack_size = product_info['pack_size_numeric_value'].values[0]
            # unit = product_info['pack_size_unit_of_measure'].values[0]
            quantity = product_info['lk_packaging_facet'].values[0]
            host = product_info['host'].values[0]
            # tariff_code = product_info['tariff_code'].values[0]
            msds_avail = skus['B'+str(i+1)].value
            # cas_number = product_info['cas_number'].values[0]
            DOT_PSN = 'N/A'
            hazard_class = product_info['hazard_class'].values[0]
            biochem_physiol_actions = product_info['biochem_physiol_actions'].values[0]
            storage_and_handling = product_info['storage_and_handling'].values[0]
            group_name = product_info['prms_group_name'].values[0]
            img_link = product_info['base_image'].values[0]
            keywords = product_info['meta_keywords'].values[0]
            storage_temp = product_info['storage_and_handling'].values[0]

            if type(pack_size) == str:
                unit = ''.join([i for i in pack_size if not i.isdigit()])
            else:
                unit = ''
            
            if sku.startswith('07') or sku.startswith('08') or sku.startswith('09') or sku.startswith('11'):
                regulatory[3][i] = 'Diagnostics'
            elif type(cas_number) == str and len(cas_number) > 0:
                regulatory[3][i] = 'Chemicals'
            elif 'preps' in unit or 'Preps' in unit:
                regulatory[3][i] = 'Consumables'
            elif sku.startswith('02') or sku.startswith('03') or sku.startswith('04') or sku.startswith('05'):
                regulatory[3][i] = 'Diagnostics'

            regulatory[4][i] = 99998
            
            if type(storage_temp) == str:
                if 'Room Temperature' in storage_temp or '15-30' in storage_temp or 'ROOM TEMPERATURE' in storage_temp or 'room Temperature' in storage_temp:
                    regulatory[6][i] = 'GWN4'
                elif '-20' in storage_temp:
                    regulatory[6][i] = 'DFD1'
                elif '-70' in storage_temp or '-80' in storage_temp:
                    regulatory[6][i] = 'DFD1'
                elif '2-8' in storage_temp or '0-5' in storage_temp or '2 - 8' in storage_temp or '0' in storage_temp or '4' in storage_temp or 'freezer' in storage_temp:
                    regulatory[6][i] = 'RFC2'
                else:
                    regulatory[6][i] = storage_temp
            else:
                regulatory[6][i] = 'GWN4'
            
            regulatory[8][i] = 'N'
            regulatory[9][i] = 'N'
            regulatory[11][i] = 'N'
            regulatory[12][i] = 'N'
            regulatory[14][i] = 'None'
            regulatory[15][i] = 'NA'
            
            tariff_code = str(tariff_code).replace('.', '')
            if len(tariff_code) >= 4:
                regulatory[29][i] = tariff_code[:4] + '999999'
                
            regulatory[30][i] = 'N'
            regulatory[31][i] = 'N'
            regulatory[32][i] = 'N'
            regulatory[35][i] = 'N'
            regulatory[36][i] = 'N'
            regulatory[37][i] = 'N'
            regulatory[39][i] = 'N'
            regulatory[40][i] = 'N'
            regulatory[43][i] = 'N'
            regulatory[58][i] = 'N'
            regulatory[61][i] = 'N'
            regulatory[62][i] = 'N'
            regulatory[63][i] = 'N'
            regulatory[64][i] = 'NA'
                
            if type(host) == str and len(host) > 0:
                regulatory[44][i] = 'Y'
            else:
                regulatory[44][i] = 'N'
            
            regulatory[47][i] = 'NA'

            if type(cas_number) == str and not cas_number == 'Not applicable':
                regulatory[48][i] = cas_number
                regulatory[49][i] = 100

    for i in range(len(fisher)-1):
        sku = fisher[1][i]
        product_info = magento.loc[magento['sku'] == sku]
        lot_info = lot_master.loc[lot_master['Product number'] == sku]
        prms_info = prms.loc[prms['SKU'] == sku]
        unspsc_info = unspsc_codes.loc[unspsc_codes['Part Number'] == sku]
        product_info_sept = magento_sept.loc[magento_sept['sku'] == sku]


        if not product_info.empty:
            if not product_info_sept.empty:
                name = product_info_sept['name'].values[0].upper()
                pack_size = product_info_sept['pack_size_joined'].values[0]
            else:
                name = product_info['name'].values[0].upper()
                pack_size = product_info['pack_size_joined'].values[0]


            if not prms_info.empty:
                # name = prms_info['Product Name'].values[0]
                price = prms_info['USD List Price'].values[0]
                hazard_statements = prms_info['UN#'].values[0]
                packing_group = prms_info['Packing Group'].values[0]
                alt_storage_temp = prms_info['Storage Temp'].values[0]
                ship_temp = prms_info['Ship Temp'].values[0]
                country_of_origin = prms_info['Country of Origin (most recent lot)'].values[0]
                if type(country_of_origin) == str and len(country_of_origin) > 0:
                    country_of_origin = pc.country_alpha3_to_country_alpha2(country_of_origin)
                    fisher[108][i] = 'Y'
                    fisher[109][i] = 'USD'
                else:
                    fisher[108][i] = 'N'
                
            else:
                # name = product_info['name'].values[0].upper()
                price = product_info['price'].values[0]
                packing_group = product_info['packing_group'].values[0]
                ship_temp = product_info['ship_conditions'].values[0]
                hazard_statements = product_info['hazard_statements'].values[0]
                
            if not unspsc_info.empty:
                unspsc = unspsc_info['UNSPSC'].values[0]
            else:
                unspsc = product_info['unspsc'].values[0]
            
            short_desc = product_info['short_description'].values[0]
            img_link = product_info['base_image'].values[0]
            # pack_size = product_info['pack_size_numeric_value'].values[0]
            # unit = product_info['pack_size_unit_of_measure'].values[0]
            quantity = product_info['lk_packaging_facet'].values[0]
            host = product_info['host'].values[0]
            tariff_code = product_info['tariff_code'].values[0]
            # msds_avail = skus['B'+str(i-1)]
            cas_number = product_info['cas_number'].values[0]
            DOT_PSN = 'NA'
            hazard_class = product_info['hazard_class'].values[0]
            biochem_physiol_actions = product_info['biochem_physiol_actions'].values[0]
            storage_and_handling = product_info['storage_and_handling'].values[0]
            group_name = product_info['prms_group_name'].values[0]
            img_link = product_info['base_image'].values[0]
            keywords = product_info['meta_keywords'].values[0]
            storage_temp = product_info['storage_and_handling'].values[0]
            if not new_magento.loc[new_magento['sku'] == sku].empty:
                weight_in_lb = new_magento.loc[new_magento['sku'] == sku]['weight'].values[0]
                
                fisher[18][i] = weight_in_lb
            
                if weight_in_lb < 0.5:
                    fisher[19][i] = 7
                    fisher[20][i] = 4
                    fisher[21][i] = 5
                    fisher[49][i] = 60
                    fisher[50][i] = 36
                elif weight_in_lb <= 1:
                    fisher[19][i] = 12
                    fisher[20][i] = 7
                    fisher[21][i] = 5
                    fisher[49][i] = 18
                    fisher[50][i] = 36
                else:
                    fisher[19][i] = 12
                    fisher[20][i] = 12
                    fisher[21][i] = 12
                    fisher[49][i] = 12
                    fisher[50][i] = 15
            

            if type(name) == str:
                name = tidyDescription(name)
                temp_name = ''
                for c in name:
                    if c.isalpha() or c.isnumeric() or c == ' ':
                        temp_name = temp_name + c
                name = temp_name
                if len(name) <= 30:
                    fisher[2][i] = name
                else:
                    fisher[2][i] = name[:30]
                
                if len(name) <= 240:
                    fisher[3][i] = name
                else:
                    fisher[3][i] = name[:240]
            # if type(short_desc) == str:
            #     short_desc = tidyDescription(short_desc)
            #     if len(short_desc) <= 240:
            #         fisher[3][i] = short_desc
            #     else:
            #         fisher[3][i] = short_desc[:240]
                    
            fisher[15][i] = 'EA'
            fisher[16][i] = 1
            fisher[17][i] = 'EA'

            if type(pack_size) == str:
                quant = ''.join([i for i in pack_size if i.isdigit()])
                unit = ''.join([i for i in pack_size if not i.isdigit()])
                fisher[22][i] = quant
                if 'preps' in unit or 'Preps' in unit:
                    unit = 'PP'
                elif 'mL' in unit or 'ml' in unit:
                    unit = 'ML'
                elif 'lb' in unit:
                    unit = 'LB'
                elif 'tests' in unit or 'Tests' in unit:
                    unit = 'TS'
                elif 'KU' in unit or 'ku' in unit:
                    unit = 'KU'
                elif 'mg' in unit:
                    unit = 'MG'
                elif 'U' in unit or 'wells' in unit or 'Bags' in unit or 'Bottle' in unit or 'Each' in unit or 'Kit' in unit or 'mCI' in unit:
                    unit = 'UN'
                elif unit == 'g':
                    unit = 'GR'
                elif unit == 'l' or unit == 'L' or 'Liter' in unit:
                    unit = 'LT'
                else:
                    unit = 'UN'
                fisher[23][i] = unit
            
            if type(price) != str:
                fisher[32][i] = price*0.7
                fisher[33][i] = price
            
            fisher[35][i] = '30'
            fisher[36][i] = '31/DEC/2020'
            
            fisher[41][i] = unspsc
            
            if type(hazard_statements) == str and len(hazard_statements) > 0:
                fisher[43][i] = 'Y'
            else:
                fisher[43][i] = 'N'
            
            fisher[44][i] = hazard_statements
            fisher[45][i] = hazard_class
            fisher[47][i] = packing_group
                      
            # if not lot_info.empty:
            #     creation_date = np.datetime64(lot_info['Creation date -'].values[0])
            #     expiration_date = np.datetime64(lot_info['Expiration date -'].values[0])
            #     shelf_life = expiration_date - creation_date
            #     shelf_life = shelf_life.astype('timedelta64[D]')/np.timedelta64(1, 'D') ########################## PLEASE DO NOT DELETE ############################################
            #     if shelf_life > 0:
            #         fisher[54][i] = 'Y'
            #         fisher[55][i] = shelf_life
            #     else:
            #         fisher[54][i] = 0
            #         fisher[55][i] = 0

            fisher[54][i] = 'N'
            fisher[55][i] = 0
            
            if type(keywords) == str:
                keywords = keywords.split(',')
                for j in range(len(keywords)):
                    if len(keywords[j]) > 11:
                        keywords[j] = keywords[j][:11]
                keywords = list(set(keywords))
                j = 0
                while j<5 and j<len(keywords):
                    # if len(keywords[j]) > 11:
                    #     fisher[j+6][i] = keywords[j][:11]
                    # else:
                    fisher[j+6][i] = keywords[j]
                    j = j+1
                    
            fisher[14][i] = '25'
            fisher[93][i] = 'N'
            fisher[42][i] = country_of_origin
            fisher[51][i] = 'Build to Order'
            fisher[52][i] = 'NA'
            fisher[58][i] = 'Y'
            fisher[59][i] = 'N'
            fisher[61][i] = 'Y'
            fisher[62][i] = 1
            fisher[63][i] = 'Y'
            fisher[65][i] = 'Y'
            fisher[78][i] = 'N'
            
            
    new_fisher = opxl.load_workbook('forms/fisher_form.xlsx')
    gen_sheet = new_fisher['General Info']
    regulatory_sheet = new_fisher['Regulatory']
    if gen_sheet.max_row < skus.max_row:
        for j in range(skus.max_row - gen_sheet.max_row+1):
            gen_sheet.insert_rows(gen_sheet.max_row-1)
    if regulatory_sheet.max_row < skus.max_row:
        regulatory_sheet.append([''])
        for j in range(skus.max_row - regulatory_sheet.max_row + 1):
            regulatory_sheet.insert_rows(regulatory_sheet.max_row)
    i = 0
    for row in gen_sheet.iter_rows(min_row=2):
        if i < len(fisher):
            for j in range(len(fisher.columns)):
                if not fisher[j][i] == 'None':
                    row[j].value = fisher[j][i]
        else:
            break
        i = i+1
    i = 0
    for row in regulatory_sheet.iter_rows(min_row=2):
        if i < len(regulatory):
            for j in range(len(regulatory.columns)):
                if not fisher[j][i] == 'None':
                    row[j].value = regulatory[j][i]
        else:
            break
        i = i+1

    # new_fisher.save('outputs/old_product_outputs/old_fisher_output.xlsx')
    new_fisher.save('../../outputs/old_product_outputs/old_fisher_output.xlsx')

def attributeLookup(attribute, product_info, product_info_sept, prms_info, lot_info, unspsc_info, origin_info):
    if not prms_info.empty:
        storage_condition = prms_info['Storage Temp'].values[0]
        ship_temp = prms_info['Ship Temp'].values[0]
        un_num = prms_info['UN#'].values[0]
        
        if storage_condition == 'AM':
            storage_condition = 'Room Temperature'
            storage_code = 'GWN4'
        elif storage_condition == 'RF':
            storage_condition = 'Refrigerated'
            storage_code = 'RFC2'
        elif storage_condition == 'FR':
            storage_condition = 'Frozen'
            storage_code = 'DFD1'
        else:
            storage_code = ''
            
        if ship_temp == 'AM':
            ship_temp = 'Ambient'
        elif ship_temp == 'CP':
            ship_temp = 'Cold Pack'
        elif ship_temp == 'DI':
            ship_temp = 'Dry Ice'
    else:
        storage_condition = ''
        storage_code = ''
        ship_temp = ''
        un_num = ''
        
    if not lot_info.empty:
        creation_date = np.datetime64(lot_info['Creation date -'].values[0])
        expiration_date = np.datetime64(lot_info['Expiration date -'].values[0])
        shelf_life = expiration_date - creation_date
        shelf_life = shelf_life.astype('timedelta64[D]')/np.timedelta64(1, 'D')
    else:
        shelf_life = ''
        
    if not origin_info.empty:
        country_of_origin = origin_info.loc[origin_info['Expiration date -'] == max(origin_info['Expiration date -'].values)]['Country of Origin'].values[0]
    elif not prms_info.empty:
        country_of_origin = prms_info['Country of Origin (most recent lot)'].values[0]
    elif not lot_info.empty:
        country_of_origin = lot_info['Country of Origin'].values[0]
    else:
        country_of_origin = ''
    
    if not product_info.empty:
        attribute_dict = {
            'Packaging': product_info['lk_packaging_facet'].values[0],
            'For Use With (Application)': product_info['application_notes'].values[0],
            'For Use With (Equipment)': product_info['application_notes'].values[0],
            'Color': product_info['color_index'].values[0],
            'Concentration': product_info['concentration'].values[0],
            'pH': product_info['ph'].values[0],
            'Percent Purity': product_info['purity'].values[0],
            'Grade': product_info['grade'].values[0],
            'Quantity': product_info['pack_size_joined'].values[0],
            'Sample Type': product_info['sample_type'].values[0],
            'Target Species': product_info['species_reactivity'].values[0],
            'Molecular Weight (g/mol)': product_info['molecular_weight'].values[0],
            'Molecular Formula': product_info['molecular_formula'].values[0],
            'Melting Point': product_info['melting_point'].values[0],
            'CAS': product_info['cas_number'].values[0],
            'Storage Requirements': storage_condition,
            'Size': product_info['pack_size_joined'].values[0],
            'Packaging Quantity': product_info['pack_size_joined'].values[0],
            'Shelf Life': shelf_life,
            'Format': product_info['format'].values[0],
            'Protein Family': product_info['protein_or_enzyme_type'].values[0],
            'Form': product_info['format'].values[0],
            'Isotype': product_info['isotype'].values[0],
            'Host Species': product_info['host'].values[0],
            'Applications': product_info['key_applications'].values[0],
            'Diameter (Metric)': product_info['dimensions'].values[0],
            'Additional Information': product_info['application_notes'].values[0],
            'Applications': product_info['application_notes'].values[0],
            'Boiling Point': product_info['boiling_point'].values[0],
            'Cell Type': product_info['application_notes'].values[0],
            'Clarity and Color': product_info['color_index'].values[0],
            'Concentration or Composition (by Analyte or Components)': product_info['concentration'].values[0],
            'Concentration or Composition Notes': product_info['concentration'].values[0],
            'Concentration Ratio': product_info['concentration'].values[0],
            'Conductivity': product_info['conductivity'].values[0],
            'Conjugate': product_info['conjugate'].values[0],
            'Culture Environment': product_info['culture_media_type'].values[0],
            'Culture Type': product_info['culture_media_type'].values[0],
            'Density': product_info['density'].values[0],
            'Description': product_info['short_description'].values[0],
            # 'Immunogen': product_info['immunogen'].values[0],
            'Purity': product_info['purity'].values[0],
            'Content And Storage': storage_code,
            'Delivery Type': ship_temp,
            'Shipping Condition': ship_temp,
            'UN Number': un_num,
            'Manufacturing Origin': country_of_origin,
            'Country of Origin': country_of_origin
        }
    
    if not product_info_sept.empty:
        attribute_dict['Sterility'] = product_info_sept['sterilization_of_solutions'].values[0]
        attribute_dict['Synonym'] = product_info_sept['alternate_names'].values[0]
    
    if not product_info.empty and attribute in attribute_dict:
        return attribute_dict[attribute]
    else:
        return ''

def fillFisher_Enrichment(filename, magento, new_magento, lot_master, prms, magento_sept, unspsc_codes, origin):
    authoring_file = pd.ExcelFile(filename)

    authoring = pd.read_excel(authoring_file, 'Core_Content', dtype=object)
    authoring.columns = np.arange(len(authoring.columns))

    attributes = pd.read_excel(authoring_file, 'Category_Attributes', dtype=object)
    attributes.columns = np.arange(len(attributes.columns))
        
    for i in range(3, len(authoring)):
        sku = authoring[3][i]
        product_info = magento.loc[magento['sku'] == sku]
        prms_info = prms.loc[prms['SKU'] == sku]
        if not product_info.empty:
            name = product_info['name'].values[0]
            short_description = product_info['short_description'].values[0]
            description = product_info['description'].values[0]
            image = product_info['base_image'].values[0]
            keywords = product_info['meta_keywords'].values[0]
            
            authoring[8][i] = name
            authoring[9][i] = short_description
            authoring[11][i] = description
            authoring[13][i] = sku
            authoring[19][i] = keywords
        
        if not prms_info.empty:
            un_number = prms_info['UN#'].values[0]
            
            authoring[28][i] = un_number
    
    
    
    for i in range(3, len(attributes)-1):
        print(i)
        sku = attributes[2][i]
        product_info = new_magento.loc[new_magento['sku'] == sku]
        product_info_sept = magento_sept.loc[magento_sept['sku'] == sku]
        prms_info = prms.loc[prms['SKU'] == sku]
        lot_info = lot_master.loc[lot_master['Product number'] == sku]
        unspsc_info = unspsc_codes.loc[unspsc_codes['Part Number'] == sku]
        origin_info = origin.loc[origin['Product number'] == sku]
        
        attribute_name = attributes[7][i]
        
        attributes[11][i] = attributeLookup(attribute_name, product_info, product_info_sept, prms_info, lot_info, unspsc_info, origin_info)

    wb_enrichment = opxl.load_workbook(filename)
    core_content = wb_enrichment['Core_Content']
    i = 3
    for row in core_content.iter_rows(min_row=5):
        for j in range(len(authoring.columns)):
            row[j].value = authoring[j][i]
        i = i+1
    
    attribute_sheet = wb_enrichment['Category_Attributes']
    i = 3
    for row in attribute_sheet.iter_rows(min_row=5):
        if i < len(attributes):
            row[11].value = attributes[11][i]
        else:
            break
        i = i+1

    # wb_enrichment.save('outputs/old_product_outputs/fisher_enrichment_output.xlsx')
    wb_enrichment.save('../../outputs/old_product_outputs/fisher_enrichment_output.xlsx')

# def fillFisher_Enrichment(filename, magento, new_magento, lot_master, prms, magento_sept):
#     authoring_file = pd.ExcelFile(filename)

#     authoring = pd.read_excel(authoring_file, 'Core_Content', dtype=object)
#     authoring.columns = np.arange(len(authoring.columns))

#     attributes = pd.read_excel(authoring_file, 'Category_Attributes', dtype=object)
#     attributes.columns = np.arange(len(attributes.columns))
        
#     for i in range(3, len(authoring)):
#         sku = authoring[3][i]
#         product_info = magento.loc[magento['sku'] == sku]
#         prms_info = prms.loc[prms['SKU'] == sku]
#         if not product_info.empty:
#             name = product_info['name'].values[0]
#             short_description = product_info['short_description'].values[0]
#             description = product_info['description'].values[0]
#             image = product_info['base_image'].values[0]
#             keywords = product_info['meta_keywords'].values[0]
            
#             authoring[8][i] = name
#             authoring[9][i] = short_description
#             authoring[11][i] = description
#             authoring[13][i] = sku
#             authoring[19][i] = keywords
        
#         if not prms_info.empty:
#             un_number = prms_info['UN#'].values[0]
            
#             authoring[28][i] = un_number
    
#     for i in range(3, len(attributes)-1):
#         sku = attributes[2][i]
#         product_info = new_magento.loc[new_magento['sku'] == sku]
#         sept_product_info = magento_sept.loc[magento_sept['sku'] == sku]
#         prms_info = prms.loc[prms['SKU'] == sku]
#         lot_info = lot_master.loc[lot_master['Product number'] == sku]
        
#         attribute_name = attributes[7][i]
#         category_name = attributes[4][i]
        
#         if not prms_info.empty:
#             storage_condition = prms_info['Storage Temp'].values[0]
#             if storage_condition == 'AM':
#                 storage_condition = 'Room Temperature'
#             elif storage_condition == 'RF':
#                 storage_condition = 'Refrigerated'
#             elif storage_condition == 'FR':
#                 storage_condition = 'Frozen'
#         else:
#             storage_condition = ''
            
#         if not lot_info.empty:
#             creation_date = np.datetime64(lot_info['Creation date -'].values[0])
#             expiration_date = np.datetime64(lot_info['Expiration date -'].values[0])
#             shelf_life = expiration_date - creation_date
#             shelf_life = shelf_life.astype('timedelta64[D]')/np.timedelta64(1, 'D')
#         else:
#             shelf_life = ''
        
#         if not product_info.empty:
#             if category_name == 'Nucleic Acid Purification Systems':
#                 if attribute_name == 'For Use With (Application)':
#                     attributes[11][i] = product_info['application_notes'].values[0]
#                 elif attribute_name == 'Kit Size':
#                     attributes[11][i] = '1 kit'
#                 elif attribute_name == 'Particle Size (Metric)':
#                     attributes[11][i] = sept_product_info['particle_size'].values[0]
#                 elif attribute_name == 'Sample Input':
#                     attributes[11][i] = sept_product_info['sample_volume'].values[0]
#                 elif attribute_name == 'Sample Type':
#                     attributes[11][i] = sept_product_info['sample_type'].values[0]
                    
#             elif category_name == 'DNA Extraction and Purification':
#                 if attribute_name == 'For Use With (Application)':
#                     attributes[11][i] = product_info['application_notes'].values[0]
#                 # elif attribute_name == 'For Use With (Equipment)':
#                 #     attributes[11][i] = product_info['application_notes'].values[0]
#                 elif attribute_name == 'Concentration':
#                     attributes[11][i] = product_info['concentration'].values[0]
#                 elif attribute_name == 'Format':
#                     attributes[11][i] = product_info['format'].values[0]
#                 elif attribute_name == 'Promoter':
#                     attributes[11][i] = product_info['biochem_physiol_actions'].values[0]
#                 elif attribute_name == 'Quantity':
#                     attributes[11][i] = product_info['pack_size_joined'].values[0]
                    
#             elif category_name == 'RNA Isolation and Purification Products':
#                 if attribute_name == 'For Use With (Application)':
#                     attributes[11][i] = product_info['key_applications'].values[0]
#                 elif attribute_name == 'Concentration':
#                     attributes[11][i] = product_info['concentration'].values[0]
#                 elif attribute_name == 'Final Product Type': 
#                     attributes[11][i] = 'RNA purification'
#                 elif attribute_name == 'Format':
#                     attributes[11][i] = product_info['format'].values[0]
#                 elif attribute_name == 'Quantity':
#                     attributes[11][i] = product_info['pack_size_joined'].values[0]
#                 elif attribute_name == 'Sample Type':
#                     attributes[11][i] = product_info['sample_type'].values[0]
                    
#             elif category_name == 'Cell Culture Dishes':
#                 # if attribute_name == 'Type':
#                 #     attributes[11][i] = product_info['culture_media_type'].values[0]
#                 if attribute_name == 'Diameter (Metric)':
#                     attributes[11][i] = product_info['dimensions'].values[0]
#                 # elif attribute_name == 'Material':
#                 #     attributes[11][i] = 'N/A'
#                 elif attribute_name == 'Sterility':
#                     attributes[11][i] = product_info['sterility'].values[0]
                    
#             elif category_name == 'Primary Antibodies':
#                 if attribute_name == 'Type':
#                     attributes[11][i] = 'Antibody'
#                 elif attribute_name == 'Applications':
#                     attributes[11][i] = product_info['key_applications'].values[0]
#                 elif attribute_name == 'Conjugate':
#                     attributes[11][i] = product_info['conjugate'].values[0]
#                 elif attribute_name == 'Format':
#                     attributes[11][i] = product_info['formulation'].values[0]
#                 elif attribute_name == 'Host Species':
#                     attributes[11][i] = product_info['host'].values[0]
#                 elif attribute_name == 'Quantity':
#                     attributes[11][i] = product_info['pack_size_joined'].values[0]
#                 elif attribute_name == 'Classification':
#                     attributes[11][i] = product_info['antibody_type'].values[0]
#                 elif attribute_name == 'Regulatory Status':
#                     attributes[11][i] = product_info['usage_statement'].values[0]
#                 elif attribute_name == 'Target Species':
#                     attributes[11][i] = product_info['species_reactivity'].values[0]
#                 elif attribute_name == 'Isotype':
#                     attributes[11][i] = product_info['isotype'].values[0]
                    
#             elif category_name == 'Additional Cell Culture Media':
#                 if attribute_name == 'Product Type':
#                     attributes[11][i] = sept_product_info['culture_media_type'].values[0]
#                 elif attribute_name == 'Protein Family':
#                     attributes[11][i] = product_info['protein_or_enzyme_type'].values[0]
#                 elif attribute_name == 'Concentration':
#                     attributes[11][i] = product_info['concentration'].values[0]
#                 elif attribute_name == 'Shelf Life':
#                     if not type(shelf_life) == str:
#                         attributes[11][i] = str(shelf_life) + ' days'
#                 elif attribute_name == 'Form':
#                     attributes[11][i] = product_info['format'].values[0]
#                 elif attribute_name == 'Research Category':
#                     attributes[11][i] = product_info['key_applications'].values[0]
#                 elif attribute_name == 'Sterility':
#                     attributes[11][i] = sept_product_info['sterilization_of_solutions'].values[0]
                    
#             elif category_name == 'Albumin Testing':
#                 if attribute_name == 'Packaging Quantity':
#                     attributes[11][i] = product_info['lk_packaging_facet'].values[0] 
#                 elif attribute_name == 'Size':
#                     attributes[11][i] = product_info['pack_size_joined'].values[0]
#                 elif attribute_name == 'Quantity':
#                     attributes[11][i] = product_info['pack_size_joined'].values[0]
#                 elif attribute_name == 'Format':
#                     attributes[11][i] = product_info['format'].values[0]
#                 elif attribute_name == 'Storage Requirements':
#                     attributes[11][i] = storage_condition
#                 elif attribute_name == 'Shelf Life':
#                     if not type(shelf_life) == str:
#                         attributes[11][i] = str(shelf_life) + ' days'
#                 elif attribute_name == 'Certifications/Compliance':
#                     attributes[11][i] = sept_product_info['usage_statement'].values[0]
#                 elif attribute_name == 'Detectable Analytes':
#                     attributes[11][i] = sept_product_info['analytes_detected'].values[0]
#                 elif attribute_name == 'Final Product Type':
#                     attributes[11][i] = sept_product_info['protein_or_enzyme_type'].values[0]
                    
#             elif category_name == 'ELISA Kits N-O':
#                 if attribute_name == 'Packaging Quantity':
#                     attributes[11][i] = product_info['pack_size_joined'].values[0]
#                 elif attribute_name == 'Product Type':
#                     attributes[11][i] = 'Elisa kit'
#                 elif attribute_name == 'Target Species':
#                     attributes[11][i] = product_info['species_reactivity'].values[0]
#                 elif attribute_name == 'Sample Type':
#                     attributes[11][i] = product_info['sample_type'].values[0]
                    
#             elif category_name == 'Phosphate Buffered Saline':
#                 if attribute_name == 'Color':
#                     attributes[11][i] = product_info['color_index'].values[0]
#                 elif attribute_name == 'Concentration':
#                     attributes[11][i] = product_info['concentration'].values[0]
#                 elif attribute_name == 'pH':
#                     attributes[11][i] = product_info['ph'].values[0]
#                 elif attribute_name == 'Product Type':
#                     attributes[11][i] = 'buffer'
#                 elif attribute_name == 'Quantity':
#                     attributes[11][i] = product_info['pack_size_joined'].values[0]
                    
#             elif category_name == 'Unclassified Organic Compounds':
#                 if attribute_name == 'Color':
#                     attributes[11][i] = product_info['color_index'].values[0]
#                 elif attribute_name == 'Boiling Point':
#                     attributes[11][i] = product_info['boiling_point'].values[0]
#                 elif attribute_name == 'Melting Point':
#                     attributes[11][i] = product_info['melting_point'].values[0]
#                 elif attribute_name == 'Molecular Weight (g/mol)':
#                     attributes[11][i] = product_info['molecular_weight'].values[0]
#                 elif attribute_name == 'Percent Purity':
#                     attributes[11][i] = product_info['purity'].values[0]
#                 elif attribute_name == 'pH':
#                     attributes[11][i] = product_info['ph'].values[0]
#                 elif attribute_name == 'Quantity':
#                     attributes[11][i] = product_info['pack_size_joined'].values[0]
#                 elif attribute_name == 'Grade':
#                     attributes[11][i] = product_info['grade'].values[0]
                    
#             elif category_name == 'Antibiotics':
#                 if attribute_name == 'Color':
#                     attributes[11][i] = product_info['grade'].values[0]
#                 elif attribute_name == 'Boiling Point':
#                     attributes[11][i] = product_info['boiling_point'].values[0]
#                 elif attribute_name == 'CAS':
#                     attributes[11][i] = 'N/A'
#                 elif attribute_name == 'Melting Point':
#                     attributes[11][i] = product_info['melting_point'].values[0]
#                 elif attribute_name == 'Molecular Formula':
#                     attributes[11][i] = product_info['antibody_category'].values[0]
#                 elif attribute_name == 'Molecular Weight (g/mol)':
#                     attributes[11][i] = product_info['molecular_weight'].values[0]
#                 elif attribute_name == 'Packaging':
#                     attributes[11][i] = product_info['lk_packaging_facet'].values[0]
#                 elif attribute_name == 'Percent Purity':
#                     attributes[11][i] = product_info['purity'].values[0]
#                 elif attribute_name == 'pH':
#                     attributes[11][i] = product_info['ph'].values[0]
#                 elif attribute_name == 'Quantity':
#                     attributes[11][i] = product_info['pack_size_joined'].values[0]
#                 elif attribute_name == 'Grade':
#                     attributes[11][i] = product_info['grade'].values[0]
#                 elif attribute_name == 'Synonym':
#                     attributes[11][i] = sept_product_info['alternate_names'].values[0]
                    
#             elif category_name == 'Liquid Scintillation Counting Cocktails and Reagents':
#                 if attribute_name == 'Boiling Point':
#                     attributes[11][i] = product_info['boiling_point'].values[0]
#                 elif attribute_name == 'CAS':
#                     attributes[11][i] = product_info['cas_number'].values[0]
#                 elif attribute_name == 'Melting Point':
#                     attributes[11][i] = product_info['melting_point'].values[0]
#                 elif attribute_name == 'Molecular Formula':
#                     attributes[11][i] = product_info['antibody_category'].values[0]
#                 elif attribute_name == 'Molecular Weight (g/mol)':
#                     attributes[11][i] = product_info['molecular_weight'].values[0]
#                 elif attribute_name == 'Packaging':
#                     attributes[11][i] = product_info['lk_packaging_facet'].values[0]
#                 elif attribute_name == 'Percent Purity':
#                     attributes[11][i] = product_info['purity'].values[0]
#                 elif attribute_name == 'pH':
#                     attributes[11][i] = product_info['ph'].values[0]
#                 elif attribute_name == 'Quantity':
#                     attributes[11][i] = product_info['pack_size_joined'].values[0]
#                 elif attribute_name == 'Grade':
#                     attributes[11][i] = product_info['grade'].values[0]
#                 elif attribute_name == 'Color':
#                     attributes[11][i] = sept_product_info['color_index'].values[0]

#     wb_enrichment = opxl.load_workbook(filename)
#     core_content = wb_enrichment['Core_Content']
#     i = 3
#     for row in core_content.iter_rows(min_row=5):
#         for j in range(len(authoring.columns)):
#             row[j].value = authoring[j][i]
#         i = i+1
    
#     attribute_sheet = wb_enrichment['Category_Attributes']
#     i = 3
#     for row in attribute_sheet.iter_rows(min_row=5):
#         if i < len(attributes):
#             row[11].value = attributes[11][i]
#         else:
#             break
#         i = i+1

#     wb_enrichment.save('../../outputs/enrichment_outputs/fisher_enrichment_output.xlsx')

def fillVWR_Enrichment(filename, magento):
    enrichment = pd.read_excel('forms/vwr_enrichment_form.xlsx', dtype = object)
    enrichment.columns = np.arange(len(enrichment.columns))
    wb = opxl.load_workbook(filename)
    skus = wb.active
    for i in range(2, skus.max_row+1):
        enrichment.loc[i+9, 4] = str(skus['A'+str(i)].value)
        
    for i in range(11, skus.max_row+10):
        sku = enrichment[4][i]
        product_info = magento.loc[magento['sku'] == sku]
        
        if not product_info.empty:
            short_desc = product_info['short_description'].values[0]
            keywords = product_info['meta_keywords'].values[0]
            name = product_info['name'].values[0].upper()
            description = product_info['description'].values[0]
            
            enrichment[16][i] = 'MP Biomedical'
            enrichment[17][i] = 'MP Biomedical'
            enrichment[18][i] = tidyDescription(str(name))
            enrichment[19][i] = tidyDescription(str(short_desc))
            enrichment[20][i] = tidyDescription(str(description))
            enrichment[21][i] = keywords
            
    new_enrichment = opxl.load_workbook('forms/vwr_enrichment_form.xlsx')
    regulatory_sheet = new_enrichment.active
    if regulatory_sheet.max_row < skus.max_row:
        regulatory_sheet.append([''])
        for j in range(skus.max_row - regulatory_sheet.max_row + 11):
            regulatory_sheet.insert_rows(regulatory_sheet.max_row)
    i = 11
    for row in regulatory_sheet.iter_rows(min_row=13):
        if i < (skus.max_row+11):
            for j in range(len(enrichment.columns)):
                row[j].value = enrichment[j][i]
        else:
            break
        i = i+1

    new_enrichment.save('../../outputs/enrichment_outputs/vwr_enrichment_output.xlsx')

def fillVWR_New(product_manager, prms2, e_marketing):
    vwr = pd.read_excel('forms/vwr_form.xlsx', dtype=object)
    vwr.columns = np.arange(len(vwr.columns))
    for i in range(1, len(product_manager)):
        vwr.loc[i+3, 4] = str(product_manager['sku'][i])

        
    for i in range(4, len(vwr)):
        sku = str(vwr[4][i])
        product_info = product_manager.loc[product_manager['sku'] == sku]
        marketing_info = e_marketing.loc[e_marketing['sku'] == sku]
        prms_info = prms2.loc[prms2['sku'] == sku]
        
        if not product_info.empty:
            name = product_info['Name'].values[0].upper().replace(',', '')
            full_desc = product_info['Description'].values[0]
            short_desc = product_info['short_description'].values[0]
            price = product_info['Price information '].values[0]
            shipping_condition = product_info['shipping condition '].values[0]
            storage_condition = product_info['storage condition '].values[0]
            country_of_origin = product_info['country_of_manufacture'].values[0]
            shelf_life = product_info['shelf life '].values[0]
            pkg_size = product_info['Package size '].values[0]
            pkg_unit = product_info['Package size unit'].values[0]
            tariff_code = product_info['tariff_code'].values[0]
            host = product_info['host'].values[0]
            keywords = product_info['keywords'].values[0]
            unspsc = product_info['unspsc'].values[0]
            hazard_statements = product_info['hazard_statements'].values[0]
            hazard_class = product_info['Hazard class '].values[0]
            psn = product_info['proper shipping name '].values[0]
            un_num = product_info['UN/NA #'].values[0]
            packing_group = product_info['Package group'].values[0]
            safety_symbol = product_info['safety_symbol'].values[0]
            cas_number = product_info['cas_number'].values[0]
            tariff_code = product_info['tariff_code'].values[0]
            ph = product_info['ph'].values[0]
            protein_or_enzyme_type = product_info['protein_or_enzyme_type'].values[0]
            #msds_avail = skus['B'+str(i+1)].value
            biochem_physiol_actions = product_info['biochem_physiol_actions'].values[0]
            
            if type(name) == str:
                name = tidyDescription(name)
                if len(name) <= 40:
                    vwr[5][i] = name
                else:
                    vwr[5][i] = name[:40]
            if type(short_desc) == str:
                short_desc = tidyDescription(short_desc)
                if len(short_desc) <= 300:
                    vwr[6][i] = short_desc
                else:
                    vwr[6][i] = short_desc[:300]
                    
            if shipping_condition == 'Cold Pack' or shipping_condition == 'Cold pack':
                vwr[107][i] = 'B'
                vwr[108][i] = 'Blue'
            elif shipping_condition == 'Dry Ice':
                vwr[107][i] = 'A'
                vwr[108][i] = 'Dry'
            else:
                vwr[107][i] = 'C'
                
            if type(storage_condition) == str:
                if 'Room Temperature' in storage_condition or '15-30' in storage_condition or 'ROOM TEMPERATURE' in storage_condition:
                    vwr[111][i] = 'C'
                elif '-20' in storage_condition:
                    vwr[111][i] = 'A'
                elif '-70' in storage_condition or '-80' in storage_condition:
                    vwr[111][i] = 'D'
                elif '2-8' in storage_condition or '0-5' in storage_condition or '2 - 8' in storage_condition or '0' in storage_condition or '4' in storage_condition:
                    vwr[111][i] = 'B'
                else:
                    vwr[111][i] = storage_condition
                
            if type(host) == str and host == 'N/A':
                vwr[61][i] = 'Synthetic'
            else:
                vwr[61][i] = 'Animal: ' + str(host)
            
            #if DOT_PSN != 'N/A': 
            vwr[68][i] = name
            vwr[69][i] = biochem_physiol_actions
            vwr[70][i] = packing_group
            vwr[71][i] = 'No'
            vwr[72][i] = 'No'
            vwr[73][i] = 'No'
            vwr[77][i] = packing_group
            
            tariff_code = str(tariff_code)
            if len(tariff_code) >= 4:
                vwr[54][i] = tariff_code[:4] + '.99.9999'
            else:
                vwr[54][i] = tariff_code
                
            vwr[7][i] = 'EA'
            vwr[8][i] = 1
            vwr[10][i] = 'USD'
            vwr[12][i] = '1/1/2021'
            vwr[15][i] = 'No'
            vwr[18][i] = pkg_unit
            vwr[19][i] = pkg_size
            vwr[20][i] = 'EA'
            vwr[21][i] = price
            vwr[22][i] = price
            vwr[43][i] = 'Y'
            vwr[58][i] = country_of_origin
            vwr[59][i] = 'No'
            vwr[60][i] = 'No'
            vwr[62][i] = host
            #vwr[65][i] = msds_avail
            #vwr[66][i] = msds_avail
            vwr[67][i] = psn
            vwr[82][i] = 'Box'
            vwr[83][i] = 'No'
            vwr[84][i] = 'Plastic'
            vwr[85][i] = cas_number
            vwr[86][i] = 'No'
            vwr[87][i] = 'No'
            vwr[88][i] = 'No'
            vwr[89][i] = 'Y'
            vwr[90][i] = 'No'
            vwr[91][i] = 'No'
            vwr[92][i] = 'N/A'
            vwr[93][i] = 'N/A'
            vwr[105][i] = 'No'
            vwr[106][i] = 'Y'
            vwr[110][i] = 'No'
            vwr[164][i] = 'N'
            vwr[109][i] = shelf_life
        
        if not marketing_info.empty:
            
            weight = marketing_info['weight'].values[0]
            meta_keywords = marketing_info['meta_keywords'].values[0]
            meta_description = marketing_info['meta_description'].values[0]
            
            vwr[39][i] = weight
            
            if weight < 0.5:
                vwr[40][i] = 7
                vwr[41][i] = 4
                vwr[42][i] = 5
            elif weight <= 1:
                vwr[40][i] = 12
                vwr[41][i] = 7
                vwr[42][i] = 5
            else:
                vwr[40][i] = 12
                vwr[41][i] = 12
                vwr[42][i] = 12
            
        if not prms_info.empty:
            pkg_size_joined = prms_info['pack_size_joined'].values[0]
            ship_conditions = prms_info['ship_conditions'].values[0]
            un_number = prms_info['un_number'].values[0]
            pack_group = prms_info['packing_group'].values[0]
            ship_hazard_code = prms_info['ship_hazard_code'].values[0]

    new_vwr = opxl.load_workbook('forms/vwr_form.xlsx')
    vwr_sheet = new_vwr.active
    if vwr_sheet.max_row < len(product_manager)+4:
        for j in range(len(product_manager) - vwr_sheet.max_row + 4):
            vwr_sheet.insert_rows(vwr_sheet.max_row-1)
    # for j in range(product_manager.max_row):
    #     vwr_sheet.insert_rows(6+j)
    i = 1
    for row in vwr_sheet.iter_rows(min_row=6):
        if i < len(product_manager):
            for j in range(2, len(vwr.columns)):
                row[j].value = vwr[j][i+3]
        else:
            break
        i = i+1

    new_vwr.save('../../outputs/new_product_outputs/new_vwr_output.xlsx')

def fillThomas_New(product_manager, prms2, e_marketing):
    thomas = pd.read_excel('forms/thomas_form.xlsx', dtype = object)
    thomas.columns = np.arange(len(thomas.columns))
    for i in range(21, len(product_manager)+20):
        thomas[1][i] = str(product_manager['sku'][i-20])
        
    for i in range(21, len(product_manager)+20):
        sku = thomas[1][i]
        product_info = product_manager.loc[product_manager['sku'] == sku]
        marketing_info = e_marketing.loc[e_marketing['sku'] == sku]
        prms_info = prms2.loc[prms2['sku'] == sku]
        if not product_info.empty:
            name = product_info['Name'].values[0].upper().replace(',', '')
            full_desc = product_info['Description'].values[0]
            short_desc = product_info['short_description'].values[0]
            price = product_info['Price information '].values[0]
            shipping_condition = product_info['shipping condition '].values[0]
            storage_condition = product_info['storage condition '].values[0]
            country_of_origin = product_info['country_of_manufacture'].values[0]
            shelf_life = product_info['shelf life '].values[0]
            pkg_size = product_info['Package size '].values[0]
            pkg_unit = product_info['Package size unit'].values[0]
            tariff_code = product_info['tariff_code'].values[0]
            host = product_info['host'].values[0]
            keywords = product_info['keywords'].values[0]
            unspsc = product_info['unspsc'].values[0]
            hazard_statements = product_info['hazard_statements'].values[0]
            hazard_class = product_info['Hazard class '].values[0]
            psn = product_info['proper shipping name '].values[0]
            un_num = product_info['UN/NA #'].values[0]
            packing_group = product_info['Package group'].values[0]
            safety_symbol = product_info['safety_symbol'].values[0]
            cas_number = product_info['cas_number'].values[0]
            tariff_code = product_info['tariff_code'].values[0]
            ph = product_info['ph'].values[0]
            protein_or_enzyme_type = product_info['protein_or_enzyme_type'].values[0]
            # msds_avail = skus['B'+str(i-1)].value
            purity = product_info['purity'].values[0]
            grade = product_info['grade'].values[0]
            molecular_weight = product_info['molecular_weight'].values[0]
            key_applications = product_info['key_applications'].values[0]
            concentration = product_info['concentration'].values[0]
            
            thomas[3][i] = 'EA'
            thomas[4][i] = 1
            
            if type(price) != str:
                if str(sku).startswith('11'):
                    thomas[5][i] = 0.89*price
                else:
                    thomas[5][i] = 0.8*price
                thomas[6][i] = price
                
            if shipping_condition == 'Dry Ice':
                thomas[13][i] = 'Ice/Dry Ice'
                thomas[14][i] = '$10.00'
                thomas[15][i] = 'ITEM'
                thomas[31][i] = 'DRY ICE'
            elif shipping_condition == 'Cold Pack' or shipping_condition == 'Cold pack':
                thomas[31][i] = 'ICE'
            elif shipping_condition == 'Ambient':
                thomas[31][i] = 'RT'
                
            if type(name) == str:
                name = tidyDescription(name)
                if (not (str(pkg_size) + str(pkg_unit)) in name) and not ((str(pkg_size) + ' ' + str(pkg_unit)) in name):
                    if len(name) > 40:
                        name = name[:39-len(str(pkg_size) + str(pkg_unit))] + ' ' + str(pkg_size) + str(pkg_unit)
                    else:
                        name = name + ' ' + str(pkg_size) + str(pkg_unit)
                if len(name) > 40:
                    name = name[:40]
                thomas[2][i] = name
                
            if type(cas_number) == str and len(cas_number) > 0:
                thomas[17][i] = 'Chemicals'
                thomas[52][i] = grade
                thomas[53][i] = str(pkg_size) + str(pkg_unit)
                thomas[54][i] = 'Bottle'
                thomas[55][i] = cas_number
                thomas[59][i] = ph
                
            thomas[21][i] = 'Y'
            thomas[22][i] = 'NONE'
            thomas[24][i] = country_of_origin
            
            tariff_code = str(tariff_code)
            if len(tariff_code) >= 4:
                thomas[23][i] = tariff_code[:4] + '.99.9999'
                
            if shelf_life > 0:
                thomas[34][i] = str(shelf_life) + ' days'
                
            if type(storage_condition) == str:
                if 'Room Temperature' in storage_condition or '15-30' in storage_condition or 'ROOM TEMPERATURE' in storage_condition:
                    thomas[32][i] = 'RT'
                    thomas[33][i] = 'N/A'
                elif '-20' in storage_condition:
                    thomas[32][i] = '-20°C'
                    thomas[33][i] = 'Freezer'
                elif '-70' in storage_condition or '-80' in storage_condition:
                    thomas[32][i] = '-80°C'
                    thomas[33][i] = 'Freezer'
                elif '2-8' in storage_condition or '0-5' in storage_condition or '2 - 8' in storage_condition or '0' in storage_condition or '4' in storage_condition:
                    thomas[32][i] = '4°C'
                    thomas[33][i] = 'Refrigerator'
                else:
                    thomas[32][i] = storage_condition
                    
            if type(name) == str:
                name = tidyDescription(name)
                thomas[64][i] = name + ' ' + str(pkg_size) + str(pkg_unit)
                thomas[66][i] = name + ' ' + str(pkg_size) + str(pkg_unit)
                
            if type(keywords) == str:
                thomas[68][i] = keywords.replace(',', ' ')
                
            specifications = ''
            if type(purity) == str and len(purity) > 0:
                specifications = specifications + 'Purity: ' + purity + '\n'
            if type(molecular_weight) == str and len(molecular_weight) > 0:
                specifications = specifications + 'Molecular Weight: ' + molecular_weight + '\n'
            if type(key_applications) == str and len(key_applications) > 0:
                specifications = specifications + 'Key Applications: ' + key_applications + '\n'
            if type(concentration) == str and len(concentration) > 0:
                specifications = specifications + 'Concentration: ' + concentration
            thomas[67][i] = specifications
            
            thomas[44][i] = 'No'
            thomas[65][i] = 'N/A'
                
    #         thomas[69][i] = img_link
            
        if not marketing_info.empty:
            weight = marketing_info['weight'].values[0]
            meta_keywords = marketing_info['meta_keywords'].values[0]
            meta_description = marketing_info['meta_description'].values[0]
            
            if weight < 0.5:
                thomas[25][i] = 7
                thomas[26][i] = 4
                thomas[27][i] = 5
                thomas[28][i] = 0.081
            elif weight <= 1:
                thomas[25][i] = 12
                thomas[26][i] = 7
                thomas[27][i] = 5
                thomas[28][i] = 0.243
            else:
                thomas[25][i] = 12
                thomas[26][i] = 12
                thomas[27][i] = 12
                thomas[28][i] = 1
                
            thomas[29][i] = weight
        
        if not prms_info.empty:
            pkg_size_joined = prms_info['pack_size_joined'].values[0]
            ship_conditions = prms_info['ship_conditions'].values[0]
            un_number = prms_info['un_number'].values[0]
            pack_group = prms_info['packing_group'].values[0]
            ship_hazard_code = prms_info['ship_hazard_code'].values[0]
            
    new_thomas = opxl.load_workbook('forms/thomas_form.xlsx')
    thomas_sheet = new_thomas.active
    if thomas_sheet.max_row < len(product_info):
        for j in range(len(product_manager) - thomas_sheet.max_row+22):
            thomas_sheet.insert_rows(thomas_sheet.max_row-1)
    i = 1
    for row in thomas_sheet.iter_rows(min_row=23):
        if i < len(product_manager)+22:
            for j in range(len(thomas.columns)):
                row[j].value = thomas[j][i+2]
        else:
            break
        i = i+1

    new_thomas.save('../../outputs/new_product_outputs/new_thomas_output.xlsx')

def fillFisher_New(product_manager, prms2, e_marketing):
    fisher_file = pd.ExcelFile('forms/fisher_form.xlsx')
    regulatory = pd.read_excel(fisher_file, 'Regulatory', dtype=object)
    regulatory.columns = np.arange(len(regulatory.columns))
    fisher = pd.read_excel(fisher_file, 'General Info', dtype=object)
    fisher.columns = np.arange(len(fisher.columns))

    for i in range(1, len(product_manager)):
        regulatory.loc[i, 1] = product_manager['sku'][i]
        fisher.loc[i, 1] = product_manager['sku'][i]

    for i in range(1, len(regulatory[1])):
        sku = regulatory[1][i]
        product_info = product_manager.loc[product_manager['sku'] == sku]
        marketing_info = e_marketing.loc[e_marketing['sku'] == sku]
        prms_info = prms2.loc[prms2['sku'] == sku]

        if not product_info.empty:
            name = product_info['Name'].values[0].upper().replace(',', '')
            full_desc = product_info['Description'].values[0]
            short_desc = product_info['short_description'].values[0]
            price = product_info['Price information '].values[0]
            shipping_condition = product_info['shipping condition '].values[0]
            storage_condition = product_info['storage condition '].values[0]
            country_of_origin = product_info['country_of_manufacture'].values[0]
            shelf_life = product_info['shelf life '].values[0]
            pkg_size = product_info['Package size '].values[0]
            pkg_unit = product_info['Package size unit'].values[0]
            tariff_code = product_info['tariff_code'].values[0]
            host = product_info['host'].values[0]
            keywords = product_info['keywords'].values[0]
            unspsc = product_info['unspsc'].values[0]
            hazard_statements = product_info['hazard_statements'].values[0]
            hazard_class = product_info['Hazard class '].values[0]
            psn = product_info['proper shipping name '].values[0]
            un_num = product_info['UN/NA #'].values[0]
            packing_group = product_info['Package group'].values[0]
            safety_symbol = product_info['safety_symbol'].values[0]
            cas_number = product_info['cas_number'].values[0]
            tariff_code = product_info['tariff_code'].values[0]
            ph = product_info['ph'].values[0]
            protein_or_enzyme_type = product_info['protein_or_enzyme_type'].values[0]
            biochem_physiol_actions = product_info['biochem_physiol_actions'].values[0]
            
            
            if type(storage_condition) == str:
                if 'Room Temperature' in storage_condition or '15-30' in storage_condition or 'ROOM TEMPERATURE' in storage_condition or 'room Temperature' in storage_condition:
                    regulatory[6][i] = 'GWN4'
                elif '-20' in storage_condition:
                    regulatory[6][i] = 'DFD1'
                elif '-70' in storage_condition or '-80' in storage_condition:
                    regulatory[6][i] = 'DFD1'
                elif '2-8' in storage_condition or '0-5' in storage_condition or '2 - 8' in storage_condition or '0' in storage_condition or '4' in storage_condition or 'freezer' in storage_condition:
                    regulatory[6][i] = 'RFN1'
                else:
                    regulatory[6][i] = storage_condition
            else:
                regulatory[6][i] = 'GWN4'
            
            regulatory[8][i] = 'N'
            regulatory[9][i] = 'N'
            regulatory[11][i] = 'N'
            regulatory[12][i] = 'N'
            
            tariff_code = str(tariff_code)
            if len(tariff_code) >= 4:
                regulatory[29][i] = tariff_code[:4] + '.99.9999'
                
            regulatory[30][i] = 'N'
            regulatory[31][i] = 'N'
            regulatory[32][i] = 'N'
            regulatory[36][i] = 'N'
            regulatory[37][i] = 'N'
            regulatory[39][i] = 'N'
            regulatory[40][i] = 'N'
            regulatory[43][i] = 'N'
                
            if type(host) == str and len(host) > 0:
                regulatory[44][i] = 'Y'
            else:
                regulatory[44][i] = 'N'

    for i in range(1, len(fisher[1])):
        sku = fisher[1][i]
        product_info = product_manager.loc[product_manager['sku'] == sku]
        marketing_info = e_marketing.loc[e_marketing['sku'] == sku]
        prms_info = prms2.loc[prms2['sku'] == sku]
            
        if not product_info.empty:
            name = product_info['Name'].values[0].upper().replace(',', '')
            full_desc = product_info['Description'].values[0]
            short_desc = product_info['short_description'].values[0]
            price = product_info['Price information '].values[0]
            shipping_condition = product_info['shipping condition '].values[0]
            storage_condition = product_info['storage condition '].values[0]
            country_of_origin = product_info['country_of_manufacture'].values[0]
            shelf_life = product_info['shelf life '].values[0]
            pkg_size = product_info['Package size '].values[0]
            pkg_unit = product_info['Package size unit'].values[0]
            tariff_code = product_info['tariff_code'].values[0]
            host = product_info['host'].values[0]
            keywords = product_info['keywords'].values[0]
            unspsc = product_info['unspsc'].values[0]
            hazard_statements = product_info['hazard_statements'].values[0]
            hazard_class = product_info['Hazard class '].values[0]
            psn = product_info['proper shipping name '].values[0]
            un_num = product_info['UN/NA #'].values[0]
            packing_group = product_info['Package group'].values[0]
            safety_symbol = product_info['safety_symbol'].values[0]
            cas_number = product_info['cas_number'].values[0]
            tariff_code = product_info['tariff_code'].values[0]
            ph = product_info['ph'].values[0]
            protein_or_enzyme_type = product_info['protein_or_enzyme_type'].values[0]
            biochem_physiol_actions = product_info['biochem_physiol_actions'].values[0]
            
            if type(name) == str:
                name = tidyDescription(name)
                if len(name) <= 30:
                    fisher[2][i] = name
                else:
                    fisher[2][i] = name[:30]
            if type(short_desc) == str:
                short_desc = tidyDescription(short_desc)
                if len(short_desc) <= 240:
                    fisher[3][i] = short_desc
                else:
                    fisher[3][i] = short_desc[:240]
                    
            fisher[15][i] = 'EA'
            fisher[16][i] = 1
            fisher[17][i] = 'EA'
            
            if type(hazard_statements) == str and len(hazard_statements) > 0:
                fisher[43][i] = 'Y'
            else:
                fisher[43][i] = 'N'
            
            if type(price) != str:
                fisher[32][i] = price*0.7
                fisher[33][i] = price
                
            fisher[35][i] = '30'
            fisher[36][i] = '31/DEC/2020'
            fisher[41][i] = unspsc
            
            # if msds_avail == 'Y':
            #     fisher[27][i] = '99998'
            # else:
            #     fisher[27][i] = '00000'
                
            fisher[44][i] = hazard_statements
            fisher[45][i] = hazard_class
            fisher[47][i] = packing_group
            
            if type(storage_condition) == str:
                if 'Room Temperature' in storage_condition or '15-30' in storage_condition or 'ROOM TEMPERATURE' in storage_condition:
                    fisher[36][i] = 'GWN4'
                elif '-20' in storage_condition:
                    fisher[36][i] = 'DFD1'
                elif '-70' in storage_condition or '-80' in storage_condition:
                    fisher[36][i] = 'DFD1'
                elif '2-8' in storage_condition or '0-5' in storage_condition or '2 - 8' in storage_condition or '0' in storage_condition or '4' in storage_condition:
                    fisher[36][i] = 'RFN1'
                else:
                    fisher[36][i] = storage_condition
                    
            fisher[43][i] = 'N'
            fisher[50][i] = 'No'
            
            if shelf_life > 0:
                fisher[54][i] = 'Y'
                fisher[55][i] = shelf_life
            else:
                fisher[54][i] = 'N'
                fisher[55][i] = 0
                
            fisher[14][i] = '25 Days'
            fisher[93][i] = 'N'
            fisher[42][i] = country_of_origin
            fisher[51][i] = 'Build to Order'
            fisher[52][i] = 'N/A'
            fisher[58][i] = 'Y'
            fisher[59][i] = 'Y'
            fisher[62][i] = 'N/A'
            fisher[63][i] = 'Y'
            fisher[65][i] = 'Y'
            fisher[78][i] = 'N'
            
        if not marketing_info.empty:
            weight = marketing_info['weight'].values[0]
            meta_keywords = marketing_info['meta_keywords'].values[0]
            meta_description = marketing_info['meta_description'].values[0]
            
            fisher[18][i] = weight
            
            if weight < 0.5:
                fisher[19][i] = 7
                fisher[20][i] = 4
                fisher[21][i] = 5
                fisher[49][i] = 60
                fisher[50][i] = 36
            elif weight <= 1:
                fisher[19][i] = 12
                fisher[20][i] = 7
                fisher[21][i] = 5
                fisher[49][i] = 18
                fisher[50][i] = 36
            else:
                fisher[19][i] = 12
                fisher[20][i] = 12
                fisher[21][i] = 12
                fisher[49][i] = 12
                fisher[50][i] = 15
                
            if type(keywords) == str:
                keywords = keywords.split(',')
                j = 0
                while j<5 and j<len(keywords):
                    fisher[j+6][i] = keywords[j]
                    j = j+1
            
    new_fisher = opxl.load_workbook('forms/fisher_form.xlsx')
    fisher_sheet = new_fisher['General Info']
    regulatory_sheet = new_fisher['Regulatory']
    if fisher_sheet.max_row < len(regulatory[1]):
        for j in range(len(regulatory[1]) - regulatory_sheet.max_row + 1):
            # regulatory_sheet.insert_rows(regulatory_sheet.max_row)
            regulatory_sheet.append([''])
    i = 1
    for row in regulatory_sheet.iter_rows(min_row=3):
        if i < len(regulatory[1]):
            for j in range(1, len(regulatory.columns)):
                row[j].value = regulatory[j][i]
        else:
            break
        i = i+1
    if fisher_sheet.max_row < len(fisher[1]):
        for j in range(len(fisher[1]) - fisher_sheet.max_row + 1):
            # fisher_sheet.insert_rows(fisher_sheet.max_row)
            fisher_sheet.append([''])
    i = 1
    for row in fisher_sheet.iter_rows(min_row=3):
        if i < len(fisher[1]):
            for j in range(1, len(fisher.columns)):
                row[j].value = fisher[j][i]
        else:
            break
        i = i+1

    new_fisher.save('../../outputs/new_product_outputs/new_fisher_output.xlsx')